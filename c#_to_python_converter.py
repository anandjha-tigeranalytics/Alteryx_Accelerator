import os
import re
import time
import openpyxl
import pyodbc
import threading
import traceback
import xml.etree.ElementTree as ET
import xml.dom.minidom as minidom
import pandas as pd
from openpyxl import load_workbook,Workbook
from xml.dom import minidom
from dataclasses import dataclass, field
from typing import List, Optional
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

@dataclass
class TaskInfo:
    PackageName: Optional[str] = None
    PackagePath: Optional[str] = None
    EventHandlerName: Optional[str] = None
    EventHandlerType: Optional[str] = None
    EventType: Optional[str] = None
    TaskName: Optional[str] = None
    TaskType: Optional[str] = None
    TaskSqlQuery: Optional[str] = None
    ContainerName: Optional[str] = None
    ContainerType: Optional[str] = None
    ContainerExpression: Optional[str] = None
    ContainerEnum: Optional[str] = None
    Variables: Optional[str] = None
    Parameters: Optional[str] = None
    Expressions: Optional[str] = None
    ExecuteProcessDetails: Optional[str] = None
    FileSystemSourcePath: Optional[str] = None
    FileSystemDestinationPath: Optional[str] = None
    SourceComponent: Optional[str] = None
    TargetComponent: Optional[str] = None
    SourceType: Optional[str] = None
    TargetType: Optional[str] = None
    TargetTable: Optional[str] = None
    SendMailTask: Optional[str] = None
    ScriptTask: Optional[str] = None
    FTPTask: Optional[str] = None
    ExecutePackage: Optional[str] = None
    ResultSetDetails: Optional[str] = None
    SeqTaskName: Optional[str] = None
    ForeachTaskName: Optional[str] = None
    ForloopTaskName: Optional[str] = None
    ConnectionName: Optional[str] = None
    SourceConnectionName: Optional[str] = None
    TargetConnectionName: Optional[str] = None
    TaskComponentDetails: Optional[str] = None

@dataclass
class ConnectionInfo:
    ConnectionName: Optional[str] = None
    ConnectionType: Optional[str] = None
    ConnectionString: Optional[str] = None
    ConnectionExpressions: Optional[str] = None
    ConnectionID: Optional[str] = None
    IsProjectConnection: Optional[str] = None

@dataclass
class ContainerInfo:
    ContainerName: Optional[str] = None
    ContainerType: Optional[str] = None
    ContainerExpression: Optional[str] = None

@dataclass
class VariableInfo:
    Name: Optional[str] = None
    Value: Optional[str] = None
    DataType: Optional[str] = None
    Namespace: Optional[str] = None
    IsParameter: int = 0

@dataclass
class TaskParameterInfo:
    ParameterName: Optional[str] = None
    ParameterType: Optional[str] = None
    DataType: Optional[str] = None
    Value: Optional[str] = None
    DtsVariableName: Optional[str] = None

@dataclass
class DataFlowTaskInfo:
    ColumnName: Optional[str] = None
    ColumnType: Optional[str] = None
    DataType: Optional[str] = None
    TargetColumn: Optional[str] = None
    componentName: Optional[str] = None
    DataConversion: Optional[str] = None
    PackageName: Optional[str] = None
    PackagePath: Optional[str] = None
    TaskName: Optional[str] = None
    isEventHandler: Optional[str] = None
    componentPropertyDetails: Optional[str] = None
    ColumnPropertyDetails: Optional[str] = None

@dataclass
class PrecedenceConstraintInfo:
    PrecedenceConstraintFrom: Optional[str] = None
    PrecedenceConstraintTo: Optional[str] = None
    PrecedenceConstraintValue: Optional[str] = None
    PrecedenceConstraintLogicalAnd: Optional[str] = None
    PrecedenceConstraintEvalOP: Optional[str] = None
    PrecedenceConstraintExpression: Optional[str] = None
    ContainerName: Optional[str] = None
    PackageName: Optional[str] = None
    PackagePath: Optional[str] = None

@dataclass
class ProjectParameterInfo:
    ParameterName: Optional[str] = None
    DataType: Optional[str] = None
    Value: Optional[str] = None

@dataclass
class PackageAnalysisResult:
    PackageName: Optional[str] = None
    CreatedDate: Optional[datetime] = None
    CreatedBy: Optional[str] = None
    Tasks: List[TaskInfo] = field(default_factory=list)
    Seqtasks: List[TaskInfo] = field(default_factory=list)
    Foreachtasks: List[TaskInfo] = field(default_factory=list)
    Forlooptasks: List[TaskInfo] = field(default_factory=list)
    Connections: List[ConnectionInfo] = field(default_factory=list)
    ExecutionTime: Optional[timedelta] = None
    PackagePath: Optional[str] = None
    Containers: List[ContainerInfo] = field(default_factory=list)
    DTSXXML: Optional[str] = None
    SequenceContainerTaskCount: List[TaskInfo] = field(default_factory=list)
    ForeachContainerTaskCount: List[TaskInfo] = field(default_factory=list)
    ForLoopContainerTaskCount: List[TaskInfo] = field(default_factory=list)
    Variables: List[VariableInfo] = field(default_factory=list)
    DataFlowTaskDetails: List[DataFlowTaskInfo] = field(default_factory=list)
    PrecedenceConstraintDetails: List[PrecedenceConstraintInfo] = field(default_factory=list)
    ExtractTaskDetails: List[TaskInfo] = field(default_factory=list)
    ProjectParameterDetails: List[ProjectParameterInfo] = field(default_factory=list)

class SSISPackageAnalyzer:
    def __init__(self, package_folder, metadata_connection_string, package_analysis_file_path, dataflow_file_path, package_details_file_path, data_save_type):
        self.container_count = 0
        self.container_task_count = 0
        self._connection_string = metadata_connection_string
        self._package_folder = package_folder
        self.processed_package_paths = set()
        self.PackagePath = ""
        self.PackageName = ""
        self.ComponentCount = 0
        self.PackageAnalysisFilePath = package_analysis_file_path
        self.DataFlowlFilePath = dataflow_file_path
        self.PackageDetailsFilePath = package_details_file_path
        self.DataSaveType = data_save_type
        self.ComponentNameCheck = []
        self.variables_metadata = []

    def truncate_table(self):
        if self.DataSaveType.upper() == "SQL":
            try:
                conn = pyodbc.connect(self._connection_string)
                cursor = conn.cursor()
                connection_query = """
                    TRUNCATE TABLE PackageAnalysisResults;
                    TRUNCATE TABLE PackageTaskDetails;
                    TRUNCATE TABLE PackageConnectionDetails;
                    TRUNCATE TABLE PackageContainerDetails;
                    TRUNCATE TABLE ProjectParameterDetails;
                    TRUNCATE TABLE PackageVariableParameterDetails;
                    TRUNCATE TABLE DataFlowTaskMappingDetails;
                    TRUNCATE TABLE PrecedenceConstraintDetails;
                    TRUNCATE TABLE EventTaskDetails;
                """
                cursor.execute(connection_query)
                conn.commit()
                cursor.close()
                conn.close()
                print("Truncated all metadata tables.")
            except Exception as e:
                self.log_error("SQL Truncate", e)
 
    def analyze_all_packages(package_folder):
        truncate_table()  # Your function to truncate the target table

        for root, dirs, files in os.walk(package_folder):
            if "\\obj\\" in root.lower():
                continue  # Skip obj folders

            try:
                package_files = [os.path.join(root, f) for f in files if f.endswith(".dtsx")]
                connection_manager_files = [os.path.join(root, f) for f in files if f.endswith(".conmgr")]
                param_files = [os.path.join(root, f) for f in files if f.endswith(".params")]

                for package_path in package_files:
                    if package_path in processed_package_paths:
                        continue
                    try:
                        processed_package_paths.add(package_path)
                        analyze_single_package(package_path)
                    except Exception as ex:
                        log_error(package_path, ex)

                for conn_path in connection_manager_files:
                    if conn_path in processed_package_paths:
                        continue
                    try:
                        processed_package_paths.add(conn_path)
                        analyze_single_connection_manager(conn_path)
                    except Exception as ex:
                        log_error(conn_path, ex)

                for param_path in param_files:
                    if param_path in processed_package_paths:
                        continue
                    try:
                        processed_package_paths.add(param_path)
                        analyze_param_manager(param_path)
                    except Exception as ex:
                        log_error(param_path, ex)

            except Exception as ex:
                print(f"Error accessing directory {root}: {ex}")

        save_update_connection_name(package_details_file_path)
        print("Completed...")

    
    def analyze_param_manager(param_file_path):
        tree = ET.parse(param_file_path)
        root = tree.getroot()
    
        ns = {'SSIS': 'www.microsoft.com/SqlServer/SSIS'}

        metadata = {
            "ProjectParameterDetails": [],
            "PackagePath": os.path.dirname(param_file_path),
            "PackageName": os.path.basename(param_file_path)
        }

        # SSIS DataType ID to readable format mapping
        ssis_data_types = {
            "3": "Boolean",
            "6": "Byte",
            "16": "DateTime",
            "15": "Decimal",
            "14": "Double",
            "7": "Int16",
            "9": "Int32",
            "11": "Int64",
            "5": "SByte",
            "13": "Single",
            "18": "String",
            "10": "Unit32",
            "12": "Unit64"
        }

        parameter_nodes = root.findall(".//SSIS:Parameter", ns)
        for param in parameter_nodes:
            param_name = param.attrib.get("{www.microsoft.com/SqlServer/SSIS}Name")
            value_node = param.find("SSIS:Properties/SSIS:Property[@SSIS:Name='Value']", ns)
            datatype_node = param.find("SSIS:Properties/SSIS:Property[@SSIS:Name='DataType']", ns)

            value = value_node.text if value_node is not None else None
            datatype_code = datatype_node.text if datatype_node is not None else None
            datatype_name = ssis_data_types.get(datatype_code, datatype_code)

            metadata["ProjectParameterDetails"].append({
                "ParameterName": param_name,
                "DataType": datatype_name,
                "Value": value
            })

        save_project_parameter_metadata(metadata, package_details_file_path)

    def analyze_single_connection_manager(connection_manager_path):
        tree = ET.parse(connection_manager_path)
        root = tree.getroot()

        # Define the namespace mapping
        ns = {'DTS': 'www.microsoft.com/SqlServer/Dts'}

        connection_string_name = ""
        connection_name = ""
        connection_id = ""
        connection_expression = ""
        connection_type = ""

        # Extract specific attributes using XPath
        conn_string_node = root.find(".//DTS:ConnectionManager/DTS:ObjectData/DTS:ConnectionManager", ns)
        connection_name_node = root.find(".//DTS:ConnectionManager", ns)
        connection_type_node = root.find(".//DTS:ConnectionManager", ns)
        connection_id_node = root.find(".//DTS:ConnectionManager", ns)

        if conn_string_node is not None:
            connection_string_name = conn_string_node.attrib.get("{www.microsoft.com/SqlServer/Dts}ConnectionString", "")

        if connection_name_node is not None:
            connection_name = connection_name_node.attrib.get("{www.microsoft.com/SqlServer/Dts}ObjectName", "")

        if connection_type_node is not None:
            connection_type = connection_type_node.attrib.get("{www.microsoft.com/SqlServer/Dts}CreationName", "")

        if connection_id_node is not None:
            connection_id = connection_id_node.attrib.get("{www.microsoft.com/SqlServer/Dts}DTSID", "")

        # Handle property expressions
        for prop_expr_node in root.findall(".//DTS:PropertyExpression", ns):
            name = prop_expr_node.attrib.get("{www.microsoft.com/SqlServer/Dts}Name", "Name not found.")
            value = prop_expr_node.text or ""
            connection_expression += f"{name} : {value} "

        metadata = {
            "Connections": [],
            "PackagePath": os.path.dirname(connection_manager_path),
            "PackageName": os.path.basename(connection_manager_path)
        }

        metadata["Connections"].append({
            "ConnectionName": connection_name,
            "ConnectionString": connection_string_name,
            "ConnectionExpressions": connection_expression.strip(),
            "ConnectionType": connection_type,
            "ConnectionID": connection_id,
            "IsProjectConnection": "1"
        })

        save_connections_metadata(metadata, package_details_file_path)

    def analyze_single_package(package_path):
        app = Application()  # Assuming a wrapper for SSIS Application
        package = app.load_package(package_path, None)  # Placeholder for loading package

        component_name_check = []

        try:
            tree = ET.parse(package_path)
            root = tree.getroot()
            traverse_xml(root)
        except Exception as ex:
            print(f"Error: {ex}")

        package_name = os.path.basename(package_path)
        package_dir = os.path.dirname(package_path)

        metadata = {
            "PackageName": package_name,
            "CreatedDate": package.creation_date,
            "CreatedBy": package.creator_name,
            "Tasks": count_package_tasks(package),
            "Connections": count_package_connections(package),
            "PackagePath": package_dir,
            "Containers": count_package_containers(package),
            "DTSXXML": ET.tostring(root, encoding="unicode"),
            "Seqtasks": [],
            "Foreachtasks": [],
            "Forlooptasks": [],
            "Variables": get_package_variables(package),
            "DataFlowTaskDetails": [],
        }

        for executable in package.executables:
            if isinstance(executable, ForEachLoop):
                metadata["Foreachtasks"].extend(
                    process_foreach_loop_container_details(executable, [], package)
                )

            elif isinstance(executable, Sequence):
                metadata["Seqtasks"].extend(
                    process_sequence_container_details(executable, [], package)
                )

            elif isinstance(executable, ForLoop):
                metadata["Forlooptasks"].extend(
                    process_for_loop_container_details(executable, [], package)
                )

            elif isinstance(executable, TaskHost):
                if isinstance(executable.inner_object, MainPipe):
                    extract_data_flow_task(executable, "0")

        metadata["SequenceContainerTaskCount"] = count_sequence_container_tasks(package)
        metadata["ForeachContainerTaskCount"] = count_foreache_container_tasks(package)
        metadata["ForLoopContainerTaskCount"] = count_forloop_container_tasks(package)
        metadata["ExecutionTime"] = measure_package_performance(package)

        save_package_metadata(
            metadata,
            package_analysis_file_path,
            package_details_file_path
        )

        extract_precedence_constraints_for_task(package)
        extract_event_handlers_for_package(package)
    
    def traverse_xml(node: ET.Element):
        if node is not None:
            for child in node:
                traverse_xml(child)
                                
    def get_package_variables(package: Package) -> list:
        variables = []

        for variable in package.variables:
            if not variable.system_variable:
                variables.append(VariableInfo(
                    name=variable.name,
                    value=str(variable.value) if variable.value is not None else None,
                    data_type=str(variable.data_type),
                    namespace=variable.namespace,
                    is_parameter=0
                ))

        for parameter in package.parameters:
            variables.append(VariableInfo(
                name=parameter.name,
                value=str(parameter.value) if parameter.value is not None else None,
                data_type=str(parameter.data_type),
                is_parameter=1
            ))

        return variables
                             
    def count_sequence_container_tasks(package: Package) -> list:
        tasks_in_sequence = []

        for executable in package.executables:
            if isinstance(executable, Sequence):
                process_container_sequence_loop(executable, tasks_in_sequence, package)

            elif isinstance(executable, ForEachLoop):
                process_container_foreach_loop(executable, tasks_in_sequence, package)

            elif isinstance(executable, ForLoop):
                process_container_for_loop(executable, tasks_in_sequence, package)

        container_task_count += len(tasks_in_sequence)

        return tasks_in_sequence

                              
    def count_foreach_container_tasks(package: Package) -> list:
        tasks_in_for_each = []

        for executable in package.executables:
            if isinstance(executable, ForEachLoop):
                process_container_foreach_loop(executable, tasks_in_for_each, package)

            elif isinstance(executable, Sequence):
                process_container_sequence_loop(executable, tasks_in_for_each, package)

            elif isinstance(executable, ForLoop):
                process_container_for_loop(executable, tasks_in_for_each, package)

        container_task_count += len(tasks_in_for_each)

        return tasks_in_for_each

                                             
    def count_forloop_container_tasks(package: Package) -> list:
        tasks_in_for_loop = []

        for executable in package.executables:
            if isinstance(executable, ForEachLoop):
                process_container_foreach_loop(executable, tasks_in_for_loop, package)

            elif isinstance(executable, Sequence):
                process_container_sequence_loop(executable, tasks_in_for_loop, package)

            elif isinstance(executable, ForLoop):
                process_container_for_loop(executable, tasks_in_for_loop, package)

        container_task_count += len(tasks_in_for_loop)
        
        return tasks_in_for_loop

    def process_container_foreach_loop(container: ForEachLoop, tasks_in_for_each: list, package):
        # Check if the container is a ForEachLoop (redundant in Python, but included for structure)
        if isinstance(container, ForEachLoop):
            pass  # Placeholder for any logic specific to ForEachLoop

        # Iterate over nested executables
        for nested_executable in container.executables:
            if isinstance(nested_executable, ForEachLoop):
                tasks_in_loop = process_foreach_loop_container_details(nested_executable, [], package)
                tasks_in_for_each.extend(tasks_in_loop)

            elif isinstance(nested_executable, Sequence):
                tasks_in_loop = process_sequence_container_details(nested_executable, [], package)
                tasks_in_for_each.extend(tasks_in_loop)

            elif isinstance(nested_executable, ForLoop):
                tasks_in_loop = process_for_loop_container_details(nested_executable, [], package)
                tasks_in_for_each.extend(tasks_in_loop)
                
                               
    def process_container_sequence_loop(container: Sequence, tasks_in_for_each: list, package):
        # Check if the container is a Sequence (redundant in Python but included for clarity)
        if isinstance(container, Sequence):
            pass  # Placeholder for any logic specific to Sequence containers

        # Iterate over nested executables
        for nested_executable in container.executables:
            if isinstance(nested_executable, ForEachLoop):
                tasks_in_loop = process_foreach_loop_container_details(nested_executable, [], package)
                tasks_in_for_each.extend(tasks_in_loop)

            elif isinstance(nested_executable, Sequence):
                tasks_in_loop = process_sequence_container_details(nested_executable, [], package)
                tasks_in_for_each.extend(tasks_in_loop)

            elif isinstance(nested_executable, ForLoop):
                tasks_in_loop = process_for_loop_container_details(nested_executable, [], package)
                tasks_in_for_each.extend(tasks_in_loop)
                

    def process_container_for_loop(container: ForLoop, tasks_in_for_each: list, package):
        # Check if the container is a ForLoop (redundant in Python, as it's already typed)
        if isinstance(container, ForLoop):
            pass  # Placeholder for any logic if needed for the base ForLoop

        # Iterate over nested executables
        for nested_executable in container.Executables:
            if isinstance(nested_executable, ForEachLoop):
                tasks_in_loop = process_foreach_loop_container_details(nested_executable, [], package)
                tasks_in_for_each.extend(tasks_in_loop)

            elif isinstance(nested_executable, Sequence):
                tasks_in_loop = process_sequence_container_details(nested_executable, [], package)
                tasks_in_for_each.extend(tasks_in_loop)

            elif isinstance(nested_executable, ForLoop):
                tasks_in_loop = process_for_loop_container_details(nested_executable, [], package)
                tasks_in_for_each.extend(tasks_in_loop)

                                                                            
    def count_package_tasks(package):
        tasks = []

        for executable in package.Executables:
            if isinstance(executable, TaskHost):
                extract_task_details(
                    executable,
                    event_handler_name="",
                    event_handler_type="",
                    event_type="",
                    event_indicator="0",
                    container_name="",
                    container_type="",
                    container_expression="",
                    container_enum_details=""
                )

                tasks.append(TaskInfo(
                TaskName=executable.Name
                ))

        return tasks
                                             
    def extract_event_handlers_for_package(package):
        if package.EventHandlers.Count > 0:
            event_handler_name = package.Name
            event_handler_type = "Package"
            event_name = ""

            for event_handler in package.EventHandlers:
                event_name = event_handler.Name

                for event_executable in event_handler.Executables:
                    if isinstance(event_executable, TaskHost):
                        extract_event_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name,
                            "",
                            "",
                            "",
                            ""
                        )

                    elif isinstance(event_executable, Sequence):
                        extract_event_sequence_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name
                        )

                    elif isinstance(event_executable, ForEachLoop):
                        extract_event_foreach_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name
                        )

                    elif isinstance(event_executable, ForLoop):
                        extract_event_for_loop_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name
                        )
                               
    def extract_event_handlers_for_sequence(sequence):
        if sequence.EventHandlers.Count > 0:
            event_handler_name = sequence.Name
            event_handler_type = "Sequence"
            event_name = ""

            for event_handler in sequence.EventHandlers:
                event_name = event_handler.Name

                for event_executable in event_handler.Executables:
                    if isinstance(event_executable, TaskHost):
                        extract_event_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name,
                            "",
                            "",
                            "",
                            ""
                        )

                    elif isinstance(event_executable, Sequence):
                        extract_event_sequence_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name
                        )

                    elif isinstance(event_executable, ForEachLoop):
                        extract_event_foreach_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name
                        )

                    elif isinstance(event_executable, ForLoop):
                        extract_event_for_loop_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name
                        )

                            
    def extract_event_handlers_for_foreach_loop(foreach_loop):
        if foreach_loop.EventHandlers.Count > 0:
            event_handler_name = foreach_loop.Name
            event_handler_type = "ForEachLoop"
            event_name = ""

            for event_handler in foreach_loop.EventHandlers:
                event_name = event_handler.Name

                for event_executable in event_handler.Executables:
                    if isinstance(event_executable, TaskHost):
                        extract_event_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name,
                            "",
                            "",
                            "",
                            ""
                        )

                    elif isinstance(event_executable, Sequence):
                        extract_event_sequence_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name
                        )

                    elif isinstance(event_executable, ForEachLoop):
                        extract_event_foreach_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name
                        )

                    elif isinstance(event_executable, ForLoop):
                        extract_event_for_loop_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name
                        )
                        
                                             
    def extract_event_handlers_for_for_loop(for_loop):
        if for_loop.EventHandlers.Count > 0:
            event_handler_name = for_loop.Name
            event_handler_type = "ForLoop"
            event_name = ""

            for event_handler in for_loop.EventHandlers:
                event_name = event_handler.Name

                for event_executable in event_handler.Executables:
                    if isinstance(event_executable, TaskHost):
                        extract_event_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name,
                            "",
                            "",
                            "",
                            ""
                        )

                    elif isinstance(event_executable, Sequence):
                        extract_event_sequence_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name
                        )

                    elif isinstance(event_executable, ForEachLoop):
                        extract_event_foreach_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name
                        )

                    elif isinstance(event_executable, ForLoop):
                        extract_event_for_loop_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name
                        )


    def extract_event_handlers_for_task(taskhost):
        if taskhost.EventHandlers.Count > 0:
            event_handler_name = taskhost.Name

            if isinstance(taskhost.InnerObject, MainPipe):
                event_handler_type = "DataFlowTask"
            elif isinstance(taskhost.InnerObject, ExecutePackageTask):
                event_handler_type = "ExecutePackageTask"
            else:
                event_handler_type = type(taskhost.InnerObject).__name__

            for event_handler in taskhost.EventHandlers:
                event_name = event_handler.Name

                for event_executable in event_handler.Executables:
                    if isinstance(event_executable, TaskHost):
                        extract_event_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name,
                            "",
                            "",
                            "",
                            ""
                        )

                    elif isinstance(event_executable, Sequence):
                        extract_event_sequence_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name
                        )

                    elif isinstance(event_executable, ForEachLoop):
                        extract_event_foreach_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name
                        )

                    elif isinstance(event_executable, ForLoop):
                        extract_event_for_loop_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name
                        )


    def extract_precedence_constraints_for_task(package):
        metadata = PackageAnalysisResult()
        metadata.PrecedenceConstraintDetails = []

        if package.PrecedenceConstraints.Count == 0:
            for executable in package.Executables:
                if isinstance(executable, Sequence):
                    extract_precedence_constraints_for_sequence(executable)
                elif isinstance(executable, ForEachLoop):
                    extract_precedence_constraints_for_foreach(executable)
                elif isinstance(executable, ForLoop):
                    extract_precedence_constraints_for_forloop(executable)
        else:
            for precedence_constraint in package.PrecedenceConstraints:
                precedence_constraint_from = ""
                precedence_constraint_to = ""
                precedence_constraint_value = str(precedence_constraint.Value)
                precedence_constraint_expression = str(precedence_constraint.Expression)
                precedence_constraint_eval_op = str(precedence_constraint.EvalOp)
                precedence_constraint_logical_and = str(precedence_constraint.LogicalAnd)

                # FROM
                if isinstance(precedence_constraint.PrecedenceExecutable, TaskHost):
                    precedence_constraint_from = precedence_constraint.PrecedenceExecutable.Name
                elif isinstance(precedence_constraint.PrecedenceExecutable, Sequence):
                    precedence_constraint_from = precedence_constraint.PrecedenceExecutable.Name
                    extract_precedence_constraints_for_sequence(precedence_constraint.PrecedenceExecutable)
                elif isinstance(precedence_constraint.PrecedenceExecutable, ForEachLoop):
                    precedence_constraint_from = precedence_constraint.PrecedenceExecutable.Name
                    extract_precedence_constraints_for_foreach(precedence_constraint.PrecedenceExecutable)
                elif isinstance(precedence_constraint.PrecedenceExecutable, ForLoop):
                    precedence_constraint_from = precedence_constraint.PrecedenceExecutable.Name
                    extract_precedence_constraints_for_forloop(precedence_constraint.PrecedenceExecutable)

                # TO
                if isinstance(precedence_constraint.ConstrainedExecutable, TaskHost):
                    precedence_constraint_to = precedence_constraint.ConstrainedExecutable.Name
                elif isinstance(precedence_constraint.ConstrainedExecutable, Sequence):
                    precedence_constraint_to = precedence_constraint.ConstrainedExecutable.Name
                    extract_precedence_constraints_for_sequence(precedence_constraint.ConstrainedExecutable)
                elif isinstance(precedence_constraint.ConstrainedExecutable, ForEachLoop):
                    precedence_constraint_to = precedence_constraint.ConstrainedExecutable.Name
                    extract_precedence_constraints_for_foreach(precedence_constraint.ConstrainedExecutable)
                elif isinstance(precedence_constraint.ConstrainedExecutable, ForLoop):
                    precedence_constraint_to = precedence_constraint.ConstrainedExecutable.Name
                    extract_precedence_constraints_for_forloop(precedence_constraint.ConstrainedExecutable)

                # Add to metadata
                metadata.PrecedenceConstraintDetails.append(PrecedenceConstraintInfo(
                    PackageName=PackageName,
                    PackagePath=PackagePath,
                    PrecedenceConstraintFrom=precedence_constraint_from,
                    PrecedenceConstraintTo=precedence_constraint_to,
                    PrecedenceConstraintValue=precedence_constraint_value,
                    PrecedenceConstraintExpression=precedence_constraint_expression,
                    PrecedenceConstraintEvalOP=precedence_constraint_eval_op,
                    PrecedenceConstraintLogicalAnd=precedence_constraint_logical_and,
                    ContainerName=""
                ))

        save_precedence_constraint_metadata(metadata, PackageDetailsFilePath)

        return metadata.PrecedenceConstraintDetails


    def extract_precedence_constraints_for_sequence(sequence):
        metadata = PackageAnalysisResult()
        metadata.PrecedenceConstraintDetails = []

        if sequence.PrecedenceConstraints.Count == 0:
            for executable in sequence.Executables:
                if isinstance(executable, Sequence):
                    extract_precedence_constraints_for_sequence(executable)
                elif isinstance(executable, ForEachLoop):
                    extract_precedence_constraints_for_foreach(executable)
                elif isinstance(executable, ForLoop):
                    extract_precedence_constraints_for_forloop(executable)
        else:
            for precedence_constraint in sequence.PrecedenceConstraints:
                precedence_constraint_from = ""
                precedence_constraint_to = ""
                precedence_constraint_value = str(precedence_constraint.Value)
                precedence_constraint_expression = str(precedence_constraint.Expression)
                precedence_constraint_eval_op = str(precedence_constraint.EvalOp)
                precedence_constraint_logical_and = str(precedence_constraint.LogicalAnd)

                # FROM
                if isinstance(precedence_constraint.PrecedenceExecutable, TaskHost):
                    precedence_constraint_from = precedence_constraint.PrecedenceExecutable.Name
                elif isinstance(precedence_constraint.PrecedenceExecutable, Sequence):
                    precedence_constraint_from = precedence_constraint.PrecedenceExecutable.Name
                    extract_precedence_constraints_for_sequence(precedence_constraint.PrecedenceExecutable)
                elif isinstance(precedence_constraint.PrecedenceExecutable, ForEachLoop):
                    precedence_constraint_from = precedence_constraint.PrecedenceExecutable.Name
                    extract_precedence_constraints_for_foreach(precedence_constraint.PrecedenceExecutable)
                elif isinstance(precedence_constraint.PrecedenceExecutable, ForLoop):
                    precedence_constraint_from = precedence_constraint.PrecedenceExecutable.Name
                    extract_precedence_constraints_for_forloop(precedence_constraint.PrecedenceExecutable)

                # TO
                if isinstance(precedence_constraint.ConstrainedExecutable, TaskHost):
                    precedence_constraint_to = precedence_constraint.ConstrainedExecutable.Name
                elif isinstance(precedence_constraint.ConstrainedExecutable, Sequence):
                    precedence_constraint_to = precedence_constraint.ConstrainedExecutable.Name
                    extract_precedence_constraints_for_sequence(precedence_constraint.ConstrainedExecutable)
                elif isinstance(precedence_constraint.ConstrainedExecutable, ForEachLoop):
                    precedence_constraint_to = precedence_constraint.ConstrainedExecutable.Name
                    extract_precedence_constraints_for_foreach(precedence_constraint.ConstrainedExecutable)
                elif isinstance(precedence_constraint.ConstrainedExecutable, ForLoop):
                    precedence_constraint_to = precedence_constraint.ConstrainedExecutable.Name
                    extract_precedence_constraints_for_forloop(precedence_constraint.ConstrainedExecutable)

                # Add to metadata
                metadata.PrecedenceConstraintDetails.append(PrecedenceConstraintInfo(
                    PackageName=PackageName,
                    PackagePath=PackagePath,
                    PrecedenceConstraintFrom=precedence_constraint_from,
                    PrecedenceConstraintTo=precedence_constraint_to,
                    PrecedenceConstraintValue=precedence_constraint_value,
                    PrecedenceConstraintExpression=precedence_constraint_expression,
                    PrecedenceConstraintEvalOP=precedence_constraint_eval_op,
                    PrecedenceConstraintLogicalAnd=precedence_constraint_logical_and,
                    ContainerName=sequence.Name
                ))

        save_precedence_constraint_metadata(metadata, PackageDetailsFilePath)
        return metadata.PrecedenceConstraintDetails


    def extract_precedence_constraints_for_foreach(for_each):
        metadata = PackageAnalysisResult()
        metadata.PrecedenceConstraintDetails = []

        if for_each.PrecedenceConstraints.Count == 0:
            for executable in for_each.Executables:
                if isinstance(executable, Sequence):
                    extract_precedence_constraints_for_sequence(executable)
                elif isinstance(executable, ForEachLoop):
                    extract_precedence_constraints_for_foreach(executable)
                elif isinstance(executable, ForLoop):
                    extract_precedence_constraints_for_forloop(executable)
        else:
            for precedence_constraint in for_each.PrecedenceConstraints:
                precedence_constraint_from = ""
                precedence_constraint_to = ""
                precedence_constraint_value = str(precedence_constraint.Value)
                precedence_constraint_expression = str(precedence_constraint.Expression)
                precedence_constraint_eval_op = str(precedence_constraint.EvalOp)
                precedence_constraint_logical_and = str(precedence_constraint.LogicalAnd)

                # FROM
                if isinstance(precedence_constraint.PrecedenceExecutable, TaskHost):
                    precedence_constraint_from = precedence_constraint.PrecedenceExecutable.Name
                elif isinstance(precedence_constraint.PrecedenceExecutable, Sequence):
                    precedence_constraint_from = precedence_constraint.PrecedenceExecutable.Name
                    extract_precedence_constraints_for_sequence(precedence_constraint.PrecedenceExecutable)
                elif isinstance(precedence_constraint.PrecedenceExecutable, ForEachLoop):
                    precedence_constraint_from = precedence_constraint.PrecedenceExecutable.Name
                    extract_precedence_constraints_for_foreach(precedence_constraint.PrecedenceExecutable)
                elif isinstance(precedence_constraint.PrecedenceExecutable, ForLoop):
                    precedence_constraint_from = precedence_constraint.PrecedenceExecutable.Name
                    extract_precedence_constraints_for_forloop(precedence_constraint.PrecedenceExecutable)

                # TO
                if isinstance(precedence_constraint.ConstrainedExecutable, TaskHost):
                    precedence_constraint_to = precedence_constraint.ConstrainedExecutable.Name
                elif isinstance(precedence_constraint.ConstrainedExecutable, Sequence):
                    precedence_constraint_to = precedence_constraint.ConstrainedExecutable.Name
                    extract_precedence_constraints_for_sequence(precedence_constraint.ConstrainedExecutable)
                elif isinstance(precedence_constraint.ConstrainedExecutable, ForEachLoop):
                    precedence_constraint_to = precedence_constraint.ConstrainedExecutable.Name
                    extract_precedence_constraints_for_foreach(precedence_constraint.ConstrainedExecutable)
                elif isinstance(precedence_constraint.ConstrainedExecutable, ForLoop):
                    precedence_constraint_to = precedence_constraint.ConstrainedExecutable.Name
                    extract_precedence_constraints_for_forloop(precedence_constraint.ConstrainedExecutable)

                # Append metadata
                metadata.PrecedenceConstraintDetails.append(PrecedenceConstraintInfo(
                    PackageName=PackageName,
                    PackagePath=PackagePath,
                    PrecedenceConstraintFrom=precedence_constraint_from,
                    PrecedenceConstraintTo=precedence_constraint_to,
                    PrecedenceConstraintValue=precedence_constraint_value,
                    PrecedenceConstraintExpression=precedence_constraint_expression,
                    PrecedenceConstraintEvalOP=precedence_constraint_eval_op,
                    PrecedenceConstraintLogicalAnd=precedence_constraint_logical_and,
                    ContainerName=for_each.Name
                ))

        save_precedence_constraint_metadata(metadata, PackageDetailsFilePath)
        return metadata.PrecedenceConstraintDetails


    def extract_precedence_constraints_for_forloop(for_loop):
        metadata = PackageAnalysisResult()
        metadata.PrecedenceConstraintDetails = []

        if for_loop.PrecedenceConstraints.Count == 0:
            for executable in for_loop.Executables:
                if isinstance(executable, Sequence):
                    extract_precedence_constraints_for_sequence(executable)
                elif isinstance(executable, ForEachLoop):
                    extract_precedence_constraints_for_foreach(executable)
                elif isinstance(executable, ForLoop):
                    extract_precedence_constraints_for_forloop(executable)
        else:
            for precedence_constraint in for_loop.PrecedenceConstraints:
                precedence_constraint_from = ""
                precedence_constraint_to = ""
                precedence_constraint_value = str(precedence_constraint.Value)
                precedence_constraint_expression = str(precedence_constraint.Expression)
                precedence_constraint_eval_op = str(precedence_constraint.EvalOp)
                precedence_constraint_logical_and = str(precedence_constraint.LogicalAnd)

                # Extract "from" executable name
                if isinstance(precedence_constraint.PrecedenceExecutable, TaskHost):
                    precedence_constraint_from = precedence_constraint.PrecedenceExecutable.Name
                elif isinstance(precedence_constraint.PrecedenceExecutable, Sequence):
                    precedence_constraint_from = precedence_constraint.PrecedenceExecutable.Name
                    extract_precedence_constraints_for_sequence(precedence_constraint.PrecedenceExecutable)
                elif isinstance(precedence_constraint.PrecedenceExecutable, ForEachLoop):
                    precedence_constraint_from = precedence_constraint.PrecedenceExecutable.Name
                    extract_precedence_constraints_for_foreach(precedence_constraint.PrecedenceExecutable)
                elif isinstance(precedence_constraint.PrecedenceExecutable, ForLoop):
                    precedence_constraint_from = precedence_constraint.PrecedenceExecutable.Name
                    extract_precedence_constraints_for_forloop(precedence_constraint.PrecedenceExecutable)

                # Extract "to" executable name
                if isinstance(precedence_constraint.ConstrainedExecutable, TaskHost):
                    precedence_constraint_to = precedence_constraint.ConstrainedExecutable.Name
                elif isinstance(precedence_constraint.ConstrainedExecutable, Sequence):
                    precedence_constraint_to = precedence_constraint.ConstrainedExecutable.Name
                    extract_precedence_constraints_for_sequence(precedence_constraint.ConstrainedExecutable)
                elif isinstance(precedence_constraint.ConstrainedExecutable, ForEachLoop):
                    precedence_constraint_to = precedence_constraint.ConstrainedExecutable.Name
                    extract_precedence_constraints_for_foreach(precedence_constraint.ConstrainedExecutable)
                elif isinstance(precedence_constraint.ConstrainedExecutable, ForLoop):
                    precedence_constraint_to = precedence_constraint.ConstrainedExecutable.Name
                    extract_precedence_constraints_for_forloop(precedence_constraint.ConstrainedExecutable)

                metadata.PrecedenceConstraintDetails.append(PrecedenceConstraintInfo(
                    PackageName=PackageName,
                    PackagePath=PackagePath,
                    PrecedenceConstraintFrom=precedence_constraint_from,
                    PrecedenceConstraintTo=precedence_constraint_to,
                    PrecedenceConstraintValue=precedence_constraint_value,
                    PrecedenceConstraintExpression=precedence_constraint_expression,
                    PrecedenceConstraintEvalOP=precedence_constraint_eval_op,
                    PrecedenceConstraintLogicalAnd=precedence_constraint_logical_and,
                    ContainerName=for_loop.Name
                ))

        save_precedence_constraint_metadata(metadata, PackageDetailsFilePath)
        return metadata.PrecedenceConstraintDetails

                        
    def extract_event_task_details(task_host, event_handler_name, event_handler_type, event_type,
                                container_name, container_type, container_expression, container_enum_details):
        extract_task_details(
            task_host,
            event_handler_name,
            event_handler_type,
            event_type,
            "1",  # Event indicator
            container_name,
            container_type,
            container_expression,
            container_enum_details
        )

    def extract_event_sequence_task_details(sequence, event_handler_name, event_handler_type, event_name):
        container_name = sequence.Name
        container_type = "Sequence"

        for event_executable in sequence.Executables:
            if isinstance(event_executable, TaskHost):
                extract_event_task_details(
                    event_executable,
                    event_handler_name,
                    event_handler_type,
                    event_name,
                    container_name,
                    container_type,
                    "",
                    ""
                )
            elif isinstance(event_executable, Sequence):
                extract_event_sequence_task_details(
                    event_executable,
                    event_handler_name,
                    event_handler_type,
                    event_name
                )
            elif isinstance(event_executable, ForEachLoop):
                extract_event_foreach_task_details(
                    event_executable,
                    event_handler_name,
                    event_handler_type,
                    event_name
                )
            elif isinstance(event_executable, ForLoop):
                extract_event_for_loop_task_details(
                    event_executable,
                    event_handler_name,
                    event_handler_type,
                    event_name
                )

    def extract_event_foreach_task_details(foreach_loop, event_handler_name, event_handler_type, event_name):
        container_name = foreach_loop.Name
        container_type = "ForEachLoop"
        container_expression = get_for_each_loop_expressions(foreach_loop)
        container_enum = get_for_each_loop_enumerator(foreach_loop)

        for event_executable in foreach_loop.Executables:
            if isinstance(event_executable, TaskHost):
                extract_event_task_details(
                    event_executable,
                    event_handler_name,
                    event_handler_type,
                    event_name,
                    container_name,
                    container_type,
                    container_expression,
                    container_enum
                )
            elif isinstance(event_executable, Sequence):
                extract_event_sequence_task_details(
                    event_executable,
                    event_handler_name,
                    event_handler_type,
                    event_name
                )
            elif isinstance(event_executable, ForEachLoop):
                extract_event_foreach_task_details(
                    event_executable,
                    event_handler_name,
                    event_handler_type,
                    event_name
                )
            elif isinstance(event_executable, ForLoop):
                extract_event_for_loop_task_details(
                    event_executable,
                    event_handler_name,
                    event_handler_type,
                    event_name
            )

                           
    def extract_event_for_loop_task_details(forloop, event_handler_name, event_handler_type, event_name):
        container_name = forloop.Name
        container_type = "ForLoop"
        container_expression = get_for_loop_expressions(forloop)
        container_enum = get_for_loop_enumerator(forloop)

        for event_executable in forloop.Executables:
            if isinstance(event_executable, TaskHost):
                extract_event_task_details(
                    event_executable,
                    event_handler_name,
                    event_handler_type,
                    event_name,
                    container_name,
                    container_type,
                    container_expression,
                    container_enum
                )

            elif isinstance(event_executable, Sequence):
                extract_event_sequence_task_details(
                    event_executable,
                    event_handler_name,
                    event_handler_type,
                    event_name
                )

            elif isinstance(event_executable, ForEachLoop):
                extract_event_foreach_task_details(
                    event_executable,
                    event_handler_name,
                    event_handler_type,
                    event_name
                )

            elif isinstance(event_executable, ForLoop):
                extract_event_for_loop_task_details(
                    event_executable,
                    event_handler_name,
                    event_handler_type,
                    event_name
                )

                                  
    def extract_variables_for_task(task_host):
        variables_used = []

        task = task_host.InnerObject
        task_type_name = type(task).__name__

        if task_type_name == "ExecuteSQLTask":
            sql_statement = getattr(task, "SqlStatementSource", "")
            connection_string = getattr(task, "Connection", "")

            if sql_statement:
                expression_variables = extract_variables_from_expression(sql_statement)
                variables_used.extend(expression_variables)
            elif connection_string:
                connection_variables = extract_variables_from_expression(connection_string)
                variables_used.extend(connection_variables)

        elif task_type_name == "FileSystemTask":
            is_source_path_variable = getattr(task, "IsSourcePathVariable", False)
            is_destination_path_variable = getattr(task, "IsDestinationPathVariable", False)

            if is_source_path_variable:
                source_path = getattr(task, "Source", "")
                variables_used.append(f"Source Path: {source_path}")
            if is_destination_path_variable:
                destination_path = getattr(task, "Destination", "")
                variables_used.append(f"Destination Path: {destination_path}")

        elif task_type_name == "ScriptTask":
            read_only_variables = getattr(task, "ReadOnlyVariables", "").split(',')
            read_write_variables = getattr(task, "ReadWriteVariables", "").split(',')

            variables_used.extend([var.strip() for var in read_only_variables if var.strip()])
            variables_used.extend([var.strip() for var in read_write_variables if var.strip()])

        return ", ".join(variables_used)

    def extract_variables_from_expression(expression: str) -> list[str]:
        """
        Extracts variable names from an expression string formatted like @[User::VariableName]
        """
        variables = []
        try:
            pattern = re.compile(r'@\[(.*?)\]')
            matches = pattern.findall(expression)
            variables = matches
        except Exception as e:
            print(f"Error extracting variables: {e}")

        return variables
                                             
    def extract_parameters_for_task(self, task_host):
        parameters_used = []

        if isinstance(task_host.InnerObject, ExecuteSQLTask):
            sql_task = task_host.InnerObject

            try:
                parameter_bindings_prop = sql_task.GetType().GetProperty("ParameterBindings", BindingFlags.Public | BindingFlags.NonPublic | BindingFlags.Instance)
                if parameter_bindings_prop is not None:
                    parameter_bindings = parameter_bindings_prop.GetValue(sql_task)

                    for binding in parameter_bindings:
                        name_prop = binding.GetType().GetProperty("ParameterName")
                        direction_prop = binding.GetType().GetProperty("ParameterDirection")
                        data_type_prop = binding.GetType().GetProperty("DataType")
                        value_prop = binding.GetType().GetProperty("Value")
                        dts_variable_prop = binding.GetType().GetProperty("DtsVariableName")

                        parameter_name = name_prop.GetValue(binding) if name_prop else ""
                        parameter_type = direction_prop.GetValue(binding) if direction_prop else ""
                        data_type = data_type_prop.GetValue(binding) if data_type_prop else ""
                        value = value_prop.GetValue(binding) if value_prop else ""
                        dts_variable_name = dts_variable_prop.GetValue(binding) if dts_variable_prop else ""

                        param_str = f"Name: {parameter_name}, Type: {parameter_type}, DataType: {data_type}, Value: {value}, DtsVariableName: {dts_variable_name}"
                        parameters_used.append(param_str)

            except Exception as ex:
                print(f"Error while extracting parameters: {str(ex)}")

        return " | ".join(parameters_used)


    def extract_data_flow_task(self, task_host, event_handle):
        metadata = PackageAnalysisResult()
        metadata.DataFlowTaskDetails = []

        if isinstance(task_host.InnerObject, MainPipe):
            data_flow_task = task_host.InnerObject
            for component in data_flow_task.ComponentMetaDataCollection:
                cm_check = f"{task_host.Name} : {self.PackageName} : {self.PackagePath} : {component.Name}"
                if cm_check in self.ComponentNameCHeck:
                    self.ComponentCount += 0
                else:
                    self.ComponentNameCHeck.add(cm_check)
                    self.ComponentCount += 1

                for input_ in component.InputCollection:
                    component_property_details = ""
                    for input_col in input_.InputColumnCollection:
                        column_property_details = ""
                        for prop in input_col.CustomPropertyCollection:
                            column_property_details += f"Property name: {prop.Name}, value: {prop.Value} , Exp: {prop.ExpressionType} "

                        metadata.DataFlowTaskDetails.append(DataFlowTaskInfo(
                            ColumnName=input_col.Name,
                            DataType=str(input_col.DataType),
                            componentName=component.Name,
                            TaskName=task_host.Name,
                            PackageName=self.PackageName,
                            PackagePath=self.PackagePath,
                            ColumnType=str(input_col.ObjectType),
                            isEventHandler=event_handle,
                            ColumnPropertyDetails=column_property_details
                        ))

                    for prop in input_.CustomPropertyCollection:
                        component_property_details += f"Property name: {prop.Name}, value: {prop.Value} , Exp: {prop.ExpressionType}"

                    if component_property_details:
                        metadata.DataFlowTaskDetails.append(DataFlowTaskInfo(
                            componentName=component.Name,
                            TaskName=task_host.Name,
                            PackageName=self.PackageName,
                            PackagePath=self.PackagePath,
                            componentPropertyDetails=f"Companaed Type: {input_.Name}, {component_property_details}",
                            isEventHandler=event_handle
                        ))

                for output in component.OutputCollection:
                    component_property_details = ""
                    for output_col in output.OutputColumnCollection:
                        if any(err in output_col.Name for err in ["Error", "ErrorCode", "ErrorColumn"]):
                            continue

                        column_property_details = ""
                        for prop in output_col.CustomPropertyCollection:
                            column_property_details += f"Property name: {prop.Name}, value: {prop.Value} , Exp: {prop.ExpressionType} "

                        metadata.DataFlowTaskDetails.append(DataFlowTaskInfo(
                            ColumnName=output_col.Name,
                            DataType=str(output_col.DataType),
                            componentName=component.Name,
                            TaskName=task_host.Name,
                            PackageName=self.PackageName,
                            PackagePath=self.PackagePath,
                            ColumnType=str(output_col.ObjectType),
                            isEventHandler=event_handle,
                            ColumnPropertyDetails=column_property_details
                        ))

                    for prop in output.CustomPropertyCollection:
                        component_property_details += f"Property name: {prop.Name}, value: {prop.Value} , Exp: {prop.ExpressionType}"

                    if component_property_details:
                        metadata.DataFlowTaskDetails.append(DataFlowTaskInfo(
                            componentName=component.Name,
                            TaskName=task_host.Name,
                            PackageName=self.PackageName,
                            PackagePath=self.PackagePath,
                            componentPropertyDetails=f"Companaed Type: {output.Name}, {component_property_details}",
                            isEventHandler=event_handle
                        ))

            for component in data_flow_task.ComponentMetaDataCollection:
                if "Data Conversion" in component.Name:
                    for input_col in component.InputCollection[0].InputColumnCollection:
                        metadata.DataFlowTaskDetails.append(DataFlowTaskInfo(
                            ColumnName=input_col.Name,
                            DataType=str(input_col.DataType),
                            componentName=component.Name,
                            TaskName=task_host.Name,
                            DataConversion=str(component.OutputCollection[0].OutputColumnCollection[0].DataType),
                            PackageName=self.PackageName,
                            PackagePath=self.PackagePath,
                            ColumnType="Data Conversion :" + str(input_col.ObjectType),
                            isEventHandler=event_handle
                        ))

                    for output_col in component.OutputCollection[0].OutputColumnCollection:
                        if any(err in output_col.Name for err in ["Error", "ErrorCode", "ErrorColumn"]):
                            continue
                        metadata.DataFlowTaskDetails.append(DataFlowTaskInfo(
                            ColumnName=output_col.Name,
                            DataType=str(output_col.DataType),
                            componentName=component.Name,
                            TaskName=task_host.Name,
                            DataConversion=str(component.OutputCollection[0].OutputColumnCollection[0].DataType),
                            PackageName=self.PackageName,
                            PackagePath=self.PackagePath,
                            ColumnType="Data Conversion :" + str(output_col.ObjectType),
                            isEventHandler=event_handle
                        ))

        self.SaveDataFlowMetadata(metadata, self.DataFlowlFilePath)
        return metadata.DataFlowTaskDetails

                                 
    def match_columns(input_column, output_column):
        """
        Checks if input and output SSIS columns match by name or data type.

        Args:
            input_column: An object representing the SSIS input column.
            output_column: An object representing the SSIS output column.

        Returns:
            bool: True if the columns match by name or data type, else False.
        """

        # Match by name
        if input_column.Name == output_column.Name:
            return True

        # Match by data type
        if input_column.DataType == output_column.DataType:
            return True

        # No match
        return False

                                             
    def extract_expressions_for_task(self, task_host):
        expressions_used = []
        expression_details = ""

        try:
            task = task_host.InnerObject
            task_type = type(task)

            if task_host.HasExpressions:
                for prop in task_host.Properties:
                    try:
                        expression = task_host.GetExpression(prop.Name)
                        if expression:
                            expressions_used.append(f"Property: {prop.Name}, Expression: {expression}")
                    except Exception as ex:
                        print(f"Error extracting expression for property {prop.Name}: {str(ex)}")

            if isinstance(task, MainPipe):  # Data Flow Task
                for component in task.ComponentMetaDataCollection:
                    for custom_property in component.CustomPropertyCollection:
                        if custom_property.Name == "Expression":
                            expressions_used.append(
                                f"Expression Name: {custom_property.Name} Expression Value: {custom_property.Value}"
                            )
            else:
                for prop_info in task_type.GetProperties():
                    try:
                        expression = task_host.GetExpression(prop_info.Name)
                        if expression:
                            expressions_used.append(f"Property: {prop_info.Name}, Expression: {expression}")
                    except Exception as ex:
                        print(f"Error reading expression from property {prop_info.Name}: {str(ex)}")

            return ", ".join(expressions_used)

        except Exception as ex:
            print(f"Error while extracting expressions: {str(ex)}")
            return ""


    def count_package_connections(self, package):
        """
        Extracts all the connection metadata from a given SSIS package and returns a list of ConnectionInfo objects.
        """
        connections = []

        for conn in package.Connections:
            connection_details = ""
            expression_details = []

            for prop in conn.Properties:
                try:
                    expression = conn.GetExpression(prop.Name)
                    if expression:
                        expression_details.append(f"{prop.Name}: {expression}")
                except Exception as ex:
                    print(f"Error accessing expression for property {prop.Name}: {str(ex)}")

            if expression_details:
                connection_details = "Expressions: " + ", ".join(expression_details)

            connections.append(ConnectionInfo(
                ConnectionName=conn.Name,
                ConnectionString=conn.ConnectionString,
                ConnectionExpressions=connection_details,
                ConnectionType=conn.CreationName,
                ConnectionID=conn.ID,
                IsProjectConnection="0"
            ))

        return connections

    def count_package_containers(self, package):
        """
        Counts all the top-level containers in the SSIS package and extracts their metadata.
        """
        containers = []
        expression_details = ""

        for executable in package.Executables:
            if isinstance(executable, DtsContainer) and not isinstance(executable, TaskHost):

                if isinstance(executable, ForEachLoop):
                    expression_details = self.get_foreach_loop_expressions(executable)
                else:
                    expression_details = ""

                # Add base container info
                containers.append(ContainerInfo(
                    ContainerName=executable.Name,
                    ContainerType=type(executable).__name__,
                    ContainerExpression=expression_details
                ))

                # Process container-specific inner executables
                if isinstance(executable, Sequence):
                    self.process_sequence_container(executable, containers)
                elif isinstance(executable, ForEachLoop):
                    self.process_foreach_loop_container(executable, containers)
                elif isinstance(executable, ForLoop):
                    self.process_for_loop_container(executable, containers)

        return containers

    
    def get_foreach_loop_expressions(self, foreach_loop):
        """
        Extracts expression metadata from a ForEachLoop container's enumerator.
        """
        expression_list = []

        try:
            enumerator = foreach_loop.ForEachEnumerator

            if hasattr(enumerator, "Properties"):
                for prop in enumerator.Properties:
                    expression = ""
                    try:
                        # Attempt to retrieve expression using the property name
                        expression = enumerator.GetExpression(prop.Name)
                    except Exception as ex:
                        print(f"Error retrieving expression for {prop.Name}: {str(ex)}")

                    if expression:
                        expression_list.append(f"Property: {prop.Name}, Expression: {expression}")
        except Exception as ex:
            print(f"Error accessing enumerator: {str(ex)}")

        return " | ".join(expression_list)


    def get_foreach_loop_enumerator(self, foreach_loop):
        """
        Extracts enumerator metadata from a ForEachLoop container, excluding default system properties.
        """
        enumerator_details = []

        try:
            enumerator = foreach_loop.ForEachEnumerator

            if enumerator:
                if isinstance(enumerator, ForEachEnumeratorHost):
                    excluded_properties = {
                        "ID",
                        "Description",
                        "CollectionEnumerator",
                        "CreationName",
                        "Name"
                    }

                    for prop in enumerator.Properties:
                        enum_name = prop.Name
                        if enum_name not in excluded_properties:
                            try:
                                value = prop.GetValue(enumerator)
                                enumerator_details.append(f"EnumeratorName: {enum_name}, EnumeratorValue: {value}")
                            except Exception as inner_ex:
                                print(f"Could not get value for property {enum_name}: {str(inner_ex)}")

        except Exception as ex:
            print(f"An error occurred: {str(ex)}")

        return " | ".join(enumerator_details)

    
    def get_for_loop_expressions(self, for_loop):
        """
        Extracts expression properties from a ForLoop container.
        """
        expressions = []

        for prop in for_loop.Properties:
            expression = ""
            try:
                expression = for_loop.GetExpression(prop.Name)
            except Exception as ex:
                print(f"Error retrieving expression for {prop.Name}: {str(ex)}")

            if expression:
                expressions.append(f"Property: {prop.Name}, Expression: {expression}")

        return " | ".join(expressions)
        

    def get_for_loop_enumerator(self, for_loop):
        """
        Returns a string summary of the Assign, Init, and Eval expressions of a ForLoop container.
        """
        try:
            return (f"AssignExpression : {for_loop.AssignExpression} | "
                    f"InitExpression : {for_loop.InitExpression} | "
                    f"EvalExpression : {for_loop.EvalExpression}")
        except Exception as ex:
            print(f"An error occurred: {str(ex)}")
            return ""
            

    def process_sequence_container_details(self, container, containers, package):
        seq_tasks = []
        container_name = container.Name
        container_type = type(container).__name__

        if len(container.EventHandlers) > 0:
            self.extract_event_handlers_for_sequence(container)

        for executable in container.Executables:
            if isinstance(executable, TaskHost):
                self.extract_task_details(
                    executable, "", "", "", "0", container_name, container_type, "", ""
                )

                seq_tasks.append(TaskInfo(
                    SeqTaskName=executable.Name
                ))

            elif isinstance(executable, DtsContainer):
                containers.append(ContainerInfo(
                    ContainerName=executable.Name,
                    ContainerType=type(executable).__name__
                ))

                if isinstance(executable, Sequence):
                    # Optionally extract actual nested task info using: self.process_sequence_container_details(...)
                    seq_tasks.extend(self.count_sequence_container_tasks(package))
                    self.process_sequence_container(executable, containers)

                elif isinstance(executable, ForEachLoop):
                    seq_tasks.extend(self.count_foreach_container_tasks(package))
                    self.process_foreach_loop_container(executable, containers)

                elif isinstance(executable, ForLoop):
                    seq_tasks.extend(self.count_forloop_container_tasks(package))
                    self.process_for_loop_container(executable, containers)

        self.containerTaskCount += len(seq_tasks)
        return seq_tasks

    
    def process_foreach_loop_container_details(self, container, containers, package):
        foreach_tasks = []
        expression_details = self.get_foreach_loop_expressions(container)
        container_name = container.Name
        container_type = type(container).__name__
        enumerator_details = self.get_foreach_loop_enumerator(container)

        if len(container.EventHandlers) > 0:
            self.extract_event_handlers_for_foreach_loop(container)

        for executable in container.Executables:
            if isinstance(executable, TaskHost):
                self.extract_task_details(
                    executable, "", "", "", "0", container_name, container_type, expression_details, enumerator_details
                )

                foreach_tasks.append(TaskInfo(
                    ForeachTaskName=executable.Name
                ))

            elif isinstance(executable, DtsContainer):
                containers.append(ContainerInfo(
                    ContainerName=executable.Name,
                    ContainerType=type(executable).__name__
                ))

                if isinstance(executable, Sequence):
                    foreach_tasks.extend(self.count_sequence_container_tasks(package))
                    self.process_sequence_container(executable, containers)

                elif isinstance(executable, ForEachLoop):
                    foreach_tasks.extend(self.count_foreach_container_tasks(package))
                    self.process_foreach_loop_container(executable, containers)

                elif isinstance(executable, ForLoop):
                    foreach_tasks.extend(self.count_forloop_container_tasks(package))
                    self.process_for_loop_container(executable, containers)

        self.containerTaskCount += len(foreach_tasks)

        return foreach_tasks


    def process_for_loop_container_details(self, container, containers, package):
        for_loop_tasks = []
        expression_details = self.get_for_loop_expressions(container)
        enumerator_details = self.get_for_loop_enumerator(container)

        container_name = container.Name
        container_type = type(container).__name__

        if len(container.EventHandlers) > 0:
            self.extract_event_handlers_for_for_loop(container)

        for executable in container.Executables:
            if isinstance(executable, TaskHost):
                self.extract_task_details(
                    executable, "", "", "", "0", container_name, container_type, expression_details, enumerator_details
                )

                for_loop_tasks.append(TaskInfo(
                    ForloopTaskName=executable.Name
                ))

            elif isinstance(executable, DtsContainer):
                containers.append(ContainerInfo(
                    ContainerName=executable.Name,
                    ContainerType=type(executable).__name__
                ))

                if isinstance(executable, Sequence):
                    for_loop_tasks.extend(self.count_sequence_container_tasks(package))
                    self.process_sequence_container(executable, containers)

                elif isinstance(executable, ForEachLoop):
                    for_loop_tasks.extend(self.count_foreach_container_tasks(package))
                    self.process_foreach_loop_container(executable, containers)

                elif isinstance(executable, ForLoop):
                    for_loop_tasks.extend(self.count_forloop_container_tasks(package))
                    self.process_for_loop_container(executable, containers)

        self.containerTaskCount += len(for_loop_tasks)

        return for_loop_tasks

    def process_sequence_container(self, container, containers):
        """
        Recursively processes a Sequence container and counts the number of TaskHost elements within it.
        """
        task_count = 0

        for executable in container.Executables:
            if isinstance(executable, TaskHost):
                task_count += 1
            elif isinstance(executable, DtsContainer):
                self.container_count += 1

                if isinstance(executable, Sequence):
                    task_count += self.process_sequence_container(executable, containers)

                elif isinstance(executable, ForEachLoop):
                    self.process_foreach_loop_container(executable, containers)

                elif isinstance(executable, ForLoop):
                    self.process_for_loop_container(executable, containers)

        return task_count


    def process_foreach_loop_container(self, container, containers):
        """
        Recursively processes a ForEachLoop container and counts the number of TaskHost elements within it.
        """
        task_count = 0

        for executable in container.Executables:
            if isinstance(executable, TaskHost):
                task_count += 1
            elif isinstance(executable, DtsContainer):
                self.container_count += 1

                if isinstance(executable, ForEachLoop):
                    task_count += self.process_foreach_loop_container(executable, containers)

                elif isinstance(executable, Sequence):
                    self.process_sequence_container(executable, containers)

                elif isinstance(executable, ForLoop):
                    self.process_for_loop_container(executable, containers)

        return task_count


    def process_for_loop_container(self, container, containers):
        """
        Recursively processes a ForLoop container and counts the number of TaskHost elements within it.
        """
        task_count = 0

        for executable in container.Executables:
            if isinstance(executable, TaskHost):
                task_count += 1
            elif isinstance(executable, DtsContainer):
                self.container_count += 1

                if isinstance(executable, ForLoop):
                    task_count += self.process_for_loop_container(executable, containers)

                elif isinstance(executable, Sequence):
                    self.process_sequence_container(executable, containers)

                elif isinstance(executable, ForEachLoop):
                    self.process_foreach_loop_container(executable, containers)

        return task_count

    def extract_task_details(
        self, task_host, event_handler_name, event_handler_type, event_type,
        event_indicator, container_name, container_type, container_expression, enum_details
    ):
        metadata = PackageAnalysisResult()
        metadata.ExtractTaskDetails = []

        sql_query = ""
        execute_process_details = ""
        source_path = destination_path = ""
        source_component_name = target_component_name = ""
        source_type = target_type = ""
        sql_table = target_sql_table = ""
        send_mail_task_details = ftp_task_details = script_task_details = ""
        execute_package_task_details = task_component_details = ""
        xml_task = bulk_insert_task = expression_task = ""
        result_set = ""
        connection_id = source_connection_id = target_connection_id = ""

        task_type = type(task_host.InnerObject)
        task_type_name = task_type.__name__

        if len(task_host.EventHandlers) > 0:
            self.extract_event_handlers_for_task(task_host)

        if isinstance(task_host.InnerObject, MainPipe):
            self.extract_data_flow_task(task_host, event_indicator)
            task_type_name = "DataFlowTask"

            for component in task_host.InnerObject.ComponentMetaDataCollection:
                if "Source" in component.Description:
                    source_component_name = component.Name
                    source_type = component.Description
                    source_connection_id = component.RuntimeConnectionCollection[0].ConnectionManagerID
                    for prop in component.CustomPropertyCollection:
                        if prop.Name == "SqlCommand":
                            sql_query = prop.Value
                        elif prop.Name == "OpenRowset":
                            sql_table = prop.Value
                    if not sql_query:
                        sql_query = sql_table

                elif "Destination" in component.Description:
                    target_component_name = component.Name
                    target_type = component.Description
                    target_connection_id = component.RuntimeConnectionCollection[0].ConnectionManagerID
                    for prop in component.CustomPropertyCollection:
                        if prop.Name == "OpenRowset":
                            target_sql_table = prop.Value

        elif task_type.__name__ == "ExecuteSQLTask":
            connection_id = getattr(task_host.InnerObject, "Connection", "")
            sql_query = getattr(task_host.InnerObject, "SqlStatementSource", "")
            sql_task = task_host.InnerObject
            if hasattr(sql_task, "ResultSetBindings"):
                for binding in sql_task.ResultSetBindings:
                    result_set += f"Result Set Column: {binding.ResultName} | SSIS Variable: {binding.DtsVariableName}  "

        elif isinstance(task_host.InnerObject, FileSystemTask):
            source_path = task_host.InnerObject.Source
            destination_path = task_host.InnerObject.Destination
            task_component_details = f"SourcePath: {source_path} | DestinationPath: {destination_path}"

        elif isinstance(task_host.InnerObject, ExecuteProcess):
            proc = task_host.InnerObject
            execute_process_details = f"Executable: {proc.Executable} | Arguments: {proc.Arguments} | WorkingDirectory: {proc.WorkingDirectory}"
            task_component_details = execute_process_details

        elif isinstance(task_host.InnerObject, SendMailTask):
            mail = task_host.InnerObject
            connection_id = mail.SmtpConnection
            send_mail_task_details = (
                f"From: {mail.FromLine} | To: {mail.ToLine} | CC: {mail.CCLine} "
                f"BCC: {mail.BCCLine} | Subject: {mail.Subject} | Body: {mail.MessageSource} | "
                f"FileAttachments: {mail.FileAttachments} | Priority: {mail.Priority}"
            )
            task_component_details = send_mail_task_details

        elif isinstance(task_host.InnerObject, FtpTask):
            ftp = task_host.InnerObject
            connection_id = ftp.Connection
            ftp_task_details = (
                f"FTP Operation: {ftp.Operation} | LocalPath: {ftp.LocalPath} | RemotePath: {ftp.RemotePath} | "
                f"OverwriteDestination: {ftp.OverwriteDestination} | IsLocalPathVariable: {ftp.IsLocalPathVariable} | "
                f"IsRemotePathVariable: {ftp.IsRemotePathVariable} | IsTransferTypeASCII: {ftp.IsTransferTypeASCII} | "
                f"StopOnOperationFailure: {ftp.StopOnOperationFailure}"
            )
            task_component_details = ftp_task_details

        elif isinstance(task_host.InnerObject, ScriptTask):
            script = task_host.InnerObject
            script_task_details = (
                f"Script Language: {script.ScriptLanguage} | EntryPoint: {script.EntryPoint} | "
                f"ReadOnlyVariables: {script.ReadOnlyVariables} | ReadWriteVariables: {script.ReadWriteVariables} | "
                f"ScriptProjectName: {script.ScriptProjectName}"
            )
            for prop in task_host.Properties:
                if prop.Name == "CodePage":
                    script_task_details += f" | Code Page: {prop.GetValue(script)}"
            task_component_details = script_task_details

        elif isinstance(task_host.InnerObject, ExecutePackageTask):
            execute_package_task_details = task_host.InnerObject.PackageName
            task_type_name = "ExecutePackageTask"
            task_component_details = execute_package_task_details

        elif isinstance(task_host.InnerObject, XMLTask):
            xml = task_host.InnerObject
            xml_task = (
                f"Source: {xml.Source} | SourceType: {xml.SourceType} | DiffAlgorithm: {xml.DiffAlgorithm} | "
                f"DiffGramDestination: {xml.DiffGramDestination} | DiffOptions: {xml.DiffOptions} | "
                f"DiffGramDestinationType: {xml.DiffGramDestinationType} | FailOnDifference: {xml.FailOnDifference} | "
                f"SaveDiffGram: {xml.SaveDiffGram} | OperationType: {xml.OperationType} | "
                f"SaveOperationResult: {xml.SaveOperationResult} | SecondOperand: {xml.SecondOperand} | "
                f"SecondOperandType: {xml.SecondOperandType}"
            )
            task_component_details = xml_task

        elif isinstance(task_host.InnerObject, BulkInsertTask):
            bulk = task_host.InnerObject
            source_connection_id = bulk.SourceConnection
            target_connection_id = bulk.DestinationConnection
            bulk_insert_task = (
                f"FormatFile: {bulk.FormatFile} | FieldTerminator: {bulk.FieldTerminator} | RowTerminator: {bulk.RowTerminator} | "
                f"DestinationTableName: {bulk.DestinationTableName} | CodePage: {bulk.CodePage} | DataFileType: {bulk.DataFileType} | "
                f"BatchSize: {bulk.BatchSize} | LastRow: {bulk.LastRow} | FirstRow: {bulk.FirstRow} | "
                f"CheckConstraints: {bulk.CheckConstraints} | KeepNulls: {bulk.KeepNulls} | KeepIdentity: {bulk.KeepIdentity} | "
                f"TableLock: {bulk.TableLock} | FireTriggers: {bulk.FireTriggers} | SortedData: {bulk.SortedData} | "
                f"MaximumErrors: {bulk.MaximumErrors}"
            )
            task_component_details = bulk_insert_task

        elif isinstance(task_host.InnerObject, ExpressionTask):
            expr = task_host.InnerObject
            expression_task = f"Expression: {expr.Expression} | ExecutionValue: {expr.ExecutionValue}"
            task_component_details = expression_task

        # Add the task info to metadata
        metadata.ExtractTaskDetails.append(TaskInfo(
            TaskName=task_host.Name,
            TaskType=task_type_name,
            TaskSqlQuery=sql_query,
            Variables=self.extract_variables_for_task(task_host),
            FileSystemSourcePath=source_path,
            FileSystemDestinationPath=destination_path,
            Parameters=self.extract_parameters_for_task(task_host),
            Expressions=self.extract_expressions_for_task(task_host),
            ExecuteProcessDetails=execute_process_details,
            SourceComponent=source_component_name,
            TargetComponent=target_component_name,
            SourceType=source_type,
            TargetType=target_type,
            TargetTable=target_sql_table,
            SendMailTask=send_mail_task_details,
            ScriptTask=script_task_details,
            FTPTask=ftp_task_details,
            ExecutePackage=execute_package_task_details,
            ResultSetDetails=result_set,
            EventHandlerName=event_handler_name,
            EventHandlerType=event_handler_type,
            EventType=event_type,
            ContainerName=container_name,
            ContainerType=container_type,
            ContainerExpression=container_expression,
            PackageName=self.PackageName,
            PackagePath=self.PackagePath,
            ContainerEnum=enum_details,
            SourceConnectionName=source_connection_id,
            TargetConnectionName=target_connection_id,
            ConnectionName=connection_id,
            TaskComponentDetails=task_component_details
        ))

        # Save based on event or normal
        if event_indicator == "1":
            self.save_event_metadata(metadata, self.PackageDetailsFilePath)
        else:
            self.save_package_task_metadata(metadata, self.PackageDetailsFilePath)

        return metadata.ExtractTaskDetails


    def measure_package_performance(self, package_obj):
        """
        Executes a mock package and returns the execution time as a timedelta object.
        Assumes 'package_obj' has an 'execute()' method.
        """
        start_time = time.time()
        package_obj.execute()  # This should be your actual package processing logic
        end_time = time.time()
        return end_time - start_time  # Returns seconds as float

    @staticmethod
    def does_workbook_exist(file_path):
        """
        Checks if an Excel workbook exists at the specified path.
        Returns True if exists, otherwise False.
        """
        return os.path.isfile(file_path)

    def save_package_metadata(self, result, analysis_file_path, details_file_path):
        complexity = ""
        complexity_count = (
            len(result.Tasks) + len(result.Foreachtasks) + len(result.Seqtasks) +
            len(result.Forlooptasks) + len(result.Containers) + self.container_count + self.ComponentCount
        )

        if complexity_count <= 5:
            complexity = "Simple"
        elif 5 < complexity_count <= 10:
            complexity = "Medium"
        elif complexity_count > 10:
            complexity = "Complex"
        else:
            complexity = "Simple"

        if self.DataSaveType.upper() == "EXCEL":
            workbook_exists = os.path.exists(analysis_file_path)
            if workbook_exists:
                workbook = load_workbook(analysis_file_path)
            else:
                workbook = Workbook()

            sheet_name = "PackageAnalysisResults"
            if sheet_name not in workbook.sheetnames:
                sheet = workbook.create_sheet(sheet_name)
                sheet.append([
                    "PackageName", "PackagePath", "TasksCount", "ConnectionsCount", "ContainerCount",
                    "ComponentCount", "ExecutionTime", "CreatedDate", "CreatedBy", "Complexity"
                ])
            else:
                sheet = workbook[sheet_name]

            row = [
                result.PackageName,
                result.PackagePath,
                len(result.Tasks) + len(result.Foreachtasks) + len(result.Seqtasks) + len(result.Forlooptasks),
                len(result.Connections),
                len(result.Containers) + self.container_count,
                self.ComponentCount,
                result.ExecutionTime,
                result.CreatedDate,
                result.CreatedBy,
                complexity
            ]
            sheet.append(row)
            workbook.save(analysis_file_path)

            # Saving Variables to 'PackageVariableParameterDetails'
            workbook_details = load_workbook(details_file_path) if os.path.exists(details_file_path) else Workbook()

            # 1. Variables
            var_sheet_name = "PackageVariableParameterDetails"
            if var_sheet_name not in workbook_details.sheetnames:
                sheet_vars = workbook_details.create_sheet(var_sheet_name)
                sheet_vars.append([
                    "PackageName", "PackagePath", "VariableOrParameterName", "DataType", "Value", "IsParameter"
                ])
            else:
                sheet_vars = workbook_details[var_sheet_name]

            for var in result.Variables:
                sheet_vars.append([
                    result.PackageName, result.PackagePath, var.Name, var.DataType, var.Value, var.IsParameter
                ])

            # 2. Connections
            conn_sheet_name = "PackageConnectionDetails"
            if conn_sheet_name not in workbook_details.sheetnames:
                sheet_conn = workbook_details.create_sheet(conn_sheet_name)
                sheet_conn.append([
                    "PackageName", "PackagePath", "ConnectionName", "ConnectionType",
                    "ConnectionExpressions", "ConnectionString", "ConnectionID", "IsProjectConnection"
                ])
            else:
                sheet_conn = workbook_details[conn_sheet_name]

            for conn in result.Connections:
                sheet_conn.append([
                    result.PackageName, result.PackagePath,
                    conn.ConnectionName, conn.ConnectionType,
                    conn.ConnectionExpressions, conn.ConnectionString,
                    conn.ConnectionID, conn.IsProjectConnection
                ])

            workbook_details.save(details_file_path)

        elif self.DataSaveType.upper() == "SQL":
            conn = pyodbc.connect(self._connection_string)
            cursor = conn.cursor()

            insert_package = """
            INSERT INTO PackageAnalysisResults 
            (PackageName, CreatedDate, TaskCount, ConnectionCount, ExecutionTime, PackageFolder, ContainerCount, DTSXXML, CreatedBy, DataFlowTaskComponentCount)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """
            cursor.execute(insert_package, (
                result.PackageName,
                result.CreatedDate,
                len(result.Tasks) + len(result.Foreachtasks) + len(result.Seqtasks) + len(result.Forlooptasks),
                len(result.Connections),
                result.ExecutionTime,
                result.PackagePath,
                len(result.Containers) + self.container_count,
                result.DTSXXML,
                result.CreatedBy,
                self.ComponentCount
            ))

            for var in result.Variables:
                insert_var = """
                INSERT INTO PackageVariableParameterDetails 
                (PackageName, VariableOrParameterName, DataType, Value, PackagePath, IsParameter)
                VALUES (?, ?, ?, ?, ?, ?)
                """
                cursor.execute(insert_var, (
                    result.PackageName, var.Name, var.DataType, var.Value, result.PackagePath, var.IsParameter
                ))

            for conn in result.Connections:
                insert_conn = """
                INSERT INTO PackageConnectionDetails 
                (PackageName, ConnectionName, ConnectionType, PackagePath, ConnectionExpressions, ConnectionString, ConnectionDTSID, IsProjectConnection)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """
                cursor.execute(insert_conn, (
                    result.PackageName,
                    conn.ConnectionName,
                    conn.ConnectionType,
                    result.PackagePath,
                    conn.ConnectionExpressions,
                    conn.ConnectionString,
                    conn.ConnectionID,
                    conn.IsProjectConnection
                ))

            conn.commit()
            cursor.close()
            conn.close()

    def save_dataflow_metadata(self, result, file_path):
        if self.DataSaveType.upper() == "EXCEL":
            for dataflow in result.DataFlowTaskDetails:
                dataflow_file = os.path.join(
                    file_path, dataflow.PackageName.replace(".dtsx", "_DFM.xlsx")
                )
                workbook_exists = os.path.exists(dataflow_file)

                if workbook_exists:
                    workbook = openpyxl.load_workbook(dataflow_file)
                else:
                    workbook = openpyxl.Workbook()
                    workbook.remove(workbook.active)

                sheet_name = "DataFlowTaskMappingDetails"
                if sheet_name not in workbook.sheetnames:
                    ws = workbook.create_sheet(sheet_name)
                    ws.append([
                        "PackageName", "PackagePath", "TaskName", "ColumnName", "ColumnType", "DataType",
                        "ComponentName", "DataConversion", "ComponentPropertyDetails",
                        "ColumnPropertyDetails", "isEventHandler"
                    ])
                else:
                    ws = workbook[sheet_name]

                # Check for existing record
                exists = False
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if (
                        row[0] == dataflow.PackageName and
                        row[1] == dataflow.PackagePath and
                        row[2] == dataflow.TaskName and
                        row[3] == dataflow.ColumnName and
                        row[4] == dataflow.ColumnType and
                        row[5] == dataflow.DataType and
                        row[6] == dataflow.componentName and
                        row[7] == dataflow.DataConversion and
                        row[8] == dataflow.componentPropertyDetails and
                        row[9] == dataflow.ColumnPropertyDetails and
                        row[10] == dataflow.isEventHandler
                    ):
                    exists = True
                    break

                if not exists:
                    ws.append([
                        dataflow.PackageName,
                        dataflow.PackagePath,
                        dataflow.TaskName,
                        dataflow.ColumnName,
                        dataflow.ColumnType,
                        dataflow.DataType,
                        dataflow.componentName,
                        dataflow.DataConversion,
                        dataflow.componentPropertyDetails,
                        dataflow.ColumnPropertyDetails,
                        dataflow.isEventHandler
                    ])
                    workbook.save(dataflow_file)

        elif self.DataSaveType.upper() == "SQL":
            conn = pyodbc.connect(self._connection_string)
            cursor = conn.cursor()

            for dataflow in result.DataFlowTaskDetails:
                insert_query = """
                    INSERT INTO DataFlowTaskMappingDetails (
                        PackageName, TaskName, ColumnName, DataType, ComponentName, DataConversion, PackagePath, 
                        ColumnType, isEventHandler, ComponentPropertyDetails, ColumnPropertyDetails
                    )
                    SELECT ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
                    WHERE NOT EXISTS (
                        SELECT 1 FROM DataFlowTaskMappingDetails
                        WHERE ISNULL(ColumnName, '') = ISNULL(?, '') AND ISNULL(DataType, '') = ISNULL(?, '')
                        AND ISNULL(PackageName, '') = ISNULL(?, '') AND ISNULL(PackagePath, '') = ISNULL(?, '')
                        AND ISNULL(ColumnType, '') = ISNULL(?, '') AND ISNULL(ComponentName, '') = ISNULL(?, '')
                        AND ISNULL(TaskName, '') = ISNULL(?, '') AND ISNULL(ComponentPropertyDetails, '') = ISNULL(?, '')
                        AND ISNULL(ColumnPropertyDetails, '') = ISNULL(?, '')
                    )
                """
                values = [
                    dataflow.PackageName,
                    dataflow.TaskName,
                    dataflow.ColumnName,
                    dataflow.DataType,
                    dataflow.componentName,
                    dataflow.DataConversion,
                    dataflow.PackagePath,
                    dataflow.ColumnType,
                    dataflow.isEventHandler,
                    dataflow.componentPropertyDetails,
                    dataflow.ColumnPropertyDetails,
                    # WHERE NOT EXISTS values
                    dataflow.ColumnName,
                    dataflow.DataType,
                    dataflow.PackageName,
                    dataflow.PackagePath,
                    dataflow.ColumnType,
                    dataflow.componentName,
                    dataflow.TaskName,
                    dataflow.componentPropertyDetails,
                    dataflow.ColumnPropertyDetails
                ]
                cursor.execute(insert_query, values)

            conn.commit()
            cursor.close()
            conn.close()

    def save_precedence_constraint_metadata(self, result, file_path):
        if self.DataSaveType.upper() == "EXCEL":
            workbook_exists = os.path.exists(file_path)
            for precedence in result.PrecedenceConstraintDetails:
                if workbook_exists:
                    workbook = openpyxl.load_workbook(file_path)
                else:
                    workbook = openpyxl.Workbook()
                    workbook.remove(workbook.active)

                sheet_name = "PrecedenceConstraintDetails"
                if sheet_name not in workbook.sheetnames:
                    ws = workbook.create_sheet(sheet_name)
                    ws.append([
                        "PackageName", "PackagePath", "PrecedenceConstraintFrom", "PrecedenceConstraintTo",
                        "PrecedenceConstraintValue", "PrecedenceConstraintExpression",
                        "PrecedenceConstraintLogicalAnd", "PrecedenceConstraintEvalOP", "ContainerName"
                    ])
                else:
                    ws = workbook[sheet_name]

                # Check if record exists
                exists = False
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if (
                        row[0] == precedence.PackageName and
                        row[1] == precedence.PackagePath and
                        row[2] == precedence.PrecedenceConstraintFrom and
                        row[3] == precedence.PrecedenceConstraintTo and
                        row[4] == precedence.PrecedenceConstraintValue and
                        row[5] == precedence.PrecedenceConstraintExpression and
                        row[6] == precedence.PrecedenceConstraintLogicalAnd and
                        row[7] == precedence.PrecedenceConstraintEvalOP and
                        row[8] == precedence.ContainerName
                    ):
                        exists = True
                        break

                if not exists:
                    ws.append([
                        precedence.PackageName, precedence.PackagePath,
                        precedence.PrecedenceConstraintFrom, precedence.PrecedenceConstraintTo,
                        precedence.PrecedenceConstraintValue, precedence.PrecedenceConstraintExpression,
                        precedence.PrecedenceConstraintLogicalAnd, precedence.PrecedenceConstraintEvalOP,
                        precedence.ContainerName
                    ])
                    workbook.save(file_path)

        elif self.DataSaveType.upper() == "SQL":
            conn = pyodbc.connect(self._connection_string)
            cursor = conn.cursor()
            for precedence in result.PrecedenceConstraintDetails:
                insert_query = """
                    INSERT INTO PrecedenceConstraintDetails (
                        PackageName, PrecedenceConstraintFrom, PrecedenceConstraintTo, 
                        PrecedenceConstraintValue, PrecedenceConstraintExpression, PrecedenceConstraintLogicalAnd,
                        PrecedenceConstraintEvalOP, ContainerName, PackagePath
                    )
                    SELECT ?, ?, ?, ?, ?, ?, ?, ?, ?
                    WHERE NOT EXISTS (
                        SELECT 1 FROM PrecedenceConstraintDetails
                        WHERE PackageName = ? AND PrecedenceConstraintFrom = ? AND PrecedenceConstraintTo = ?
                        AND PrecedenceConstraintValue = ? AND PrecedenceConstraintExpression = ?
                        AND PrecedenceConstraintLogicalAnd = ? AND PrecedenceConstraintEvalOP = ?
                        AND ContainerName = ? AND PackagePath = ?
                    )
                """
                values = [
                    precedence.PackageName, precedence.PrecedenceConstraintFrom, precedence.PrecedenceConstraintTo,
                    precedence.PrecedenceConstraintValue, precedence.PrecedenceConstraintExpression,
                    precedence.PrecedenceConstraintLogicalAnd, precedence.PrecedenceConstraintEvalOP,
                    precedence.ContainerName, precedence.PackagePath,
                    # For WHERE NOT EXISTS part
                    precedence.PackageName, precedence.PrecedenceConstraintFrom, precedence.PrecedenceConstraintTo,
                    precedence.PrecedenceConstraintValue, precedence.PrecedenceConstraintExpression,
                    precedence.PrecedenceConstraintLogicalAnd, precedence.PrecedenceConstraintEvalOP,
                    precedence.ContainerName, precedence.PackagePath
                ]
                cursor.execute(insert_query, values)
            conn.commit()
            cursor.close()
            conn.close()

    def save_event_metadata(self, result, file_path):
        if self.DataSaveType.upper() == "EXCEL":
            workbook_exists = os.path.exists(file_path)
            for task in result.ExtractTaskDetails:
                if not task.TaskName:
                    continue

                if workbook_exists:
                    workbook = load_workbook(file_path)
                else:
                    workbook = openpyxl.Workbook()
                    workbook.remove(workbook.active)

                sheet_name = "EventHandlerTaskDetails"
                if sheet_name not in workbook.sheetnames:
                    ws = workbook.create_sheet(sheet_name)
                    ws.append([
                        "PackageName", "PackagePath", "EventHandlerName", "EventHandlerType", "EventType",
                        "TaskName", "TaskType", "ContainerName", "ContainerType", "ContainerExpression",
                        "TaskConnectionName", "SqlQuery", "Variables", "Parameters", "Expressions",
                        "DataFlowDaskSourceName", "DataFlowTaskSourceType", "DataFlowTaskTargetName", "DataFlowTaskTargetType",
                        "DataFlowTaskTargetTable", "DataFlowDaskSourceConnectionName", "DataFlowDaskTargetConnectionName",
                        "SendMailTaskDetails", "ResultSetDetails", "TaskComponentDetails"
                    ])
                ws = workbook[sheet_name]
                ws.append([
                    task.PackageName, task.PackagePath, task.EventHandlerName, task.EventHandlerType, task.EventType,
                    task.TaskName, task.TaskType, task.ContainerName, task.ContainerType, task.ContainerExpression,
                    task.ConnectionName, task.TaskSqlQuery, task.Variables, task.Parameters, task.Expressions,
                    task.SourceComponent, task.SourceType, task.TargetComponent, task.TargetType,
                    task.TargetTable, task.SourceConnectionName, task.TargetConnectionName,
                    task.SendMailTask, task.ResultSetDetails, task.TaskComponentDetails
                ])
                workbook.save(file_path)

        elif self.DataSaveType.upper() == "SQL":
            conn = pyodbc.connect(self._connection_string)
            cursor = conn.cursor()
            for task in result.ExtractTaskDetails:
                if not task.TaskName:
                    continue
                insert_query = """
                    INSERT INTO EventTaskDetails (
                        PackageName, TaskName, TaskType, SqlQuery, ContainerName, PackagePath,
                        Variables, Parameters, Expressions, DataFlowDaskSourceName,
                        DataFlowTaskSourceType, DataFlowTaskTargetName, DataFlowTaskTargetType,
                        DataFlowTaskTargetTable, SendMailTaskDetails, ResultSetDetails,
                        ContainerType, ContainerExpression, EventHandlerName, EventHandlerType,
                        EventType, DataFlowDaskSourceConnectionName, DataFlowDaskTargetConnectionName,
                        TaskConnectionName, TaskComponentDetails
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """
                cursor.execute(insert_query, (
                    task.PackageName, task.TaskName, task.TaskType, task.TaskSqlQuery, task.ContainerName,
                    task.PackagePath, task.Variables, task.Parameters, task.Expressions, task.SourceComponent,
                    task.SourceType, task.TargetComponent, task.TargetType, task.TargetTable, task.SendMailTask,
                    task.ResultSetDetails, task.ContainerType, task.ContainerExpression, task.EventHandlerName,
                    task.EventHandlerType, task.EventType, task.SourceConnectionName, task.TargetConnectionName,
                    task.ConnectionName, task.TaskComponentDetails
                ))
            conn.commit()
            cursor.close()
            conn.close()

    def save_package_task_metadata(self, result, file_path):
        if self.DataSaveType.upper() == "EXCEL":
            tasks = result.get("ExtractTaskDetails", [])
            workbook_exists = os.path.exists(file_path)

            for task in tasks:
                if task.get("TaskName"):
                    wb = load_workbook(file_path) if workbook_exists else Workbook()
                    if not wb.sheetnames:
                        wb.remove(wb.active)

                    sheet_name = "PackageTaskDetails"
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                    else:
                        ws = wb.create_sheet(sheet_name)
                        ws.append([
                            "PackageName", "PackagePath", "TaskName", "TaskType", "ContainerName",
                            "TaskConnectionName", "SqlQuery", "Variables", "Parameters", "Expressions",
                            "DataFlowDaskSourceName", "DataFlowTaskSourceType", "DataFlowTaskTargetName",
                            "DataFlowTaskTargetType", "DataFlowTaskTargetTable", "DataFlowDaskSourceConnectionName",
                            "DataFlowDaskTargetConnectionName", "ResultSetDetails", "TaskComponentDetails"
                        ])

                    rows = list(ws.iter_rows(min_row=2, values_only=True))
                    record_exists = any(
                        row[0] == task.get("PackageName") and
                        row[1] == task.get("PackagePath") and
                        row[2] == task.get("TaskName") and
                        row[3] == task.get("TaskType") and
                        row[4] == task.get("ContainerName")
                        for row in rows
                    )

                    if not record_exists:
                        ws.append([
                            task.get("PackageName"), task.get("PackagePath"), task.get("TaskName"),
                            task.get("TaskType"), task.get("ContainerName"), task.get("ConnectionName"),
                            task.get("TaskSqlQuery"), task.get("Variables"), task.get("Parameters"),
                            task.get("Expressions"), task.get("SourceComponent"), task.get("SourceType"),
                            task.get("TargetComponent"), task.get("TargetType"), task.get("TargetTable"),
                            task.get("SourceConnectionName"), task.get("TargetConnectionName"),
                            task.get("ResultSetDetails"), task.get("TaskComponentDetails")
                        ])
                        wb.save(file_path)

                if task.get("ContainerName"):
                    wb = load_workbook(file_path) if os.path.exists(file_path) else Workbook()
                    if not wb.sheetnames:
                        wb.remove(wb.active)
                    sheet_name = "PackageContainerDetails"
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                    else:
                        ws = wb.create_sheet(sheet_name)
                        ws.append([
                            "PackageName", "PackagePath", "ContainerName",
                            "ContainerType", "ContainerExpressions", "ContainerEnumerator"
                        ])

                    rows = list(ws.iter_rows(min_row=2, values_only=True))
                    record_exists = any(
                        row[0] == task.get("PackageName") and
                        row[1] == task.get("PackagePath") and
                        row[2] == task.get("ContainerName") and
                        row[3] == task.get("ContainerType") and
                        row[4] == task.get("ContainerExpression") and
                        row[5] == task.get("ContainerEnum")
                        for row in rows
                    )

                    if not record_exists:
                        ws.append([
                            task.get("PackageName"), task.get("PackagePath"), task.get("ContainerName"),
                            task.get("ContainerType"), task.get("ContainerExpression"), task.get("ContainerEnum")
                        ])
                        wb.save(file_path)

        elif self.DataSaveType.upper() == "SQL":
            conn = pyodbc.connect(self._connection_string)
            cursor = conn.cursor()

            for task in result.get("ExtractTaskDetails", []):
                if task.get("TaskName"):
                    task_query = """
                        INSERT INTO PackageTaskDetails (
                            PackageName, TaskName, TaskType, SqlQuery, ContainerName, PackagePath,
                            Variables, Parameters, Expressions, DataFlowDaskSourceName, DataFlowTaskSourceType,
                            DataFlowTaskTargetName, DataFlowTaskTargetType, DataFlowTaskTargetTable,
                            ResultSetDetails, DataFlowDaskSourceConnectionName,
                            DataFlowDaskTargetConnectionName, TaskConnectionName, TaskComponentDetails
                        )
                        SELECT ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
                        WHERE NOT EXISTS (
                            SELECT 1 FROM PackageTaskDetails
                            WHERE ContainerName = ? AND PackageName = ? AND PackagePath = ? AND TaskName = ?
                        )
                    """
                    values = [
                        task.get("PackageName"), task.get("TaskName"), task.get("TaskType"),
                        task.get("TaskSqlQuery"), task.get("ContainerName"), task.get("PackagePath"),
                        task.get("Variables"), task.get("Parameters"), task.get("Expressions"),
                        task.get("SourceComponent"), task.get("SourceType"), task.get("TargetComponent"),
                        task.get("TargetType"), task.get("TargetTable"), task.get("ResultSetDetails"),
                        task.get("SourceConnectionName"), task.get("TargetConnectionName"),
                        task.get("ConnectionName"), task.get("TaskComponentDetails"),
                        task.get("ContainerName"), task.get("PackageName"), task.get("PackagePath"), task.get("TaskName")
                    ]
                    cursor.execute(task_query, values)

                if task.get("ContainerName"):
                    container_query = """
                        INSERT INTO PackageContainerDetails (
                            PackageName, ContainerName, ContainerType,
                            ContainerExpressions, ContainerEnumerator, PackagePath
                        )
                        SELECT ?, ?, ?, ?, ?, ?
                        WHERE NOT EXISTS (
                            SELECT 1 FROM PackageContainerDetails
                            WHERE ContainerName = ? AND ContainerType = ? AND PackageName = ?
                            AND PackagePath = ? AND ContainerExpressions = ISNULL(?, '')
                            AND ContainerEnumerator = ISNULL(?, '')
                        )
                    """
                    values = [
                        task.get("PackageName"), task.get("ContainerName"), task.get("ContainerType"),
                        task.get("ContainerExpression"), task.get("ContainerEnum"), task.get("PackagePath"),
                        task.get("ContainerName"), task.get("ContainerType"), task.get("PackageName"),
                        task.get("PackagePath"), task.get("ContainerExpression"), task.get("ContainerEnum")
                    ]
                    cursor.execute(container_query, values)

            conn.commit()
            cursor.close()
            conn.close()
            print("Saved Package Task and Container metadata to SQL Server.")

    def save_project_parameter_metadata(self, result, file_path):
        if self.DataSaveType.upper() == "EXCEL":
            for param_info in result.get("ProjectParameterDetails", []):
                if not result.get("PackageName"):
                    continue

                if os.path.exists(file_path):
                    wb = load_workbook(file_path)
                else:
                    wb = Workbook()
                    wb.remove(wb.active)

                if "ProjectParameterDetails" in wb.sheetnames:
                    ws = wb["ProjectParameterDetails"]
                else:
                    ws = wb.create_sheet("ProjectParameterDetails")
                    ws.append(["ProjectPath", "ParameterName", "ParameterValue", "ParameterDataType"])

                row = [
                    result.get("PackagePath", ""),
                    param_info.get("ParameterName", ""),
                    param_info.get("Value", ""),
                    param_info.get("DataType", "")
                ]
                ws.append(row)
                wb.save(file_path)

            print(f"Saved project parameters to Excel: {file_path}")

        elif self.DataSaveType.upper() == "SQL":
            conn = pyodbc.connect(self._connection_string)
            cursor = conn.cursor()

            for param_info in result.get("ProjectParameterDetails", []):
                if not result.get("PackageName"):
                    continue

                cursor.execute("""
                    INSERT INTO ProjectParameterDetails (
                        ParameterName, ParameterValue, ParameterDataType, ProjectPath
                    ) VALUES (?, ?, ?, ?)
                """, (
                    param_info.get("ParameterName", ""),
                    param_info.get("Value", ""),
                    param_info.get("DataType", ""),
                    result.get("PackagePath", "")
                ))

            conn.commit()
            cursor.close()
            conn.close()
            print("Saved project parameters to SQL Server.")

    def save_connections_metadata(self, result, file_path):
        if self.DataSaveType.upper() == "EXCEL":
            for conn in result.get("Connections", []):
                if os.path.exists(file_path):
                    wb = load_workbook(file_path)
                else:
                    wb = Workbook()
                    wb.remove(wb.active)

                if "PackageConnectionDetails" in wb.sheetnames:
                    ws = wb["PackageConnectionDetails"]
                else:
                    ws = wb.create_sheet("PackageConnectionDetails")
                    ws.append([
                        "PackageName", "PackagePath", "ConnectionName", "ConnectionType",
                        "ConnectionExpressions", "ConnectionString", "ConnectionID", "IsProjectConnection"
                    ])

                row = [
                    result.get("PackageName", ""),
                    result.get("PackagePath", ""),
                    conn.get("ConnectionName", ""),
                    conn.get("ConnectionType", ""),
                    conn.get("ConnectionExpressions", ""),
                    conn.get("ConnectionString", ""),
                    conn.get("ConnectionID", ""),
                    conn.get("IsProjectConnection", "")
                ]
                ws.append(row)
                wb.save(file_path)

            print(f"Saved connection metadata to Excel: {file_path}")

        elif self.DataSaveType.upper() == "SQL":
            conn_db = pyodbc.connect(self._connection_string)
            cursor = conn_db.cursor()

            for conn_info in result.get("Connections", []):
                cursor.execute("""
                    INSERT INTO PackageConnectionDetails (
                        PackageName, ConnectionName, ConnectionType, PackagePath, 
                        ConnectionExpressions, ConnectionString, ConnectionDTSID, IsProjectConnection
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    result.get("PackageName", ""),
                    conn_info.get("ConnectionName", ""),
                    conn_info.get("ConnectionType", ""),
                    result.get("PackagePath", ""),
                    conn_info.get("ConnectionExpressions", ""),
                    conn_info.get("ConnectionString", ""),
                    conn_info.get("ConnectionID", ""),
                    conn_info.get("IsProjectConnection", "")
                ))

            conn_db.commit()
            cursor.close()
            conn_db.close()
            print("Saved connection metadata to SQL Server.")

    def log_error(self, file_path, exception):
        print(f"Error processing {file_path}: {str(exception)}")
        traceback.print_exc()

    def save_update_connection_name(self, file_path):
        if self.DataSaveType.upper() == "EXCEL":
            wb = load_workbook(file_path)

            sheet1 = wb["PackageTaskDetails"]
            sheet2 = wb["PackageConnectionDetails"]
            sheet3 = wb["EventHandlerTaskDetails"]

            for row2 in range(2, sheet2.max_row + 1):
                conn_pkg_path = sheet2.cell(row=row2, column=2).value
                conn_name = sheet2.cell(row=row2, column=3).value
                conn_dtsid = sheet2.cell(row=row2, column=7).value

                for row1 in range(2, sheet1.max_row + 1):
                    task_pkg_path = sheet1.cell(row=row1, column=2).value
                    if conn_pkg_path == task_pkg_path:
                        if conn_dtsid == sheet1.cell(row=row1, column=6).value:
                            sheet1.cell(row=row1, column=6).value = conn_name
                        elif conn_dtsid == sheet1.cell(row=row1, column=16).value:
                            sheet1.cell(row=row1, column=16).value = conn_name
                        elif conn_dtsid == sheet1.cell(row=row1, column=17).value:
                            sheet1.cell(row=row1, column=17).value = conn_name

                for row3 in range(2, sheet3.max_row + 1):
                    handler_pkg_path = sheet3.cell(row=row3, column=2).value
                    if conn_pkg_path == handler_pkg_path:
                        if conn_dtsid == sheet3.cell(row=row3, column=11).value:
                            sheet3.cell(row=row3, column=11).value = conn_name
                        elif conn_dtsid == sheet3.cell(row=row3, column=21).value:
                            sheet3.cell(row=row3, column=21).value = conn_name
                        elif conn_dtsid == sheet3.cell(row=row3, column=22).value:
                            sheet3.cell(row=row3, column=22).value = conn_name

            wb.save(file_path)
            print("Connection names updated in Excel.")

        elif self.DataSaveType.upper() == "SQL":
            conn = pyodbc.connect(self._connection_string)
            cursor = conn.cursor()

            connection_query = """
            UPDATE task SET task.TaskConnectionName = conn.ConnectionName
            FROM PackageTaskDetails task
            INNER JOIN PackageConnectionDetails conn WITH (NOLOCK)
                ON conn.PackagePath = task.PackagePath
                AND task.TaskConnectionName = conn.ConnectionDTSID
            WHERE ISNULL(task.TaskConnectionName, '') <> '';

            UPDATE task SET 
                task.DataFlowDaskSourceConnectionName = sconn.ConnectionName, 
                task.DataFlowDaskTargetConnectionName = tconn.ConnectionName
            FROM PackageTaskDetails task
            INNER JOIN PackageConnectionDetails sconn WITH (NOLOCK)
                ON sconn.PackagePath = task.PackagePath
                AND task.DataFlowDaskSourceConnectionName = sconn.ConnectionDTSID
            INNER JOIN PackageConnectionDetails tconn WITH (NOLOCK)
                ON tconn.PackagePath = task.PackagePath
                AND task.DataFlowDaskTargetConnectionName = tconn.ConnectionDTSID
            WHERE ISNULL(task.DataFlowDaskSourceConnectionName, '') <> '';

            UPDATE task SET task.TaskConnectionName = conn.ConnectionName
            FROM EventTaskDetails task
            INNER JOIN PackageConnectionDetails conn WITH (NOLOCK)
                ON conn.PackagePath = task.PackagePath
                AND task.TaskConnectionName = conn.ConnectionDTSID
            WHERE ISNULL(task.TaskConnectionName, '') <> '';

            UPDATE task SET 
                task.DataFlowDaskSourceConnectionName = sconn.ConnectionName,
                task.DataFlowDaskTargetConnectionName = tconn.ConnectionName
            FROM EventTaskDetails task
            INNER JOIN PackageConnectionDetails sconn WITH (NOLOCK)
                ON sconn.PackagePath = task.PackagePath
                AND task.DataFlowDaskSourceConnectionName = sconn.ConnectionDTSID
            INNER JOIN PackageConnectionDetails tconn WITH (NOLOCK)
                ON tconn.PackagePath = task.PackagePath
                AND task.DataFlowDaskTargetConnectionName = tconn.ConnectionDTSID
            WHERE ISNULL(task.DataFlowDaskSourceConnectionName, '') <> '';

            -- Reset precedence constraints
            UPDATE task SET
                task.ONSuccessPrecedenceConstrainttoTask = '',
                task.ONSuccessPrecedenceConstraintExpression = '',
                task.ONSuccessPrecedenceConstraintEvalOP = '',
                task.ONSuccessPrecedenceConstraintLogicalAnd = '',
                task.ONFailurePrecedenceConstrainttoTask = '',
                task.ONFailurePrecedenceConstraintExpression = '',
                task.ONFailurePrecedenceConstraintEvalOP = '',
                task.ONFailurePrecedenceConstraintLogicalAnd = '',
                task.ONCompletionPrecedenceConstrainttoTask = '',
                task.ONCompletionPrecedenceConstraintExpression = '',
                task.ONCompletionPrecedenceConstraintEvalOP = '',
                task.ONCompletionPrecedenceConstraintLogicalAnd = ''
            FROM PackageTaskDetails task;

            -- Apply precedence constraints (Success, Failure, Completion)
            UPDATE task SET
                task.ONSuccessPrecedenceConstrainttoTask = pcd.PrecedenceConstraintto,
                task.ONSuccessPrecedenceConstraintExpression = pcd.PrecedenceConstraintExpression,
                task.ONSuccessPrecedenceConstraintEvalOP = pcd.PrecedenceConstraintEvalOP,
                task.ONSuccessPrecedenceConstraintLogicalAnd = pcd.PrecedenceConstraintLogicalAnd
            FROM PackageTaskDetails task
            INNER JOIN PrecedenceConstraintDetails pcd WITH (NOLOCK)
                ON pcd.PrecedenceConstraintFrom = task.TaskName
                AND pcd.PackageName = task.PackageName
                AND pcd.PackagePath = task.PackagePath
                AND ISNULL(pcd.ContainerName, '') = ISNULL(task.ContainerName, '')
            WHERE pcd.PrecedenceConstraintValue = 'Success';

            UPDATE task SET
                task.ONFailurePrecedenceConstrainttoTask = pcd.PrecedenceConstraintto,
                task.ONFailurePrecedenceConstraintExpression = pcd.PrecedenceConstraintExpression,
                task.ONFailurePrecedenceConstraintEvalOP = pcd.PrecedenceConstraintEvalOP,
                task.ONFailurePrecedenceConstraintLogicalAnd = pcd.PrecedenceConstraintLogicalAnd
            FROM PackageTaskDetails task
            INNER JOIN PrecedenceConstraintDetails pcd WITH (NOLOCK)
                ON pcd.PrecedenceConstraintFrom = task.TaskName
                AND pcd.PackageName = task.PackageName
                AND pcd.PackagePath = task.PackagePath
                AND ISNULL(pcd.ContainerName, '') = ISNULL(task.ContainerName, '')
            WHERE pcd.PrecedenceConstraintValue = 'Failure';

            UPDATE task SET
                task.ONCompletionPrecedenceConstrainttoTask = pcd.PrecedenceConstraintto,
                task.ONCompletionPrecedenceConstraintExpression = pcd.PrecedenceConstraintExpression,
                task.ONCompletionPrecedenceConstraintEvalOP = pcd.PrecedenceConstraintEvalOP,
                task.ONCompletionPrecedenceConstraintLogicalAnd = pcd.PrecedenceConstraintLogicalAnd
            FROM PackageTaskDetails task
            INNER JOIN PrecedenceConstraintDetails pcd WITH (NOLOCK)
                ON pcd.PrecedenceConstraintFrom = task.TaskName
                AND pcd.PackageName = task.PackageName
                AND pcd.PackagePath = task.PackagePath
                AND ISNULL(pcd.ContainerName, '') = ISNULL(task.ContainerName, '')
            WHERE pcd.PrecedenceConstraintValue = 'Completion';

            -- Set Complexity Classification
            UPDATE PA SET PA.Complexcity = CASE 
                WHEN Final.TaskCount + Final.ContainerCount + Final.ComponentCount < 5 THEN 'Simple'
                WHEN Final.TaskCount + Final.ContainerCount + Final.ComponentCount BETWEEN 5 AND 10 THEN 'Medium'
                WHEN Final.TaskCount + Final.ContainerCount + Final.ComponentCount > 10 THEN 'Complex'
                ELSE 'Simple'
            END
            FROM PackageAnalysisResults PA
            LEFT JOIN (
                SELECT PackageName, PackagePath,
                       SUM(TaskCount) TaskCount,
                       SUM(ContainerCount) ContainerCount,
                       SUM(ComponentCount) ComponentCount
                FROM (
                    SELECT PT.PackageName, PT.PackagePath,
                           SUM(CASE WHEN TaskType <> 'ExecutePackageTask' THEN 1 ELSE 0 END) AS TaskCount,
                           0 AS ContainerCount,
                           0 AS ComponentCount
                    FROM PackageTaskDetails PT
                    GROUP BY PT.PackageName, PT.PackagePath

                    UNION ALL

                    SELECT PT.PackageName, PT.PackagePath,
                           1 AS TaskCount, 0, 0
                    FROM PackageTaskDetails PT
                    WHERE TaskType = 'ExecutePackageTask'

                    UNION ALL

                    SELECT PC.PackageName, PC.PackagePath,
                           0, 1, 0
                    FROM PackageContainerDetails PC
                    WHERE PC.ContainerType = 'Sequence'

                    UNION ALL

                    SELECT PC.PackageName, PC.PackagePath,
                           0, COUNT(1), 0
                    FROM PackageContainerDetails PC
                    WHERE PC.ContainerType <> 'Sequence'
                    GROUP BY PackageName, PackagePath

                    UNION ALL

                    SELECT PC.PackageName, PC.PackagePath,
                           0, 0, COUNT(DISTINCT ComponentName)
                    FROM DataFlowTaskMappingDetails PC
                    GROUP BY PackageName, PackagePath
                ) A
                GROUP BY PackageName, PackagePath
            ) Final ON Final.PackageName = PA.PackageName
            AND Final.PackagePath = PA.PackageFolder;
            """

            cursor.execute(connection_query)
            conn.commit()
            cursor.close()
            conn.close()
            print("Connection names updated in SQL.")

class Program:
    @staticmethod
    def delete_all_files_in_directory(directory_path):
        try:
            if os.path.exists(directory_path) and os.path.isdir(directory_path):
                for file_name in os.listdir(directory_path):
                    file_path = os.path.join(directory_path, file_name)
                    if os.path.isfile(file_path):
                        os.remove(file_path)
                        # print(f"Deleted: {file_path}")
            # print("All files have been deleted.")
            else:
                print("Directory does not exist.")
        except Exception as ex:
            print(f"An error occurred: {ex}")

def main():
    connection_string = ""
    output_folder = ""
    package_folder = input("Enter the Package Folder path:\n")
    data_save_type = input("Enter the Data Save Type (SQL or EXCEL):\n")

    if data_save_type.upper() == "SQL":
        connection_string = input("Enter the Connection String:\n")
    elif data_save_type.upper() == "EXCEL":
        output_folder = input("Enter the Output Folder path:\n")
    else:
        print("Wrong Input")
        threading.Event().wait(5)
        return

    package_analysis_file_path = os.path.join(output_folder, "PackageAnalysisResult.xlsx")
    dataflow_file_path = output_folder
    package_details_file_path = os.path.join(output_folder, "PackageDetails.xlsx")

    if data_save_type.upper() == "EXCEL":
        Program.delete_all_files_in_directory(dataflow_file_path)

    analyzer = SSISPackageAnalyzer(
        package_folder,
        connection_string,
        package_analysis_file_path,
        dataflow_file_path,
        package_details_file_path,
        data_save_type
    )
    analyzer.analyze_all_packages()
    print("Running...")


if __name__ == "__main__":
    main()
