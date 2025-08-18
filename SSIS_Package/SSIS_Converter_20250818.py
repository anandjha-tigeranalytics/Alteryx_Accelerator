import os
import re
import time
import openpyxl
import pyodbc
import threading
import traceback
import xml.etree.ElementTree as ET
import xml.dom.minidom as minidom
import pandas as pd
from openpyxl import load_workbook,Workbook
from xml.dom import minidom
from dataclasses import dataclass, field
from typing import List, Optional
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Border, Side

# auto adjusting excel columns
def auto_adjust_excel_columns(ws):
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[col_letter].width = adjusted_width

# formatting excel file
def format_excel_file(filename):

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    wb = load_workbook(filename)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Make header bold
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Adjust column widths
        auto_adjust_excel_columns(ws)

        # Apply thin border
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border

    wb.save(filename)
    # print(f"Formatted Excel file: {filename}")


@dataclass
class TaskInfo:
    PackageName: Optional[str] = None
    PackagePath: Optional[str] = None
    EventHandlerName: Optional[str] = None
    EventHandlerType: Optional[str] = None
    EventType: Optional[str] = None
    TaskName: Optional[str] = None
    TaskType: Optional[str] = None
    TaskSqlQuery: Optional[str] = None
    ContainerName: Optional[str] = None
    ContainerType: Optional[str] = None
    ContainerExpression: Optional[str] = None
    ContainerEnum: Optional[str] = None
    Variables: Optional[str] = None
    Parameters: Optional[str] = None
    Expressions: Optional[str] = None
    ExecuteProcessDetails: Optional[str] = None
    FileSystemSourcePath: Optional[str] = None
    FileSystemDestinationPath: Optional[str] = None
    SourceComponent: Optional[str] = None
    TargetComponent: Optional[str] = None
    SourceType: Optional[str] = None
    TargetType: Optional[str] = None
    TargetTable: Optional[str] = None
    SendMailTask: Optional[str] = None
    ScriptTask: Optional[str] = None
    FTPTask: Optional[str] = None
    ExecutePackage: Optional[str] = None
    ResultSetDetails: Optional[str] = None
    SeqTaskName: Optional[str] = None
    ForeachTaskName: Optional[str] = None
    ForloopTaskName: Optional[str] = None
    ConnectionName: Optional[str] = None
    SourceConnectionName: Optional[str] = None
    TargetConnectionName: Optional[str] = None
    TaskComponentDetails: Optional[str] = None

@dataclass
class ConnectionInfo:
    ConnectionName: Optional[str] = None
    ConnectionType: Optional[str] = None
    ConnectionString: Optional[str] = None
    ConnectionExpressions: Optional[str] = None
    ConnectionID: Optional[str] = None
    IsProjectConnection: Optional[str] = None

@dataclass
class ContainerInfo:
    ContainerName: Optional[str] = None
    ContainerType: Optional[str] = None
    ContainerExpression: Optional[str] = None

@dataclass
class VariableInfo:
    Name: Optional[str] = None
    Value: Optional[str] = None
    DataType: Optional[str] = None
    Namespace: Optional[str] = None
    IsParameter: int = 0

@dataclass
class TaskParameterInfo:
    ParameterName: Optional[str] = None
    ParameterType: Optional[str] = None
    DataType: Optional[str] = None
    Value: Optional[str] = None
    DtsVariableName: Optional[str] = None

@dataclass
class DataFlowTaskInfo:
    ColumnName: Optional[str] = None
    ColumnType: Optional[str] = None
    DataType: Optional[str] = None
    TargetColumn: Optional[str] = None
    componentName: Optional[str] = None
    DataConversion: Optional[str] = None
    PackageName: Optional[str] = None
    PackagePath: Optional[str] = None
    TaskName: Optional[str] = None
    isEventHandler: Optional[str] = None
    componentPropertyDetails: Optional[str] = None
    ColumnPropertyDetails: Optional[str] = None

@dataclass
class PrecedenceConstraintInfo:
    PrecedenceConstraintFrom: Optional[str] = None
    PrecedenceConstraintTo: Optional[str] = None
    PrecedenceConstraintValue: Optional[str] = None
    PrecedenceConstraintLogicalAnd: Optional[str] = None
    PrecedenceConstraintEvalOP: Optional[str] = None
    PrecedenceConstraintExpression: Optional[str] = None
    ContainerName: Optional[str] = None
    PackageName: Optional[str] = None
    PackagePath: Optional[str] = None

@dataclass
class ProjectParameterInfo:
    ParameterName: Optional[str] = None
    DataType: Optional[str] = None
    Value: Optional[str] = None

@dataclass
class PackageAnalysisResult:
    PackageName: Optional[str] = None
    CreatedDate: Optional[datetime] = None
    CreatedBy: Optional[str] = None
    Tasks: List[TaskInfo] = field(default_factory=list)
    Seqtasks: List[TaskInfo] = field(default_factory=list)
    Foreachtasks: List[TaskInfo] = field(default_factory=list)
    Forlooptasks: List[TaskInfo] = field(default_factory=list)
    Connections: List[ConnectionInfo] = field(default_factory=list)
    ExecutionTime: Optional[timedelta] = None
    PackagePath: Optional[str] = None
    Containers: List[ContainerInfo] = field(default_factory=list)
    DTSXXML: Optional[str] = None
    SequenceContainerTaskCount: List[TaskInfo] = field(default_factory=list)
    ForeachContainerTaskCount: List[TaskInfo] = field(default_factory=list)
    ForLoopContainerTaskCount: List[TaskInfo] = field(default_factory=list)
    Variables: List[VariableInfo] = field(default_factory=list)
    DataFlowTaskDetails: List[DataFlowTaskInfo] = field(default_factory=list)
    PrecedenceConstraintDetails: List[PrecedenceConstraintInfo] = field(default_factory=list)
    ExtractTaskDetails: List[TaskInfo] = field(default_factory=list)
    ProjectParameterDetails: List[ProjectParameterInfo] = field(default_factory=list)

class TaskHost:
    def __init__(self, name="Task"):
        self.Name = name
        self.InnerObject = None

class ForEachLoop:
    pass

class Sequence:
    pass

class ForLoop:
    pass

@dataclass
class CustomProperty:
    Name: str
    Value: str
    ExpressionType: Optional[str] = None

@dataclass
class InputColumn:
    Name: str
    DataType: str
    #ObjectType: str
    #CustomPropertyCollection: List[CustomProperty]

@dataclass
class OutputColumn:
    Name: str
    DataType: str
    #ObjectType: str
    #CustomPropertyCollection: List[CustomProperty]

@dataclass
class Input:
    Name: str
    InputColumnCollection: List[InputColumn]
    #CustomPropertyCollection: List[CustomProperty]

@dataclass
class Output:
    Name: str
    OutputColumnCollection: List[OutputColumn]
    #CustomPropertyCollection: List[CustomProperty]

@dataclass
class VariableParameter:
    Name: str
    DataType: str
    Value: str
    IsParameter: int

@dataclass
class Component:
    Name: str
    ComponentClassID: str
    InputCollection: List[Input]
    OutputCollection: List[Output]

class MainPipe:
    def __init__(self):
        self.ComponentMetaDataCollection: List[Component] = []

class SSISPackageAnalyzer:
    def __init__(self, package_folder, metadata_connection_string, package_analysis_file_path, dataflow_file_path, package_details_file_path, data_save_type):
        self.container_count = 0
        self.container_task_count = 0
        self._connection_string = metadata_connection_string
        self._package_folder = package_folder
        self.processed_package_paths = set()
        self.PackagePath = ""
        self.PackageName = ""
        self.ComponentCount = 0
        self.PackageAnalysisFilePath = package_analysis_file_path
        self.DataFlowlFilePath = dataflow_file_path
        self.PackageDetailsFilePath = package_details_file_path
        self.DataSaveType = data_save_type
        self.ComponentNameCheck = set()
        self.variables_metadata = []

    def truncate_table(self):
        if self.DataSaveType.upper() == "SQL":
            try:
                conn = pyodbc.connect(self._connection_string)
                cursor = conn.cursor()
                tables_to_truncate = [
                    "PackageAnalysisResults",
                    "PackageTaskDetails",
                    "PackageConnectionDetails",
                    "PackageContainerDetails",
                    "ProjectParameterDetails",
                    "PackageVariableParameterDetails",
                    "DataFlowTaskMappingDetails",
                    "PrecedenceConstraintDetails",
                    "EventTaskDetails"
                ]

                for table in tables_to_truncate:
                    try:
                        cursor.execute(f"TRUNCATE TABLE {table};")
                    except Exception as inner_ex:
                        print(f"Warning: Could not truncate {table}: {inner_ex}")

                conn.commit()
                cursor.close()
                conn.close()
                print("Truncated all metadata tables.")

            except Exception as e:
                self.log_error("SQL Truncate", e)

    def extract_variables_and_parameters(self, dtsx_file_path: str) -> List[VariableParameter]:
        result = []
        try:
            tree = ET.parse(dtsx_file_path)
            root = tree.getroot()
            namespace = {'DTS': 'www.microsoft.com/SqlServer/Dts'}

            # Extract <DTS:Variable> with Namespace="User" only
            for variable in root.findall('.//DTS:Variables/DTS:Variable', namespace):
                ns_attr = variable.attrib.get('{www.microsoft.com/SqlServer/Dts}Namespace', '')
                if ns_attr != "User":
                    continue  # Skip system variables

                name = variable.attrib.get('{www.microsoft.com/SqlServer/Dts}ObjectName', '')
                value_elem = variable.find('DTS:VariableValue', namespace)
                value = value_elem.text.strip() if value_elem is not None and value_elem.text else ''
                result.append(VariableParameter(Name=name, DataType="String", Value=value, IsParameter=0))

            # Extract <DTS:PackageParameter>
            for param in root.findall('.//DTS:PackageParameters/DTS:PackageParameter', namespace):
                name = param.attrib.get('{www.microsoft.com/SqlServer/Dts}ObjectName', '')
                prop_elem = param.find('DTS:Property', namespace)
                value = prop_elem.text.strip() if prop_elem is not None and prop_elem.text else ''
                result.append(VariableParameter(Name=name, DataType="String", Value=value, IsParameter=1))

        except Exception as e:
            print(f"Error extracting variables/parameters from {dtsx_file_path}: {e}")
        return result

 
    def analyze_all_packages(self, package_folder):
        self.truncate_table()  # Now correctly using self

        for root, dirs, files in os.walk(package_folder):
            if "\\obj\\" in root.lower():
                continue  # Skip obj folders

            try:
                package_files = [os.path.join(root, f) for f in files if f.endswith(".dtsx")]
                connection_manager_files = [os.path.join(root, f) for f in files if f.endswith(".conmgr")]
                param_files = [os.path.join(root, f) for f in files if f.endswith(".params")]

                for package_path in package_files:
                    if package_path in self.processed_package_paths:
                        continue
                    try:
                        self.processed_package_paths.add(package_path)
                        self.analyze_single_package(package_path)
                        self.generate_dataflow_mapping_excel(package_path, self.DataFlowlFilePath)

                        #  NEW: Extract Package Task Details from .dtsx
                        result = self.extract_package_task_details(package_path)
                        self.save_package_task_metadata({"ExtractTaskDetails": result}, self.PackageDetailsFilePath)

                        # Extract Container Details and merge with result
                        container_results = self.extract_package_container_details(package_path)
                        self.save_package_container_metadata(container_results, self.PackageDetailsFilePath)


                        # Extract Event Handler Task Details
                        event_details = self.extract_event_handler_task_details(package_path)
                        if event_details:
                            self.save_event_metadata(event_details, self.PackageDetailsFilePath)

                        #  Save to Excel or SQL using existing function
                        

                    except Exception as ex:
                        self.log_error(package_path, ex)

                for conn_path in connection_manager_files:
                    if conn_path in self.processed_package_paths:
                        continue
                    try:
                        self.processed_package_paths.add(conn_path)
                        
                        self.analyze_single_connection_manager(conn_path)
                    except Exception as ex:
                        self.log_error(conn_path, ex)

                for param_path in param_files:
                    if param_path in self.processed_package_paths:
                        continue
                    try:
                        self.processed_package_paths.add(param_path)
                        self.analyze_param_manager(param_path)
                    except Exception as ex:
                        self.log_error(param_path, ex)

            except Exception as ex:
                print(f"Error accessing directory {root}: {ex}")

        self.save_update_connection_name(self.PackageDetailsFilePath)
        print("Completed...")
    
    ## extarcting individual dtsx files into xlsx
    def generate_dataflow_mapping_excel(self, dtsx_file_path, output_folder):
        try:
            tree = ET.parse(dtsx_file_path)
            root = tree.getroot()
            namespace = {'DTS': 'www.microsoft.com/SqlServer/Dts'}

            package_name = os.path.basename(dtsx_file_path)
            package_path = os.path.dirname(dtsx_file_path)

            rows = []

            # Define columns to ignore
            ignore_columns = {"ErrorCode", "ErrorColumn", "Flat File Source Error Output Column"}

            for executable in root.findall(".//DTS:Executable", namespace):
                exec_type = executable.attrib.get('{www.microsoft.com/SqlServer/Dts}ExecutableType', '')
                task_name = executable.attrib.get('{www.microsoft.com/SqlServer/Dts}ObjectName', 'Unknown')

                if exec_type == "Microsoft.Pipeline":
                    for component in executable.findall(".//{*}component"):
                        component_name = component.attrib.get('name', '')
                        ref_id = component.attrib.get('refId', '')
                        component_class_id = component.attrib.get('componentClassID', '')

                        # Process input columns
                        for inputs in component.findall(".//{*}inputs/{*}input"):
                            input_cols = inputs.findall(".//{*}inputColumn")
                            for col in input_cols:
                                col_name = col.attrib.get('cachedName', '').strip()
                                if col_name not in ignore_columns:  # Skip unwanted columns
                                    rows.append({
                                        "PackageName": package_name,
                                        "PackagePath": package_path,
                                        "TaskName": task_name,
                                        "ColumnName": col_name,
                                        "ColumnType": "OT_INPUTCOLUMN",
                                        "DataType": f"DT_{col.attrib.get('cachedDataType', '').upper()}",
                                        "ComponentName": component_name,
                                        "DataConversion": "",
                                        "ComponentPropertyDetails": "",
                                        "ColumnPropertyDetails": "",
                                        "isEventHandler": 0
                                    })

                        for outputs in component.findall(".//{*}outputs/{*}output"):
                            output_cols = outputs.findall(".//{*}outputColumn")
                            for col in output_cols:
                                col_name = col.attrib.get('name', '').strip()
                                if col_name not in ignore_columns:  # Skip unwanted columns
                                    rows.append({
                                        "PackageName": package_name,
                                        "PackagePath": package_path,
                                        "TaskName": task_name,
                                        "ColumnName": col_name,
                                        "ColumnType": "OT_OUTPUTCOLUMN",
                                        "DataType": f"DT_{col.attrib.get('dataType', '').upper()}",
                                        "ComponentName": component_name,
                                        "DataConversion": "",
                                        "ComponentPropertyDetails": "",
                                        "ColumnPropertyDetails": "",
                                        "isEventHandler": 0
                                    })

            if rows:
                df = pd.DataFrame(rows)
                excel_filename = os.path.splitext(package_name)[0] + "_DFM.xlsx"
                full_path = os.path.join(output_folder, excel_filename)

                # Save with custom sheet name
                with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name="DataFlowTaskMappingDetails")

                # Format the file after writing
                format_excel_file(full_path)

                print(f"Generated DFM file: {full_path}")

        except Exception as e:
            print(f"Error parsing {dtsx_file_path}: {e}")


    def extract_package_task_details(self, package_path):
        details = []
        try:
            tree = ET.parse(package_path)
            root = tree.getroot()
            ns = {
                'DTS': 'www.microsoft.com/SqlServer/Dts',
                'SQLTask': 'www.microsoft.com/sqlserver/dts/tasks/sqltask'
            }

            package_name = os.path.basename(package_path)
            package_folder = os.path.dirname(package_path)

            for executable in root.findall(".//DTS:Executable", ns):
                ref_id = executable.get("{www.microsoft.com/SqlServer/Dts}refId", "")
                task_name = executable.get("{www.microsoft.com/SqlServer/Dts}ObjectName", "").strip()
                task_type = executable.get("{www.microsoft.com/SqlServer/Dts}Description", "").strip()

                # Skip if refId contains EventHandlers[OnError]
                if "EventHandlers[OnError]" in ref_id:
                    continue

                # Skip if ObjectName is "PackageError"
                if task_name in ("PackageError","Foreach Loop Container"):
                    continue

                # Skip if the task name contains 'Sequence'
                if "Sequence" in task_name:
                    continue

                # Extract container name if hierarchy is valid
                container_name = ""
                if ref_id.startswith("Package\\"):
                    parts = ref_id.split("\\")
                    if len(parts) >= 3:
                        # e.g., Package\Foreach Loop Container\Archive Task
                        container_name = parts[-2]  # second last part
                    else:
                        container_name = ""  # no container

                # Initialize placeholders
                task_sql_query = ""
                expressions = []
                variables = []
                parameters = []
                source_component = ""
                source_type = ""
                target_component = ""
                target_type = ""
                target_table = ""
                task_connection_name = ""
                result_set_details = ""
                task_component_details = ""
                source_connection_name = ""
                target_connection_name = ""

                # Extract Expressions and SQLQuery from DTS:PropertyExpression
                for prop in executable.findall(".//DTS:PropertyExpression", ns):
                    prop_name = prop.attrib.get("{www.microsoft.com/SqlServer/Dts}Name", "").strip()
                    prop_value = prop.text.strip() if prop.text else ""
                    prop_value = (
                        prop_value.replace("&lt;", "<")
                                  .replace("&gt;", ">")
                                  .replace("&#xA;", "\n")
                    )

                    # Only consider if property is SqlStatementSource and not a package error
                    if prop_name == "SqlStatementSource":
                        continue

                    # Add to expressions only if not SqlStatementSource
                    if prop_name or prop_value:
                        expressions.append(f"Property: {prop_name}, Expression: {prop_value}")

                    # If SQL Query not already extracted, set it here
                    if not task_sql_query and "insert into packageError" not in prop_value.lower():
                        task_sql_query = prop_value

                # If still not found, check inside SQLTask
                object_data = executable.find("DTS:ObjectData", ns)
                if object_data is not None and not task_sql_query:
                    sql_node = object_data.find("SQLTask:SqlTaskData", ns)
                    if sql_node is not None:
                        sql_text = sql_node.get("{www.microsoft.com/sqlserver/dts/tasks/sqltask}SqlStatementSource", "")
                        if sql_text and "insert into packageError" not in sql_text.lower():
                            task_sql_query = (
                                sql_text.replace("&lt;", "<")
                                        .replace("&gt;", ">")
                                        .replace("&#xA;", "\n")
                            )

                
                # --- Extract TaskConnectionName inline ---
                task_connection_guid = ""

                if object_data is not None:
                    sql_node = object_data.find("SQLTask:SqlTaskData", ns)
                    if sql_node is not None:
                        raw_conn_id = sql_node.get("{www.microsoft.com/sqlserver/dts/tasks/sqltask}Connection", "")
                        task_connection_guid = raw_conn_id.strip().replace("{", "").replace("}", "")

                        if task_connection_guid:
                            # print(f"[DEBUG] Searching for connections for Task: {task_name}, GUID: {task_connection_guid}")
                            all_conn_nodes = root.findall(".//connection")
                            for conn_node in all_conn_nodes:
                                conn_mgr_id = conn_node.attrib.get("connectionManagerID", "")
                                conn_ref_id = conn_node.attrib.get("connectionManagerRefId", "")
                                conn_name = conn_node.attrib.get("name", "")

                                if conn_mgr_id.startswith(f"{{{task_connection_guid}}}"):
                                    # Final connection name comes from connectionManagerRefId
                                    if conn_ref_id.startswith("Project.ConnectionManagers[") and conn_ref_id.endswith("]"):
                                        task_connection_name = conn_ref_id.split("[")[1].replace("]", "")
                                        # print(f"[DEBUG] Found connection for {task_name} -> {task_connection_name}")
                                        break
                    

                # Extract Variables (from FileSystemData under ObjectData)
                object_data = executable.find("DTS:ObjectData", ns)
                if object_data is not None:
                    file_system_data = object_data.find("FileSystemData")
                    if file_system_data is not None:
                        # Extract TaskSourcePath if it's marked as a variable
                        if file_system_data.attrib.get("TaskIsSourceVariable", "").lower() == "true":
                            source_path_var = file_system_data.attrib.get("TaskSourcePath", "").strip()
                            if source_path_var:
                                variables.append(f"Source Path: {source_path_var}")

                        # Extract TaskDestinationPath if it's marked as a variable
                        if file_system_data.attrib.get("TaskIsDestinationVariable", "").lower() == "true":
                            dest_path_var = file_system_data.attrib.get("TaskDestinationPath", "").strip()
                            if dest_path_var:
                                variables.append(f"Destination Path: {dest_path_var}")

                # Extract Parameters (from SQLTask:ParameterBinding inside SQLTask:SqlTaskData)
                if object_data is not None:
                    sql_task_node = object_data.find("SQLTask:SqlTaskData", ns)
                    if sql_task_node is not None:
                        for param_binding in sql_task_node.findall("SQLTask:ParameterBinding", ns):
                            param_name = param_binding.get("{www.microsoft.com/sqlserver/dts/tasks/sqltask}ParameterName", "")
                            direction = param_binding.get("{www.microsoft.com/sqlserver/dts/tasks/sqltask}ParameterDirection", "")
                            data_type = param_binding.get("{www.microsoft.com/sqlserver/dts/tasks/sqltask}DataType", "")
                            dts_variable = param_binding.get("{www.microsoft.com/sqlserver/dts/tasks/sqltask}DtsVariableName", "")

                            parameters.append(
                                f"Name: {param_name}, Type: {direction}, DataType: {data_type}, Value: , DtsVariableName: {dts_variable}"
                            )


                # Extract DataFlow components
                # DataFlowTask Connection Names
                source_conn_name = ""
                target_conn_name = ""

                if object_data is not None:
                    pipeline = object_data.find("pipeline")
                    if pipeline is not None:
                        for component in pipeline.findall(".//component"):
                            comp_desc = component.get("description", "")
                            comp_name = component.get("name", "")

                            if "Source" in comp_desc:
                                source_component = comp_name
                                source_type = comp_desc
                            elif "Destination" in comp_desc:
                                target_component = comp_name
                                target_type = comp_desc

                            connections = component.find("connections")
                            if connections is not None:
                                for conn in connections.findall("connection"):
                                    ref_id = conn.get("connectionManagerRefId", "")
                                    if not ref_id:
                                        continue
                                    conn_clean = ref_id.split("[")[-1].replace("]", "").strip()

                                    # Match Source
                                    if any(x in comp_desc for x in ["OLE DB Source", "Flat File Source", "ODBC Source"]):
                                        source_conn_name = conn_clean
                                    # Match Target
                                    elif any(x in comp_desc for x in ["OLE DB Destination", "Flat File Destination", "ODBC Destination"]):
                                        target_conn_name = conn_clean

                # Assign values to output fields
                source_connection_name = source_conn_name
                target_connection_name = target_conn_name


                # Target Table from OpenRowSet
                open_rowset = executable.findall(".//property[@name='OpenRowset']")
                if open_rowset:
                    target_table = open_rowset[0].text

                # ResultSetDetails
                result_set_parts = []
                sql_task_node = executable.find("DTS:ObjectData/SQLTask:SqlTaskData", ns)

                if sql_task_node is not None:
                    result_bindings = sql_task_node.findall("SQLTask:ResultBinding", ns)
                    for result in result_bindings:
                        result_name = result.attrib.get("{www.microsoft.com/sqlserver/dts/tasks/sqltask}ResultName", "").strip()
                        variable_name = result.attrib.get("{www.microsoft.com/sqlserver/dts/tasks/sqltask}DtsVariableName", "").strip()
                        if result_name and variable_name:
                            result_set_parts.append(
                                f"Result Set Column: {result_name} | SSIS Variable: {variable_name}"
                            )
                result_set_details = "\n".join(result_set_parts)

                # TaskComponentDetails (cmd.exe)
                executable_type = executable.attrib.get("{www.microsoft.com/SqlServer/Dts}ExecutableType", "")
                object_data = executable.find("DTS:ObjectData", ns)

                if executable_type.endswith("Task") and object_data is not None:
                    # Case 1: ExecutePackageTask → extract only <PackageName> text
                    exec_pkg = object_data.find("ExecutePackageTask")
                    if exec_pkg is not None:
                        pkg_name_tag = exec_pkg.find("PackageName")
                        if pkg_name_tag is not None and pkg_name_tag.text:
                            task_component_details = pkg_name_tag.text.strip()

                    # Case 2: FileSystemTask → extract source and destination path
                    fs_data = object_data.find("FileSystemData")
                    if fs_data is not None:
                        src_path = fs_data.attrib.get("TaskSourcePath", "").strip()
                        dst_path = fs_data.attrib.get("TaskDestinationPath", "").strip()
                        parts = []
                        if src_path:
                            parts.append(f"SourcePath: {src_path}")
                        if dst_path:
                            parts.append(f"DestinationPath: {dst_path}")
                        task_component_details = " | ".join(parts)


                details.append({
                    "PackageName": package_name,
                    "PackagePath": package_folder,
                    "TaskName": task_name,
                    "TaskType": task_type,
                    "ContainerName": container_name,
                    "TaskConnectionName": task_connection_name,
                    "TaskSqlQuery": task_sql_query,
                    "Variables": "; ".join(variables),
                    "Parameters": "; ".join(parameters),
                    "Expressions": "; ".join(expressions),
                    "SourceComponent": source_component,
                    "SourceType": source_type,
                    "TargetComponent": target_component,
                    "TargetType": target_type,
                    "TargetTable": target_table,
                    "SourceConnectionName": source_connection_name,
                    "TargetConnectionName": target_connection_name,
                    "ResultSetDetails": result_set_details,
                    "TaskComponentDetails": task_component_details
                })

            return details

        except Exception as e:
            print(f"Error parsing {package_path}: {e}")
            return []


    def extract_package_container_details(self, package_path):
        details = []
        try:
            tree = ET.parse(package_path)
            root = tree.getroot()
            ns = {'DTS': 'www.microsoft.com/SqlServer/Dts'}

            package_name = os.path.basename(package_path)
            package_folder = os.path.dirname(package_path)

            for container in root.findall(".//DTS:Executable", ns):
                exec_type = container.get("{www.microsoft.com/SqlServer/Dts}ExecutableType", "")

                # Only interested in Sequence and ForeachLoop
                if exec_type not in ("STOCK:SEQUENCE", "STOCK:FOREACHLOOP"):
                    continue
                
                container_name = container.get("{www.microsoft.com/SqlServer/Dts}ObjectName", "")

                # ContainerType from Description (remove " Container" and normalize casing)
                raw_desc = container.get("{www.microsoft.com/SqlServer/Dts}Description", "")
                if "Foreach" in raw_desc:
                    container_type = "ForEachLoop"
                elif "Sequence" in raw_desc:
                    container_type = "Sequence"
                else:
                    container_type = ""

                expression = ""
                enumerator = ""

                # Extract only if <DTS:ForEachEnumerator> exists
                for_each_enum = container.find(".//DTS:ForEachEnumerator", ns)
                if for_each_enum is not None:
                    # Extract PropertyExpression for expression
                    prop_expr = for_each_enum.find(".//DTS:PropertyExpression", ns)
                    if prop_expr is not None:
                        expr_name = prop_expr.get("{www.microsoft.com/SqlServer/Dts}Name", "")
                        expr_value = prop_expr.text.strip() if prop_expr.text else ""
                        expression = f"Property : {expr_name}, Expression : {expr_value}"

                    # Extract enumerator details
                    enum_parts = []
                    dir_expr = for_each_enum.find(".//DTS:PropertyExpression", ns)
                    if dir_expr is not None:
                        dir_value = dir_expr.text.strip() if dir_expr.text else ""
                        enum_parts.append(f"EnumeratorName : Directory, EnumeratorValue: {dir_value}")

                    # Then extract from ForEachFileEnumeratorProperties
                    fefe_props = for_each_enum.find(".//ForEachFileEnumeratorProperties")
                    if fefe_props is not None:
                        for fefe in fefe_props.findall(".//FEFEProperty"):
                            for key, val in fefe.attrib.items():
                                enum_name = key
                                enum_value = "True" if key == "Recurse" and val == "1" else val
                                enum_parts.append(f"EnumeratorName : {enum_name}, EnumeratorValue: {enum_value}")
                        
                    enumerator = " | ".join(enum_parts)

                details.append({
                    "PackageName": package_name,
                    "PackagePath": package_folder,
                    "ContainerName": container_name,
                    "ContainerType": container_type,
                    "ContainerExpression": expression,
                    "ContainerEnum": enumerator
                    })

            return details

        except Exception as e:
            print(f"Error extracting PackageContainerDetails from {package_path}: {e}")
            return []

    # Helper: Extract schema names from SQL
    @staticmethod
    def extract_schema_list(sql_text):
        return [schema for schema in re.findall(r'\[([^\]]+)\]', sql_text)]

    # Helper: Extract first schema/table name
    @staticmethod
    def extract_schema_from_sql(sql_text):
        match = re.search(r'FROM\s+\[?([^\]]+)\]?', sql_text, re.IGNORECASE)
        return match.group(1) if match else ""

    # Helper: Extract CMD details
    @staticmethod
    def extract_cmd_details(executable):
        cmd = executable.attrib.get("{www.microsoft.com/SqlServer/Dts}ExecutableType", "")
        args = executable.attrib.get("{www.microsoft.com/SqlServer/Dts}Arguments", "")
        work_dir = executable.attrib.get("{www.microsoft.com/SqlServer/Dts}WorkingDirectory", "")
        return f"Executable: {cmd} | Arguments: {args} | WorkingDirectory: {work_dir}"



    def analyze_param_manager(self, param_file_path):
        try:
            tree = ET.parse(param_file_path)
            root = tree.getroot()
            ns = {'SSIS': 'www.microsoft.com/SqlServer/SSIS'}

            # Metadata container
            metadata = {
                "ProjectParameterDetails": [],
                "ProjectPath": os.path.dirname(param_file_path),
                "ProjectFile": os.path.basename(param_file_path)
            }

            # SSIS DataType mapping
            ssis_data_types = {
                "3": "Boolean", "6": "Byte", "16": "DateTime", "15": "Decimal", "14": "Double",
                "7": "Int16", "9": "Int32", "11": "Int64", "5": "SByte", "13": "Single",
                "18": "String", "10": "UInt32", "12": "UInt64"
            }

            # Find all parameters in .params file
            parameter_nodes = root.findall(".//SSIS:Parameter", ns)
            for param in parameter_nodes:
                param_name = param.attrib.get("{www.microsoft.com/SqlServer/SSIS}Name")
                value_node = param.find("SSIS:Properties/SSIS:Property[@SSIS:Name='Value']", ns)
                datatype_node = param.find("SSIS:Properties/SSIS:Property[@SSIS:Name='DataType']", ns)

                value = value_node.text if value_node is not None else ""
                datatype_code = datatype_node.text if datatype_node is not None else ""
                datatype_name = ssis_data_types.get(datatype_code, datatype_code)

                metadata["ProjectParameterDetails"].append({
                    "ParameterName": param_name,
                    "Value": value,
                    "DataType": datatype_name
                })

            # Save extracted parameter details
            self.save_project_parameter_metadata(metadata, self.PackageDetailsFilePath)

        except Exception as e:
            print(f"Error analyzing param file {param_file_path}: {e}")

    
    def extract_event_handler_task_details(self, package_path):
        details = []
        try:
            tree = ET.parse(package_path)
            root = tree.getroot()
            ns = {
                  'DTS': 'www.microsoft.com/SqlServer/Dts',
                  'SQLTask': 'www.microsoft.com/sqlserver/dts/tasks/sqltask'
                 }

            package_name = os.path.basename(package_path)
            package_folder = os.path.dirname(package_path)

            # Find all EventHandlers
            for event_handler in root.findall(".//DTS:EventHandler", ns):
                event_type = event_handler.get("{www.microsoft.com/SqlServer/Dts}EventName", "")
                if event_type != "OnError":  # Only process OnError handlers
                    continue

                # Extract EventHandlerName from refId → after "Package\" and before ".EventHandlers"
                ref_id = event_handler.get("{www.microsoft.com/SqlServer/Dts}refId", "")
                event_handler_name = ""
                event_handler_type = ""

                # Extract EventHandlerName
                if ref_id.startswith("Package\\"):
                    parts = ref_id[len("Package\\"):].split("\\")

                    # Handle EventHandlerName
                    if len(parts) == 1 and ".EventHandlers" in parts[0]:
                        event_handler_name = parts[0].split(".EventHandlers")[0]
                    elif len(parts) >= 2 and ".EventHandlers" in parts[-1]:
                        event_handler_name = parts[-2]

                    # Handle EventHandlerType
                    if len(parts) == 1 and ".EventHandlers" in parts[0]:
                        event_handler_type = "Sequence"
                    elif len(parts) >= 2 and ".EventHandlers" in parts[-1]:
                        event_handler_type = "ExecuteSQLTask"

                # Loop through tasks inside EventHandler
                for executable in event_handler.findall(".//DTS:Executable", ns):
                    task_name = executable.get("{www.microsoft.com/SqlServer/Dts}ObjectName", "")
                    task_type_raw = executable.get("{www.microsoft.com/SqlServer/Dts}ExecutableType", "")
                    task_type = task_type_raw.split(".")[-1] if task_type_raw else ""

                    # Extract Expressions
                    expressions = []
                    for prop in executable.findall(".//DTS:PropertyExpression", ns):
                        prop_name = prop.get("{www.microsoft.com/SqlServer/Dts}Name", "")
                        expressions.append(f"{prop_name}: {prop.text.strip() if prop.text else ''}")

                    # Extract SQL Query from SqlTaskData
                    sql_query = ""
                    sql_node = executable.find(".//SQLTask:SqlTaskData", ns)
                    if sql_node is not None:
                        sql_query = sql_node.attrib.get("{www.microsoft.com/sqlserver/dts/tasks/sqltask}SqlStatementSource", "")

                    # Extract TaskConnectionName (from SQL query if available)
                    task_connection_name = ""
                    if "insert into" in sql_query.lower():
                        task_connection_name = "ODS_Mode"

                    # Append final record
                    details.append({
                        "PackageName": package_name,
                        "PackagePath": package_folder,
                        "EventHandlerName": event_handler_name,
                        "EventHandlerType": event_handler_type,
                        "EventType": event_type,
                        "TaskName": task_name,
                        "TaskType": task_type,
                        "ContainerName": "",
                        "ContainerType": "",
                        "ContainerExpression": "",
                        "TaskConnectionName": task_connection_name,
                        "SqlQuery": sql_query,
                        "Variables": "",
                        "Parameters": "",
                        "Expressions": "; ".join(expressions),
                        "DataFlowDaskSourceName": "",
                        "DataFlowTaskSourceType": "",
                        "DataFlowTaskTargetName": "",
                        "DataFlowTaskTargetType": "",
                        "DataFlowTaskTargetTable": "",
                        "DataFlowDaskSourceConnectionName": "",
                        "DataFlowDaskTargetConnectionName": "",
                        "SendMailTaskDetails": "",
                        "ResultSetDetails": "",
                        "TaskComponentDetails": ""
                    })

            return details

        except Exception as e:
            print(f"Error parsing EventHandlers from {package_path}: {e}")
            return []


    def parse_main_pipe(self,main_pipe_elem, namespaces):
        main_pipe = MainPipe()
        #print(main_pipe.ComponentMetaDataCollection)
        for comp_elem in main_pipe_elem.findall(".//{*}component", namespaces):
            comp_name = comp_elem.attrib.get("name", "")
            comp_class_id = comp_elem.attrib.get("componentClassID", "")
            print (comp_name)
            #print (comp_class_id)

            # INPUTS
            inputs = []
            for input_elem in comp_elem.findall(".inputs/input", namespaces):
                input_name = input_elem.attrib.get("name")
                #print(input_name)
                input_cols = []
                for col_elem in input_elem.findall(".inputColumns/inputColumn", namespaces):
                    col_name = col_elem.attrib.get("cachedName")
                    col_type = col_elem.attrib.get("cachedDataType")
                    #obj_type = col_elem.attrib.get("objectType")
                    '''props = [
                        CustomProperty(p.attrib.get("name", ""), p.attrib.get("value", ""), p.attrib.get("expressionType"))
                        for p in col_elem.findall(".properties/property", namespaces)
                    ]'''
                    input_cols.append(InputColumn(col_name, col_type))

                '''input_props = [
                    CustomProperty(p.attrib.get("name", ""), p.attrib.get("value", ""), p.attrib.get("expressionType"))
                    for p in input_elem.findall(".//pipeline:property", namespaces)
                ]'''
                inputs.append(Input(input_name, input_cols))
                #print(inputs)
            # OUTPUTS
            outputs = []
            for output_elem in comp_elem.findall(".outputs/output", namespaces):
                output_name = output_elem.attrib.get("name", "")
                output_cols = []
                for col_elem in output_elem.findall(".outputColumns/outputColumn", namespaces):
                    col_name = col_elem.attrib.get("name", "")
                    col_type = col_elem.attrib.get("dataType", "")
                    '''obj_type = col_elem.attrib.get("objectType", "")
                    props = [
                        CustomProperty(p.attrib.get("name", ""), p.attrib.get("value", ""), p.attrib.get("expressionType"))
                        for p in col_elem.findall(".//pipeline:property", namespaces)
                    ]'''
                    output_cols.append(OutputColumn(col_name, col_type))

                '''output_props = [
                    CustomProperty(p.attrib.get("name", ""), p.attrib.get("value", ""), p.attrib.get("expressionType"))
                    for p in output_elem.findall(".//pipeline:property", namespaces)
                ]'''
                outputs.append(Output(output_name, output_cols))
                #print(outputs)
            component = Component(comp_name, comp_class_id, inputs, outputs)
            main_pipe.ComponentMetaDataCollection.append(component)
        return main_pipe

    def analyze_single_connection_manager(self, connection_manager_path):
        try:
            tree = ET.parse(connection_manager_path)
            root = tree.getroot()

            ns = {'DTS': 'www.microsoft.com/SqlServer/Dts'}

            base_node = root  # DTS:ConnectionManager
            object_data_node = root.find("DTS:ObjectData", ns)
            inner_conn_node = object_data_node.find("DTS:ConnectionManager", ns) if object_data_node is not None else None

            connection_name = base_node.attrib.get("{www.microsoft.com/SqlServer/Dts}ObjectName", "")
            connection_type = base_node.attrib.get("{www.microsoft.com/SqlServer/Dts}CreationName", "")
            connection_id = base_node.attrib.get("{www.microsoft.com/SqlServer/Dts}DTSID", "")
            connection_string = inner_conn_node.attrib.get("{www.microsoft.com/SqlServer/Dts}ConnectionString", "") if inner_conn_node is not None else ""

            # Collect expressions
            connection_expression = ""
            for expr_node in root.findall(".//DTS:PropertyExpression", ns):
                name = expr_node.attrib.get("{www.microsoft.com/SqlServer/Dts}Name", "")
                value = expr_node.text or ""
                connection_expression += f"{name}: {value} "

            metadata = {
                "Connections": [],
                "PackagePath": os.path.dirname(connection_manager_path),
                "PackageName": os.path.basename(connection_manager_path)
            }

            metadata["Connections"].append({
                "ConnectionName": connection_name,
                "ConnectionType": connection_type,
                "ConnectionString": connection_string,
                "ConnectionExpressions": connection_expression.strip(),
                "ConnectionID": connection_id,
                "IsProjectConnection": "1"
            })

            self.save_connections_metadata(metadata, self.PackageDetailsFilePath)

        except Exception as e:
            self.log_error(connection_manager_path, e)


    def analyze_single_package(self,package_path):
        # Reset component count and set for this package
        self.ComponentCount = 0
        self.ComponentNameCheck = set()

        class MockPackage:
            def __init__(self, path):
                self.path = path
                self.creation_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                self.creator_name = "Unknown"   # Default value
                self.executables = []            # Used in task/container processing
                self.connections = []            # Used in connection processing
                self.variables = []              # Used in variable extraction
                self.event_handlers = []         # Used in event handler extraction
                self.parameters = []             # Used in parameters extraction 
                self.PrecedenceConstraints = []  # Used in PrecedenceConstraints extraction 

                # Try extracting creator name from the XML itself
                try:
                    tree = ET.parse(path)
                    root = tree.getroot()
                    self.creator_name = root.attrib.get('{www.microsoft.com/SqlServer/Dts}CreatorName', 'Unknown')
                    self.root = root
                except Exception as e:
                    print(f"Could not extract creator name from {path}: {e}")
                    self.root = None

            def execute(self):
                # Simulate some execution delay
                time.sleep(0.5)  # Sleep for half a second     

        def extract_tasks_from_xml(root, package):
            namespace = {'DTS': 'www.microsoft.com/SqlServer/Dts',
                        'pipeline': 'www.microsoft.com/SqlServer/Dts/Pipeline'}
            for executable in root.findall(".//{*}Executable"):
                exec_type = executable.attrib.get("{www.microsoft.com/SqlServer/Dts}ExecutableType")
                task_name = executable.attrib.get("{www.microsoft.com/SqlServer/Dts}ObjectName") or executable.attrib.get("Name")
                
                if exec_type and task_name:
                    task_host = TaskHost(task_name)
                    #print(exec_type)
                    # Instantiate correct object based on task type
                    if exec_type == "Microsoft.ForEachLoopContainer":
                        task_host.InnerObject = ForEachLoop()
                    elif exec_type == "Microsoft.ForLoopContainer":
                        task_host.InnerObject = ForLoop()
                    elif exec_type == "STOCK:SEQUENCE":
                        task_host.InnerObject = Sequence()
                    elif exec_type == "Microsoft.Pipeline":
                        main_pipe_elem = executable.find(".//{*}ObjectData/{*}pipeline")
                        #print(main_pipe_elem)
                        if main_pipe_elem is not None:
                            #components = main_pipe_elem.findall(".//{*}component", namespace)
                            #print(len(components))
                            #task_host = TaskHost(name=executable.attrib["ObjectName"])
                            task_host.InnerObject = self.parse_main_pipe(main_pipe_elem, namespace)
                    else:
                        task_host.InnerObject = exec_type  # fallback to string

                    package.executables.append(task_host)

        package = MockPackage(package_path)

        if package.root is None:
            print(f"Skipping file due to XML parsing error: {package_path}")
            return
        
        # Traverse and parse for task details
        extract_tasks_from_xml(package.root, package)
        #print(package.executables)

        try:
            # tree = ET.parse(package_path)
            # root = tree.getroot()
            SSISPackageAnalyzer.traverse_xml(package.root)
        except Exception as ex:
            print(f"Error in XML traversal: {ex}")

        # package_name = os.path.basename(package_path)
        # package_dir = os.path.dirname(package_path)

        result = PackageAnalysisResult()
        result.Variables = self.extract_variables_and_parameters(package_path)
        result.PackageName = os.path.basename(package_path)
        result.PackagePath = os.path.dirname(package_path)
        result.CreatedDate = package.creation_date
        result.CreatedBy = package.creator_name
        result.ExecutionTime = self.measure_package_performance(package)
        result.Tasks = self.count_package_tasks(package)
        result.Connections = self.count_package_connections(package)
        result.Containers = self.count_sequence_container(package) #self.count_package_containers(package)
        result.DTSXXML = ET.tostring(package.root, encoding="unicode")
        # result.Variables = self.get_package_variables(package)

        for executable in package.executables:
            if isinstance(executable.InnerObject, ForEachLoop):
                result.Foreachtasks.extend(self.process_foreach_loop_container_details(executable, [], package))
       #     elif isinstance(executable.InnerObject, Sequence):
        #        result.Seqtasks.extend(self.process_sequence_container_details(executable, [], package))
            elif isinstance(executable.InnerObject, ForLoop):
                result.Forlooptasks.extend(self.process_for_loop_container_details(executable, [], package))
            elif isinstance(executable.InnerObject, MainPipe):
                self.extract_data_flow_task(executable, "0")

        result.SequenceContainerTaskCount = self.count_sequence_container_tasks(package)
        result.ForeachContainerTaskCount = self.count_foreach_container_tasks(package)
        result.ForLoopContainerTaskCount = self.count_forloop_container_tasks(package)
        result.ExecutionTime = self.measure_package_performance(package)

        # Save metadata
        self.save_package_metadata(result, self.PackageAnalysisFilePath, self.PackageDetailsFilePath)

        # Fix: define package_name from pat
        package_name = os.path.basename(package_path)

        self.extract_precedence_constraints_for_task(package_path,package_name)
        self.extract_event_handlers_for_package(package)

    
    @staticmethod
    def traverse_xml(node: ET.Element):
        if node is not None:
            for child in node:
                SSISPackageAnalyzer.traverse_xml(child)

                                
    def get_package_variables(self,package) -> list:
        variables = []

        for variable in package.variables:
            if not variable.system_variable:
                variables.append(VariableInfo(
                    name=variable.name,
                    value=str(variable.value) if variable.value is not None else None,
                    data_type=str(variable.data_type),
                    namespace=variable.namespace,
                    is_parameter=0
                ))

        for parameter in package.parameters:
            variables.append(VariableInfo(
                name=parameter.name,
                value=str(parameter.value) if parameter.value is not None else None,
                data_type=str(parameter.data_type),
                is_parameter=1
            ))

        return variables
                             
    def count_sequence_container_tasks(self,package) -> list:
        tasks_in_sequence = []

        for executable in package.executables:
            if isinstance(executable, Sequence):
                process_container_sequence_loop(executable, tasks_in_sequence, package)

            elif isinstance(executable, ForEachLoop):
                process_container_foreach_loop(executable, tasks_in_sequence, package)

            elif isinstance(executable, ForLoop):
                process_container_for_loop(executable, tasks_in_sequence, package)

        self.container_task_count += len(tasks_in_sequence)

        return tasks_in_sequence

                              
    def count_foreach_container_tasks(self,package) -> list:
        tasks_in_for_each = []

        for executable in package.executables:
            if isinstance(executable, ForEachLoop):
                process_container_foreach_loop(executable, tasks_in_for_each, package)

            elif isinstance(executable, Sequence):
                process_container_sequence_loop(executable, tasks_in_for_each, package)

            elif isinstance(executable, ForLoop):
                process_container_for_loop(executable, tasks_in_for_each, package)

        self.container_task_count += len(tasks_in_for_each)

        return tasks_in_for_each

                                             
    def count_forloop_container_tasks(self,package) -> list:
        tasks_in_for_loop = []

        for executable in package.executables:
            if isinstance(executable, ForEachLoop):
                process_container_foreach_loop(executable, tasks_in_for_loop, package)

            elif isinstance(executable, Sequence):
                process_container_sequence_loop(executable, tasks_in_for_loop, package)

            elif isinstance(executable, ForLoop):
                process_container_for_loop(executable, tasks_in_for_loop, package)

        self.container_task_count += len(tasks_in_for_loop)
        
        return tasks_in_for_loop

    def process_container_foreach_loop(container, tasks_in_for_each: list, package):
        # Check if the container is a ForEachLoop (redundant in Python, but included for structure)
        if isinstance(container, ForEachLoop):
            pass  # Placeholder for any logic specific to ForEachLoop

        # Iterate over nested executables
        for nested_executable in container.executables:
            if isinstance(nested_executable, ForEachLoop):
                tasks_in_loop = process_foreach_loop_container_details(nested_executable, [], package)
                tasks_in_for_each.extend(tasks_in_loop)

            elif isinstance(nested_executable, Sequence):
                tasks_in_loop = process_sequence_container_details(nested_executable, [], package)
                tasks_in_for_each.extend(tasks_in_loop)

            elif isinstance(nested_executable, ForLoop):
                tasks_in_loop = process_for_loop_container_details(nested_executable, [], package)
                tasks_in_for_each.extend(tasks_in_loop)
                
                               
    def process_container_sequence_loop(container, tasks_in_for_each: list, package):
        # Check if the container is a Sequence (redundant in Python but included for clarity)
        if isinstance(container, Sequence):
            pass  # Placeholder for any logic specific to Sequence containers

        # Iterate over nested executables
        for nested_executable in container.executables:
            if isinstance(nested_executable, ForEachLoop):
                tasks_in_loop = process_foreach_loop_container_details(nested_executable, [], package)
                tasks_in_for_each.extend(tasks_in_loop)

            elif isinstance(nested_executable, Sequence):
                tasks_in_loop = process_sequence_container_details(nested_executable, [], package)
                tasks_in_for_each.extend(tasks_in_loop)

            elif isinstance(nested_executable, ForLoop):
                tasks_in_loop = process_for_loop_container_details(nested_executable, [], package)
                tasks_in_for_each.extend(tasks_in_loop)
                

    def process_container_for_loop(container, tasks_in_for_each: list, package):
        # Check if the container is a ForLoop (redundant in Python, as it's already typed)
        if isinstance(container, ForLoop):
            pass  # Placeholder for any logic if needed for the base ForLoop

        # Iterate over nested executables
        for nested_executable in container.executables:
            if isinstance(nested_executable, ForEachLoop):
                tasks_in_loop = process_foreach_loop_container_details(nested_executable, [], package)
                tasks_in_for_each.extend(tasks_in_loop)

            elif isinstance(nested_executable, Sequence):
                tasks_in_loop = process_sequence_container_details(nested_executable, [], package)
                tasks_in_for_each.extend(tasks_in_loop)

            elif isinstance(nested_executable, ForLoop):
                tasks_in_loop = process_for_loop_container_details(nested_executable, [], package)
                tasks_in_for_each.extend(tasks_in_loop)

    def recursive_count(self, executable_element, req_creation_name):
            namespace = {'DTS': 'www.microsoft.com/SqlServer/Dts'}
            count = 0
            creation_name = executable_element.attrib.get('{www.microsoft.com/SqlServer/Dts}CreationName', '')

            # Count only Execute SQL and Execute Package tasks
            if creation_name in req_creation_name:
                count += 1

            # Recurse into nested Executables
            for executables in executable_element.findall('DTS:Executables', namespace):
                for child_executable in executables.findall('DTS:Executable', namespace):
                    count += self.recursive_count(child_executable,req_creation_name)

            # Recurse into Event Handlers as well
            for event_handlers in executable_element.findall('DTS:EventHandlers', namespace):
                for event_handler in event_handlers.findall('DTS:EventHandler', namespace):
                    for handler_executables in event_handler.findall('DTS:Executables', namespace):
                        for handler_executable in handler_executables.findall('DTS:Executable', namespace):
                            count += self.recursive_count(handler_executable,req_creation_name)

            return count
                                                                            
    def count_package_tasks(self, package):
        """
        Recursively count only Execute SQL and Execute Package tasks (exclude containers and Data Flow tasks).
        """
        namespace = {'DTS': 'www.microsoft.com/SqlServer/Dts'}
        total_count = 0
        req_creation_name = ["Microsoft.ExecuteSQLTask", "Microsoft.ExecutePackageTask"]
        if package.root is not None:
            # Start only from the top-level Executables ONCE
            top_executables = package.root.find('DTS:Executables', namespace)
            if top_executables is not None:
                for executable in top_executables.findall('DTS:Executable', namespace):
                    total_count += self.recursive_count(executable,req_creation_name)

        return total_count
    
    def count_sequence_container(self, package):
        """
        Recursively count only Execute SQL and Execute Package tasks (exclude containers and Data Flow tasks).
        """
        namespace = {'DTS': 'www.microsoft.com/SqlServer/Dts'}
        total_count = 0
        req_creation_name = ["STOCK:SEQUENCE"]
        if package.root is not None:
            # Start only from the top-level Executables ONCE
            top_executables = package.root.find('DTS:Executables', namespace)
            if top_executables is not None:
                for executable in top_executables.findall('DTS:Executable', namespace):
                    total_count += self.recursive_count(executable,req_creation_name)

        return total_count

    def count_package_connections(self, package):
        """
        Extracts all the connection metadata from a given SSIS package (.dtsx) and returns a list of connection info.
        This version also extracts deeply nested <connection> tags under <components>.
        """
        connection_managers = []
        namespace = {'DTS': 'www.microsoft.com/SqlServer/Dts'}
        connection_data = []
        
        def recurse_connections(element):
            # Standard DTS connection managers
            for conn_mgrs in element.findall('DTS:ConnectionManagers', namespace):
                for conn in conn_mgrs.findall('DTS:ConnectionManager', namespace):
                    conn_name = conn.attrib.get('{www.microsoft.com/SqlServer/Dts}ObjectName')
                    conn_type = conn.attrib.get('{www.microsoft.com/SqlServer/Dts}CreationName')
                    conn_id = conn.attrib.get('{www.microsoft.com/SqlServer/Dts}DTSID')

                    # Extract property expressions
                    conn_exprs = []
                    for expr_node in conn.findall('DTS:PropertyExpression', namespace):
                        expr_name = expr_node.attrib.get('{www.microsoft.com/SqlServer/Dts}Name', '').strip()
                        expr_value = expr_node.text.strip() if expr_node.text else ''
                        if expr_name and expr_value:
                            conn_exprs.append(f"{expr_name}: {expr_value}")

                    # Find connection string inside nested ObjectData/ConnectionManager
                    conn_string = ""
                    object_data = conn.find('DTS:ObjectData', namespace)
                    if object_data is not None:
                        inner_conn = object_data.find('DTS:ConnectionManager', namespace)
                        if inner_conn is not None:
                            conn_string = inner_conn.attrib.get('{www.microsoft.com/SqlServer/Dts}ConnectionString', '')

                    connection_data.append({
                        "ConnectionName": conn_name,
                        "ConnectionString": conn_string,
                        "ConnectionExpressions": "; ".join(conn_exprs),
                        "ConnectionType": conn_type,
                        "ConnectionID": conn_id,
                        "IsProjectConnection": "0"
                    })

            for child in element:
                recurse_connections(child)
        

        recurse_connections(package.root)
        return connection_data
        #return connection_managers

    def count_package_containers(self, package):
        """
        Counts all the top-level containers in the SSIS package and extracts their metadata.
        """
        containers = []
        expression_details = ""

        for executable in package.executables:
            if isinstance(executable, DtsContainer) and not isinstance(executable, TaskHost):

                if isinstance(executable, ForEachLoop):
                    expression_details = self.get_foreach_loop_expressions(executable)
                else:
                    expression_details = ""

                # Add base container info
                containers.append(ContainerInfo(
                    ContainerName=executable.Name,
                    ContainerType=type(executable).__name__,
                    ContainerExpression=expression_details
                ))

                # Process container-specific inner executables
                if isinstance(executable, Sequence):
                    self.process_sequence_container(executable, containers)
                elif isinstance(executable, ForEachLoop):
                    self.process_foreach_loop_container(executable, containers)
                elif isinstance(executable, ForLoop):
                    self.process_for_loop_container(executable, containers)
        print(containers)
        return containers


    def extract_event_handlers_for_package(self,package):
        if len(package.event_handlers) > 0:
            event_handler_name = package.Name
            event_handler_type = "Package"
            event_name = ""

            for event_handler in package.EventHandlers:
                event_name = event_handler.Name

                for event_executable in event_handler.Executables:
                    if isinstance(event_executable, TaskHost):
                        self.extract_event_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name,
                            "",
                            "",
                            "",
                            ""
                        )

                    elif isinstance(event_executable, Sequence):
                        self.extract_event_sequence_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name
                        )

                    elif isinstance(event_executable, ForEachLoop):
                        self.extract_event_foreach_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name
                        )

                    elif isinstance(event_executable, ForLoop):
                        self.extract_event_for_loop_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name
                        )
                               
    def extract_event_handlers_for_sequence(sequence):
        if len(sequence.event_handlers) > 0:
            event_handler_name = sequence.Name
            event_handler_type = "Sequence"
            event_name = ""

            for event_handler in sequence.EventHandlers:
                event_name = event_handler.Name

                for event_executable in event_handler.Executables:
                    if isinstance(event_executable, TaskHost):
                        self.extract_event_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name,
                            "",
                            "",
                            "",
                            ""
                        )

                    elif isinstance(event_executable, Sequence):
                        self.extract_event_sequence_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name
                        )

                    elif isinstance(event_executable, ForEachLoop):
                        self.extract_event_foreach_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name
                        )

                    elif isinstance(event_executable, ForLoop):
                        self.extract_event_for_loop_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name
                        )

                            
    def extract_event_handlers_for_foreach_loop(foreach_loop):
        if len(foreach_loop.event_handlers) > 0:
            event_handler_name = foreach_loop.Name
            event_handler_type = "ForEachLoop"
            event_name = ""

            for event_handler in foreach_loop.EventHandlers:
                event_name = event_handler.Name

                for event_executable in event_handler.Executables:
                    if isinstance(event_executable, TaskHost):
                        self.extract_event_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name,
                            "",
                            "",
                            "",
                            ""
                        )

                    elif isinstance(event_executable, Sequence):
                        self.extract_event_sequence_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name
                        )

                    elif isinstance(event_executable, ForEachLoop):
                        self.extract_event_foreach_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name
                        )

                    elif isinstance(event_executable, ForLoop):
                        self.extract_event_for_loop_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name
                        )
                        
                                             
    def extract_event_handlers_for_for_loop(for_loop):
        if len(for_loop.event_handlers) > 0:
            event_handler_name = for_loop.Name
            event_handler_type = "ForLoop"
            event_name = ""

            for event_handler in for_loop.EventHandlers:
                event_name = event_handler.Name

                for event_executable in event_handler.Executables:
                    if isinstance(event_executable, TaskHost):
                        self.extract_event_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name,
                            "",
                            "",
                            "",
                            ""
                        )

                    elif isinstance(event_executable, Sequence):
                        self.extract_event_sequence_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name
                        )

                    elif isinstance(event_executable, ForEachLoop):
                        self.extract_event_foreach_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name
                        )

                    elif isinstance(event_executable, ForLoop):
                        self.extract_event_for_loop_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name
                        )


    def extract_event_handlers_for_task(taskhost):
        if len(taskhost.event_handlers) > 0:
            event_handler_name = taskhost.Name

            if isinstance(taskhost.InnerObject, MainPipe):
                event_handler_type = "DataFlowTask"
            elif isinstance(taskhost.InnerObject, ExecutePackageTask):
                event_handler_type = "ExecutePackageTask"
            else:
                event_handler_type = type(taskhost.InnerObject).__name__

            for event_handler in taskhost.EventHandlers:
                event_name = event_handler.Name

                for event_executable in event_handler.Executables:
                    if isinstance(event_executable, TaskHost):
                        self.extract_event_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name,
                            "",
                            "",
                            "",
                            ""
                        )

                    elif isinstance(event_executable, Sequence):
                        self.extract_event_sequence_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name
                        )

                    elif isinstance(event_executable, ForEachLoop):
                        self.extract_event_foreach_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name
                        )

                    elif isinstance(event_executable, ForLoop):
                        self.extract_event_for_loop_task_details(
                            event_executable,
                            event_handler_name,
                            event_handler_type,
                            event_name
                        )


    def extract_precedence_constraints_for_task(self, package_path, package_name):
        metadata = PackageAnalysisResult()
        metadata.PrecedenceConstraintDetails = []

        try:
            tree = ET.parse(package_path)
            root = tree.getroot()

            namespace = {'DTS': 'www.microsoft.com/SqlServer/Dts'}
            precedence_constraints = root.findall(".//DTS:PrecedenceConstraint", namespace)

            for pc in precedence_constraints:
                ref_id = pc.attrib.get("{www.microsoft.com/SqlServer/Dts}refId", "")
                from_full = pc.attrib.get("{www.microsoft.com/SqlServer/Dts}From", "")
                to_full = pc.attrib.get("{www.microsoft.com/SqlServer/Dts}To", "")
                logical_and = pc.attrib.get("{www.microsoft.com/SqlServer/Dts}LogicalAnd", "False")
                expression = pc.attrib.get("{www.microsoft.com/SqlServer/Dts}Expression", "")
                object_name = pc.attrib.get("{www.microsoft.com/SqlServer/Dts}ObjectName", "")

                from_task = from_full.split("\\")[-1] if from_full else ""
                to_task = to_full.split("\\")[-1] if to_full else ""

                # container_name = from_full.split("\\")[1] if len(from_full.split("\\")) >= 3 else ""
                if ".PrecedenceConstraints[" in ref_id:
                    parts = ref_id.split("\\")
                    if len(parts) >= 2:
                        container_segment = parts[-1].split(".PrecedenceConstraints")[0]
                        container_name = container_segment.strip()

                raw_value = pc.attrib.get('{www.microsoft.com/SqlServer/Dts}Value', None)

                if raw_value == "1":
                    precedence_value = "Failure"
                elif raw_value == "2":
                    precedence_value = "Completion"
                else:
                    precedence_value = "Success"
           
                if expression:
                    eval_op = "ExpressionAndConstraint"
                else:
                    if object_name:
                        # Remove trailing digits (and optional spaces) from end
                        eval_op = re.sub(r"\s*\d+$", "", object_name.strip())
                    else:
                        eval_op = ""

                metadata.PrecedenceConstraintDetails.append(PrecedenceConstraintInfo(
                    PackageName=package_name,
                    PackagePath=package_path,
                    PrecedenceConstraintFrom=from_task,
                    PrecedenceConstraintTo=to_task,
                    PrecedenceConstraintValue=precedence_value,
                    PrecedenceConstraintExpression=expression,
                    PrecedenceConstraintEvalOP=eval_op,
                    PrecedenceConstraintLogicalAnd=logical_and,
                    ContainerName=container_name
                ))

            self.save_precedence_constraint_metadata(metadata, self.PackageDetailsFilePath)
            return metadata.PrecedenceConstraintDetails
        
        except Exception as e:
            print(f"Error parsing precedence constraints from {package_path}: {str(e)}")
            return []


    def extract_precedence_constraints_for_sequence(self, package_path, package_name):
        return self.extract_precedence_constraints_for_task(package_path, package_name)


    def extract_precedence_constraints_for_foreach(self, package_path, package_name):
        return self.extract_precedence_constraints_for_task(package_path, package_name)

    def extract_precedence_constraints_for_forloop(self, package_path, package_name):
        return self.extract_precedence_constraints_for_task(package_path, package_name)

                        
    def extract_event_task_details(task_host, event_handler_name, event_handler_type, event_type,
                                container_name, container_type, container_expression, container_enum_details):
        self.extract_task_details(
            task_host,
            event_handler_name,
            event_handler_type,
            event_type,
            "1",  # Event indicator
            container_name,
            container_type,
            container_expression,
            container_enum_details
        )

    def extract_event_sequence_task_details(sequence, event_handler_name, event_handler_type, event_name):
        container_name = sequence.Name
        container_type = "Sequence"

        for event_executable in sequence.Executables:
            if isinstance(event_executable, TaskHost):
                self.extract_event_task_details(
                    event_executable,
                    event_handler_name,
                    event_handler_type,
                    event_name,
                    container_name,
                    container_type,
                    "",
                    ""
                )
            elif isinstance(event_executable, Sequence):
                extract_event_sequence_task_details(
                    event_executable,
                    event_handler_name,
                    event_handler_type,
                    event_name
                )
            elif isinstance(event_executable, ForEachLoop):
                extract_event_foreach_task_details(
                    event_executable,
                    event_handler_name,
                    event_handler_type,
                    event_name
                )
            elif isinstance(event_executable, ForLoop):
                extract_event_for_loop_task_details(
                    event_executable,
                    event_handler_name,
                    event_handler_type,
                    event_name
                )

    def extract_event_foreach_task_details(foreach_loop, event_handler_name, event_handler_type, event_name):
        container_name = foreach_loop.Name
        container_type = "ForEachLoop"
        container_expression = get_for_each_loop_expressions(foreach_loop)
        container_enum = get_for_each_loop_enumerator(foreach_loop)

        for event_executable in foreach_loop.Executables:
            if isinstance(event_executable, TaskHost):
                self.extract_event_task_details(
                    event_executable,
                    event_handler_name,
                    event_handler_type,
                    event_name,
                    container_name,
                    container_type,
                    container_expression,
                    container_enum
                )
            elif isinstance(event_executable, Sequence):
                extract_event_sequence_task_details(
                    event_executable,
                    event_handler_name,
                    event_handler_type,
                    event_name
                )
            elif isinstance(event_executable, ForEachLoop):
                extract_event_foreach_task_details(
                    event_executable,
                    event_handler_name,
                    event_handler_type,
                    event_name
                )
            elif isinstance(event_executable, ForLoop):
                extract_event_for_loop_task_details(
                    event_executable,
                    event_handler_name,
                    event_handler_type,
                    event_name
            )

                           
    def extract_event_for_loop_task_details(forloop, event_handler_name, event_handler_type, event_name):
        container_name = forloop.Name
        container_type = "ForLoop"
        container_expression = get_for_loop_expressions(forloop)
        container_enum = get_for_loop_enumerator(forloop)

        for event_executable in forloop.Executables:
            if isinstance(event_executable, TaskHost):
                self.extract_event_task_details(
                    event_executable,
                    event_handler_name,
                    event_handler_type,
                    event_name,
                    container_name,
                    container_type,
                    container_expression,
                    container_enum
                )

            elif isinstance(event_executable, Sequence):
                extract_event_sequence_task_details(
                    event_executable,
                    event_handler_name,
                    event_handler_type,
                    event_name
                )

            elif isinstance(event_executable, ForEachLoop):
                extract_event_foreach_task_details(
                    event_executable,
                    event_handler_name,
                    event_handler_type,
                    event_name
                )

            elif isinstance(event_executable, ForLoop):
                extract_event_for_loop_task_details(
                    event_executable,
                    event_handler_name,
                    event_handler_type,
                    event_name
                )

                                  
    def extract_variables_for_task(task_host):
        variables_used = []

        task = task_host.InnerObject
        task_type_name = type(task).__name__

        if task_type_name == "ExecuteSQLTask":
            sql_statement = getattr(task, "SqlStatementSource", "")
            connection_string = getattr(task, "Connection", "")

            if sql_statement:
                expression_variables = extract_variables_from_expression(sql_statement)
                variables_used.extend(expression_variables)
            elif connection_string:
                connection_variables = extract_variables_from_expression(connection_string)
                variables_used.extend(connection_variables)

        elif task_type_name == "FileSystemTask":
            is_source_path_variable = getattr(task, "IsSourcePathVariable", False)
            is_destination_path_variable = getattr(task, "IsDestinationPathVariable", False)

            if is_source_path_variable:
                source_path = getattr(task, "Source", "")
                variables_used.append(f"Source Path: {source_path}")
            if is_destination_path_variable:
                destination_path = getattr(task, "Destination", "")
                variables_used.append(f"Destination Path: {destination_path}")

        elif task_type_name == "ScriptTask":
            read_only_variables = getattr(task, "ReadOnlyVariables", "").split(',')
            read_write_variables = getattr(task, "ReadWriteVariables", "").split(',')

            variables_used.extend([var.strip() for var in read_only_variables if var.strip()])
            variables_used.extend([var.strip() for var in read_write_variables if var.strip()])

        return ", ".join(variables_used)

    def extract_variables_from_expression(expression: str) -> list[str]:
        """
        Extracts variable names from an expression string formatted like @[User::VariableName]
        """
        variables = []
        try:
            pattern = re.compile(r'@\[(.*?)\]')
            matches = pattern.findall(expression)
            variables = matches
        except Exception as e:
            print(f"Error extracting variables: {e}")

        return variables
                                             
    def extract_parameters_for_task(self, task_host):
        parameters_used = []

        if isinstance(task_host.InnerObject, ExecuteSQLTask):
            sql_task = task_host.InnerObject

            try:
                parameter_bindings_prop = sql_task.GetType().GetProperty("ParameterBindings", BindingFlags.Public | BindingFlags.NonPublic | BindingFlags.Instance)
                if parameter_bindings_prop is not None:
                    parameter_bindings = parameter_bindings_prop.GetValue(sql_task)

                    for binding in parameter_bindings:
                        name_prop = binding.GetType().GetProperty("ParameterName")
                        direction_prop = binding.GetType().GetProperty("ParameterDirection")
                        data_type_prop = binding.GetType().GetProperty("DataType")
                        value_prop = binding.GetType().GetProperty("Value")
                        dts_variable_prop = binding.GetType().GetProperty("DtsVariableName")

                        parameter_name = name_prop.GetValue(binding) if name_prop else ""
                        parameter_type = direction_prop.GetValue(binding) if direction_prop else ""
                        data_type = data_type_prop.GetValue(binding) if data_type_prop else ""
                        value = value_prop.GetValue(binding) if value_prop else ""
                        dts_variable_name = dts_variable_prop.GetValue(binding) if dts_variable_prop else ""

                        param_str = f"Name: {parameter_name}, Type: {parameter_type}, DataType: {data_type}, Value: {value}, DtsVariableName: {dts_variable_name}"
                        parameters_used.append(param_str)

            except Exception as ex:
                print(f"Error while extracting parameters: {str(ex)}")

        return " | ".join(parameters_used)


    def extract_data_flow_task(self, task_host, event_handle):
        metadata = PackageAnalysisResult()
        metadata.DataFlowTaskDetails = []
        
        if isinstance(task_host.InnerObject, MainPipe):
            data_flow_task = task_host.InnerObject
    
            for component in data_flow_task.ComponentMetaDataCollection:
                cm_check = f"{task_host.Name} :: {component.Name} :: {self.PackageName} : {self.PackagePath}"
               
                if cm_check not in self.ComponentNameCheck:
                    self.ComponentNameCheck.add(cm_check)
                    self.ComponentCount += 1
                    
        return metadata.DataFlowTaskDetails
        '''
                for input_ in component.InputCollection:
                    component_property_details = ""
                    for input_col in input_.InputColumnCollection:
                        column_property_details = ""
                        for prop in input_col.CustomPropertyCollection:
                            column_property_details += f"Property name: {prop.Name}, value: {prop.Value} , Exp: {prop.ExpressionType} "

                        metadata.DataFlowTaskDetails.append(DataFlowTaskInfo(
                            ColumnName=input_col.Name,
                            DataType=str(input_col.DataType),
                            componentName=component.Name,
                            TaskName=task_host.Name,
                            PackageName=self.PackageName,
                            PackagePath=self.PackagePath,
                            ColumnType=str(input_col.ObjectType),
                            isEventHandler=event_handle,
                            ColumnPropertyDetails=column_property_details
                        ))

                    for prop in input_.CustomPropertyCollection:
                        component_property_details += f"Property name: {prop.Name}, value: {prop.Value} , Exp: {prop.ExpressionType}"

                    if component_property_details:
                        metadata.DataFlowTaskDetails.append(DataFlowTaskInfo(
                            componentName=component.Name,
                            TaskName=task_host.Name,
                            PackageName=self.PackageName,
                            PackagePath=self.PackagePath,
                            componentPropertyDetails=f"Companaed Type: {input_.Name}, {component_property_details}",
                            isEventHandler=event_handle
                        ))

                for output in component.OutputCollection:
                    component_property_details = ""
                    for output_col in output.OutputColumnCollection:
                        if any(err in output_col.Name for err in ["Error", "ErrorCode", "ErrorColumn"]):
                            continue

                        column_property_details = ""
                        for prop in output_col.CustomPropertyCollection:
                            column_property_details += f"Property name: {prop.Name}, value: {prop.Value} , Exp: {prop.ExpressionType} "

                        metadata.DataFlowTaskDetails.append(DataFlowTaskInfo(
                            ColumnName=output_col.Name,
                            DataType=str(output_col.DataType),
                            componentName=component.Name,
                            TaskName=task_host.Name,
                            PackageName=self.PackageName,
                            PackagePath=self.PackagePath,
                            ColumnType=str(output_col.ObjectType),
                            isEventHandler=event_handle,
                            ColumnPropertyDetails=column_property_details
                        ))

                    for prop in output.CustomPropertyCollection:
                        component_property_details += f"Property name: {prop.Name}, value: {prop.Value} , Exp: {prop.ExpressionType}"

                    if component_property_details:
                        metadata.DataFlowTaskDetails.append(DataFlowTaskInfo(
                            componentName=component.Name,
                            TaskName=task_host.Name,
                            PackageName=self.PackageName,
                            PackagePath=self.PackagePath,
                            componentPropertyDetails=f"Companaed Type: {output.Name}, {component_property_details}",
                            isEventHandler=event_handle
                        ))

            for component in data_flow_task.ComponentMetaDataCollection:
                if "Data Conversion" in component.Name:
                    for input_col in component.InputCollection[0].InputColumnCollection:
                        metadata.DataFlowTaskDetails.append(DataFlowTaskInfo(
                            ColumnName=input_col.Name,
                            DataType=str(input_col.DataType),
                            componentName=component.Name,
                            TaskName=task_host.Name,
                            DataConversion=str(component.OutputCollection[0].OutputColumnCollection[0].DataType),
                            PackageName=self.PackageName,
                            PackagePath=self.PackagePath,
                            ColumnType="Data Conversion :" + str(input_col.ObjectType),
                            isEventHandler=event_handle
                        ))

                    for output_col in component.OutputCollection[0].OutputColumnCollection:
                        if any(err in output_col.Name for err in ["Error", "ErrorCode", "ErrorColumn"]):
                            continue
                        metadata.DataFlowTaskDetails.append(DataFlowTaskInfo(
                            ColumnName=output_col.Name,
                            DataType=str(output_col.DataType),
                            componentName=component.Name,
                            TaskName=task_host.Name,
                            DataConversion=str(component.OutputCollection[0].OutputColumnCollection[0].DataType),
                            PackageName=self.PackageName,
                            PackagePath=self.PackagePath,
                            ColumnType="Data Conversion :" + str(output_col.ObjectType),
                            isEventHandler=event_handle
                        ))
        print(metadata)
        self.save_dataflow_metadata(metadata, self.DataFlowlFilePath)'''
        

                                 
    def match_columns(input_column, output_column):
        """
        Checks if input and output SSIS columns match by name or data type.

        Args:
            input_column: An object representing the SSIS input column.
            output_column: An object representing the SSIS output column.

        Returns:
            bool: True if the columns match by name or data type, else False.
        """

        # Match by name
        if input_column.Name == output_column.Name:
            return True

        # Match by data type
        if input_column.DataType == output_column.DataType:
            return True

        # No match
        return False

                                             
    def extract_expressions_for_task(self, task_host):
        expressions_used = []
        expression_details = ""

        try:
            task = task_host.InnerObject
            task_type = type(task)

            if task_host.HasExpressions:
                for prop in task_host.Properties:
                    try:
                        expression = task_host.GetExpression(prop.Name)
                        if expression:
                            expressions_used.append(f"Property: {prop.Name}, Expression: {expression}")
                    except Exception as ex:
                        print(f"Error extracting expression for property {prop.Name}: {str(ex)}")

            if isinstance(task, MainPipe):  # Data Flow Task
                for component in task.ComponentMetaDataCollection:
                    for custom_property in component.CustomPropertyCollection:
                        if custom_property.Name == "Expression":
                            expressions_used.append(
                                f"Expression Name: {custom_property.Name} Expression Value: {custom_property.Value}"
                            )
            else:
                for prop_info in task_type.GetProperties():
                    try:
                        expression = task_host.GetExpression(prop_info.Name)
                        if expression:
                            expressions_used.append(f"Property: {prop_info.Name}, Expression: {expression}")
                    except Exception as ex:
                        print(f"Error reading expression from property {prop_info.Name}: {str(ex)}")

            return ", ".join(expressions_used)

        except Exception as ex:
            print(f"Error while extracting expressions: {str(ex)}")
            return ""


    
    
    def get_foreach_loop_expressions(self, foreach_loop):
        """
        Extracts expression metadata from a ForEachLoop container's enumerator.
        """
        expression_list = []

        try:
            enumerator = foreach_loop.ForEachEnumerator

            if hasattr(enumerator, "Properties"):
                for prop in enumerator.Properties:
                    expression = ""
                    try:
                        # Attempt to retrieve expression using the property name
                        expression = enumerator.GetExpression(prop.Name)
                    except Exception as ex:
                        print(f"Error retrieving expression for {prop.Name}: {str(ex)}")

                    if expression:
                        expression_list.append(f"Property: {prop.Name}, Expression: {expression}")
        except Exception as ex:
            print(f"Error accessing enumerator: {str(ex)}")

        return " | ".join(expression_list)


    def get_foreach_loop_enumerator(self, foreach_loop):
        """
        Extracts enumerator metadata from a ForEachLoop container, excluding default system properties.
        """
        enumerator_details = []

        try:
            enumerator = foreach_loop.ForEachEnumerator

            if enumerator:
                if isinstance(enumerator, ForEachEnumeratorHost):
                    excluded_properties = {
                        "ID",
                        "Description",
                        "CollectionEnumerator",
                        "CreationName",
                        "Name"
                    }

                    for prop in enumerator.Properties:
                        enum_name = prop.Name
                        if enum_name not in excluded_properties:
                            try:
                                value = prop.GetValue(enumerator)
                                enumerator_details.append(f"EnumeratorName: {enum_name}, EnumeratorValue: {value}")
                            except Exception as inner_ex:
                                print(f"Could not get value for property {enum_name}: {str(inner_ex)}")

        except Exception as ex:
            print(f"An error occurred: {str(ex)}")

        return " | ".join(enumerator_details)

    
    def get_for_loop_expressions(self, for_loop):
        """
        Extracts expression properties from a ForLoop container.
        """
        expressions = []

        for prop in for_loop.Properties:
            expression = ""
            try:
                expression = for_loop.GetExpression(prop.Name)
            except Exception as ex:
                print(f"Error retrieving expression for {prop.Name}: {str(ex)}")

            if expression:
                expressions.append(f"Property: {prop.Name}, Expression: {expression}")

        return " | ".join(expressions)
        

    def get_for_loop_enumerator(self, for_loop):
        """
        Returns a string summary of the Assign, Init, and Eval expressions of a ForLoop container.
        """
        try:
            return (f"AssignExpression : {for_loop.AssignExpression} | "
                    f"InitExpression : {for_loop.InitExpression} | "
                    f"EvalExpression : {for_loop.EvalExpression}")
        except Exception as ex:
            print(f"An error occurred: {str(ex)}")
            return ""
            

    def process_sequence_container_details(self, container, containers, package):
        seq_tasks = []
        container_name = container.Name
        container_type = type(container).__name__

        if len(container.EventHandlers) > 0:
            self.extract_event_handlers_for_sequence(container)

        for executable in container.Executables:
            if isinstance(executable, TaskHost):
                self.extract_task_details(
                    executable, "", "", "", "0", container_name, container_type, "", ""
                )

                seq_tasks.append(TaskInfo(
                    SeqTaskName=executable.Name
                ))

            elif isinstance(executable, DtsContainer):
                containers.append(ContainerInfo(
                    ContainerName=executable.Name,
                    ContainerType=type(executable).__name__
                ))

                if isinstance(executable, Sequence):
                    # Optionally extract actual nested task info using: self.process_sequence_container_details(...)
                    seq_tasks.extend(self.count_sequence_container_tasks(package))
                    self.process_sequence_container(executable, containers)

                elif isinstance(executable, ForEachLoop):
                    seq_tasks.extend(self.count_foreach_container_tasks(package))
                    self.process_foreach_loop_container(executable, containers)

                elif isinstance(executable, ForLoop):
                    seq_tasks.extend(self.count_forloop_container_tasks(package))
                    self.process_for_loop_container(executable, containers)

        self.containerTaskCount += len(seq_tasks)
        return seq_tasks

    
    def process_foreach_loop_container_details(self, container, containers, package):
        foreach_tasks = []
        expression_details = self.get_foreach_loop_expressions(container)
        container_name = container.Name
        container_type = type(container).__name__
        enumerator_details = self.get_foreach_loop_enumerator(container)

        if len(container.EventHandlers) > 0:
            self.extract_event_handlers_for_foreach_loop(container)

        for executable in container.Executables:
            if isinstance(executable, TaskHost):
                self.extract_task_details(
                    executable, "", "", "", "0", container_name, container_type, expression_details, enumerator_details
                )

                foreach_tasks.append(TaskInfo(
                    ForeachTaskName=executable.Name
                ))

            elif isinstance(executable, DtsContainer):
                containers.append(ContainerInfo(
                    ContainerName=executable.Name,
                    ContainerType=type(executable).__name__
                ))

                if isinstance(executable, Sequence):
                    foreach_tasks.extend(self.count_sequence_container_tasks(package))
                    self.process_sequence_container(executable, containers)

                elif isinstance(executable, ForEachLoop):
                    foreach_tasks.extend(self.count_foreach_container_tasks(package))
                    self.process_foreach_loop_container(executable, containers)

                elif isinstance(executable, ForLoop):
                    foreach_tasks.extend(self.count_forloop_container_tasks(package))
                    self.process_for_loop_container(executable, containers)

        self.containerTaskCount += len(foreach_tasks)

        return foreach_tasks


    def process_for_loop_container_details(self, container, containers, package):
        for_loop_tasks = []
        expression_details = self.get_for_loop_expressions(container)
        enumerator_details = self.get_for_loop_enumerator(container)

        container_name = container.Name
        container_type = type(container).__name__

        if len(container.EventHandlers) > 0:
            self.extract_event_handlers_for_for_loop(container)

        for executable in container.Executables:
            if isinstance(executable, TaskHost):
                self.extract_task_details(
                    executable, "", "", "", "0", container_name, container_type, expression_details, enumerator_details
                )

                for_loop_tasks.append(TaskInfo(
                    ForloopTaskName=executable.Name
                ))

            elif isinstance(executable, DtsContainer):
                containers.append(ContainerInfo(
                    ContainerName=executable.Name,
                    ContainerType=type(executable).__name__
                ))

                if isinstance(executable, Sequence):
                    for_loop_tasks.extend(self.count_sequence_container_tasks(package))
                    self.process_sequence_container(executable, containers)

                elif isinstance(executable, ForEachLoop):
                    for_loop_tasks.extend(self.count_foreach_container_tasks(package))
                    self.process_foreach_loop_container(executable, containers)

                elif isinstance(executable, ForLoop):
                    for_loop_tasks.extend(self.count_forloop_container_tasks(package))
                    self.process_for_loop_container(executable, containers)

        self.containerTaskCount += len(for_loop_tasks)

        return for_loop_tasks

    def process_sequence_container(self, container, containers):
        """
        Recursively processes a Sequence container and counts the number of TaskHost elements within it.
        """
        task_count = 0

        for executable in container.Executables:
            if isinstance(executable, TaskHost):
                task_count += 1
            elif isinstance(executable, DtsContainer):
                self.container_count += 1

                if isinstance(executable, Sequence):
                    task_count += self.process_sequence_container(executable, containers)

                elif isinstance(executable, ForEachLoop):
                    self.process_foreach_loop_container(executable, containers)

                elif isinstance(executable, ForLoop):
                    self.process_for_loop_container(executable, containers)

        return task_count


    def process_foreach_loop_container(self, container, containers):
        """
        Recursively processes a ForEachLoop container and counts the number of TaskHost elements within it.
        """
        task_count = 0

        for executable in container.Executables:
            if isinstance(executable, TaskHost):
                task_count += 1
            elif isinstance(executable, DtsContainer):
                self.container_count += 1

                if isinstance(executable, ForEachLoop):
                    task_count += self.process_foreach_loop_container(executable, containers)

                elif isinstance(executable, Sequence):
                    self.process_sequence_container(executable, containers)

                elif isinstance(executable, ForLoop):
                    self.process_for_loop_container(executable, containers)

        return task_count


    def process_for_loop_container(self, container, containers):
        """
        Recursively processes a ForLoop container and counts the number of TaskHost elements within it.
        """
        task_count = 0

        for executable in container.Executables:
            if isinstance(executable, TaskHost):
                task_count += 1
            elif isinstance(executable, DtsContainer):
                self.container_count += 1

                if isinstance(executable, ForLoop):
                    task_count += self.process_for_loop_container(executable, containers)

                elif isinstance(executable, Sequence):
                    self.process_sequence_container(executable, containers)

                elif isinstance(executable, ForEachLoop):
                    self.process_foreach_loop_container(executable, containers)

        return task_count

    def extract_task_details(
        self, task_host, event_handler_name, event_handler_type, event_type,
        event_indicator, container_name, container_type, container_expression, enum_details
    ):
        metadata = PackageAnalysisResult()
        metadata.ExtractTaskDetails = []

        sql_query = ""
        execute_process_details = ""
        source_path = destination_path = ""
        source_component_name = target_component_name = ""
        source_type = target_type = ""
        sql_table = target_sql_table = ""
        send_mail_task_details = ftp_task_details = script_task_details = ""
        execute_package_task_details = task_component_details = ""
        xml_task = bulk_insert_task = expression_task = ""
        result_set = ""
        connection_id = source_connection_id = target_connection_id = ""

        task_type = type(task_host.InnerObject)
        task_type_name = task_type.__name__

        if len(task_host.EventHandlers) > 0:
            self.extract_event_handlers_for_task(task_host)

        if isinstance(task_host.InnerObject, MainPipe):
            self.extract_data_flow_task(task_host, event_indicator)
            task_type_name = "DataFlowTask"

            for component in task_host.InnerObject.ComponentMetaDataCollection:
                if "Source" in component.Description:
                    source_component_name = component.Name
                    source_type = component.Description
                    source_connection_id = component.RuntimeConnectionCollection[0].ConnectionManagerID
                    for prop in component.CustomPropertyCollection:
                        if prop.Name == "SqlCommand":
                            sql_query = prop.Value
                        elif prop.Name == "OpenRowset":
                            sql_table = prop.Value
                    if not sql_query:
                        sql_query = sql_table

                elif "Destination" in component.Description:
                    target_component_name = component.Name
                    target_type = component.Description
                    target_connection_id = component.RuntimeConnectionCollection[0].ConnectionManagerID
                    for prop in component.CustomPropertyCollection:
                        if prop.Name == "OpenRowset":
                            target_sql_table = prop.Value

        elif task_type.__name__ == "ExecuteSQLTask":
            connection_id = getattr(task_host.InnerObject, "Connection", "")
            sql_query = getattr(task_host.InnerObject, "SqlStatementSource", "")
            sql_task = task_host.InnerObject
            if hasattr(sql_task, "ResultSetBindings"):
                for binding in sql_task.ResultSetBindings:
                    result_set += f"Result Set Column: {binding.ResultName} | SSIS Variable: {binding.DtsVariableName}  "

        elif isinstance(task_host.InnerObject, FileSystemTask):
            source_path = task_host.InnerObject.Source
            destination_path = task_host.InnerObject.Destination
            task_component_details = f"SourcePath: {source_path} | DestinationPath: {destination_path}"

        elif isinstance(task_host.InnerObject, ExecuteProcess):
            proc = task_host.InnerObject
            execute_process_details = f"Executable: {proc.Executable} | Arguments: {proc.Arguments} | WorkingDirectory: {proc.WorkingDirectory}"
            task_component_details = execute_process_details

        elif isinstance(task_host.InnerObject, SendMailTask):
            mail = task_host.InnerObject
            connection_id = mail.SmtpConnection
            send_mail_task_details = (
                f"From: {mail.FromLine} | To: {mail.ToLine} | CC: {mail.CCLine} "
                f"BCC: {mail.BCCLine} | Subject: {mail.Subject} | Body: {mail.MessageSource} | "
                f"FileAttachments: {mail.FileAttachments} | Priority: {mail.Priority}"
            )
            task_component_details = send_mail_task_details

        elif isinstance(task_host.InnerObject, FtpTask):
            ftp = task_host.InnerObject
            connection_id = ftp.Connection
            ftp_task_details = (
                f"FTP Operation: {ftp.Operation} | LocalPath: {ftp.LocalPath} | RemotePath: {ftp.RemotePath} | "
                f"OverwriteDestination: {ftp.OverwriteDestination} | IsLocalPathVariable: {ftp.IsLocalPathVariable} | "
                f"IsRemotePathVariable: {ftp.IsRemotePathVariable} | IsTransferTypeASCII: {ftp.IsTransferTypeASCII} | "
                f"StopOnOperationFailure: {ftp.StopOnOperationFailure}"
            )
            task_component_details = ftp_task_details

        elif isinstance(task_host.InnerObject, ScriptTask):
            script = task_host.InnerObject
            script_task_details = (
                f"Script Language: {script.ScriptLanguage} | EntryPoint: {script.EntryPoint} | "
                f"ReadOnlyVariables: {script.ReadOnlyVariables} | ReadWriteVariables: {script.ReadWriteVariables} | "
                f"ScriptProjectName: {script.ScriptProjectName}"
            )
            for prop in task_host.Properties:
                if prop.Name == "CodePage":
                    script_task_details += f" | Code Page: {prop.GetValue(script)}"
            task_component_details = script_task_details

        elif isinstance(task_host.InnerObject, ExecutePackageTask):
            execute_package_task_details = task_host.InnerObject.PackageName
            task_type_name = "ExecutePackageTask"
            task_component_details = execute_package_task_details

        elif isinstance(task_host.InnerObject, XMLTask):
            xml = task_host.InnerObject
            xml_task = (
                f"Source: {xml.Source} | SourceType: {xml.SourceType} | DiffAlgorithm: {xml.DiffAlgorithm} | "
                f"DiffGramDestination: {xml.DiffGramDestination} | DiffOptions: {xml.DiffOptions} | "
                f"DiffGramDestinationType: {xml.DiffGramDestinationType} | FailOnDifference: {xml.FailOnDifference} | "
                f"SaveDiffGram: {xml.SaveDiffGram} | OperationType: {xml.OperationType} | "
                f"SaveOperationResult: {xml.SaveOperationResult} | SecondOperand: {xml.SecondOperand} | "
                f"SecondOperandType: {xml.SecondOperandType}"
            )
            task_component_details = xml_task

        elif isinstance(task_host.InnerObject, BulkInsertTask):
            bulk = task_host.InnerObject
            source_connection_id = bulk.SourceConnection
            target_connection_id = bulk.DestinationConnection
            bulk_insert_task = (
                f"FormatFile: {bulk.FormatFile} | FieldTerminator: {bulk.FieldTerminator} | RowTerminator: {bulk.RowTerminator} | "
                f"DestinationTableName: {bulk.DestinationTableName} | CodePage: {bulk.CodePage} | DataFileType: {bulk.DataFileType} | "
                f"BatchSize: {bulk.BatchSize} | LastRow: {bulk.LastRow} | FirstRow: {bulk.FirstRow} | "
                f"CheckConstraints: {bulk.CheckConstraints} | KeepNulls: {bulk.KeepNulls} | KeepIdentity: {bulk.KeepIdentity} | "
                f"TableLock: {bulk.TableLock} | FireTriggers: {bulk.FireTriggers} | SortedData: {bulk.SortedData} | "
                f"MaximumErrors: {bulk.MaximumErrors}"
            )
            task_component_details = bulk_insert_task

        elif isinstance(task_host.InnerObject, ExpressionTask):
            expr = task_host.InnerObject
            expression_task = f"Expression: {expr.Expression} | ExecutionValue: {expr.ExecutionValue}"
            task_component_details = expression_task

        # Add the task info to metadata
        metadata.ExtractTaskDetails.append(TaskInfo(
            TaskName=task_host.Name,
            TaskType=task_type_name,
            TaskSqlQuery=sql_query,
            Variables=self.extract_variables_for_task(task_host),
            FileSystemSourcePath=source_path,
            FileSystemDestinationPath=destination_path,
            Parameters=self.extract_parameters_for_task(task_host),
            Expressions=self.extract_expressions_for_task(task_host),
            ExecuteProcessDetails=execute_process_details,
            SourceComponent=source_component_name,
            TargetComponent=target_component_name,
            SourceType=source_type,
            TargetType=target_type,
            TargetTable=target_sql_table,
            SendMailTask=send_mail_task_details,
            ScriptTask=script_task_details,
            FTPTask=ftp_task_details,
            ExecutePackage=execute_package_task_details,
            ResultSetDetails=result_set,
            EventHandlerName=event_handler_name,
            EventHandlerType=event_handler_type,
            EventType=event_type,
            ContainerName=container_name,
            ContainerType=container_type,
            ContainerExpression=container_expression,
            PackageName=self.PackageName,
            PackagePath=self.PackagePath,
            ContainerEnum=enum_details,
            SourceConnectionName=source_connection_id,
            TargetConnectionName=target_connection_id,
            ConnectionName=connection_id,
            TaskComponentDetails=task_component_details
        ))

        # Save based on event or normal
        if event_indicator == "1":
            self.save_event_metadata(metadata, self.PackageDetailsFilePath)
        else:
            self.save_package_task_metadata(metadata, self.PackageDetailsFilePath)

            # Save container details separately
            container_metadata = [{
                "PackageName": self.PackageName,
                "PackagePath": self.PackagePath,
                "ContainerName": container_name,
                "ContainerType": container_type,
                "ContainerExpression": container_expression,
                "ContainerEnum": enum_details
            }]
            self.save_package_container_metadata(container_metadata, self.PackageDetailsFilePath)


        return metadata.ExtractTaskDetails


    def measure_package_performance(self, package_obj):
        """
        Executes a mock package and returns the execution time as a timedelta object.
        Assumes 'package_obj' has an 'execute()' method.
        """
        start_time = time.time()
        package_obj.execute()  # This should be your actual package processing logic
        end_time = time.time()
        return end_time - start_time  # Returns seconds as float

    @staticmethod
    def does_workbook_exist(file_path):
        """
        Checks if an Excel workbook exists at the specified path.
        Returns True if exists, otherwise False.
        """
        return os.path.isfile(file_path)

    def save_package_metadata(self, result, analysis_file_path, details_file_path):
        #print(result.Connections) -- used for testing purpose
        task_count = (
            result.Tasks +             
            len(result.Foreachtasks) +
            len(result.Seqtasks) +
            len(result.Forlooptasks)
        )

        complexity_count = (
            task_count +
            result.Containers +
            self.container_count +
            self.ComponentCount
        )

        if complexity_count <= 5:
            complexity = "Simple"
        elif 5 < complexity_count <= 10:
            complexity = "Medium"
        elif complexity_count > 10:
            complexity = "Complex"
        else:
            complexity = "Simple"

        if self.DataSaveType.upper() == "EXCEL":
            workbook_exists = os.path.exists(analysis_file_path)
            if workbook_exists:
                workbook = load_workbook(analysis_file_path)
            else:
                workbook = Workbook()

            sheet_name = "PackageAnalysisResults"
            if sheet_name not in workbook.sheetnames:
                sheet = workbook.create_sheet(sheet_name)
                sheet.append([
                    "PackageName", "PackagePath", "TasksCount", "ConnectionsCount", "ContainerCount",
                    "ComponentCount", "ExecutionTime", "CreatedDate", "CreatedBy", "Complexity"
                ])
            else:
                sheet = workbook[sheet_name]

            row = [
                result.PackageName,
                result.PackagePath,
                task_count,
                len(result.Connections),
                result.Containers + self.container_count,
                self.ComponentCount,
                result.ExecutionTime,
                result.CreatedDate,
                result.CreatedBy,
                complexity
            ]
            sheet.append(row)
            workbook.save(analysis_file_path)
            format_excel_file(analysis_file_path)

            # Saving Variables to 'PackageVariableParameterDetails'
            workbook_details = load_workbook(details_file_path) if os.path.exists(details_file_path) else Workbook()

            # 1. Variables
            var_sheet_name = "PackageVariableParameterDetails"
            if var_sheet_name not in workbook_details.sheetnames:
                sheet_vars = workbook_details.create_sheet(var_sheet_name)
                sheet_vars.append([
                    "PackageName", "PackagePath", "VariableOrParameterName", "DataType", "Value", "IsParameter"
                ])
            else:
                sheet_vars = workbook_details[var_sheet_name]

            for var in result.Variables:
                sheet_vars.append([
                    result.PackageName, result.PackagePath, var.Name, var.DataType, var.Value, var.IsParameter
                ])

            # 2. Connections
            conn_sheet_name = "PackageConnectionDetails"
            if conn_sheet_name not in workbook_details.sheetnames:
                sheet_conn = workbook_details.create_sheet(conn_sheet_name)
                sheet_conn.append([
                    "PackageName", "PackagePath", "ConnectionName", "ConnectionType",
                    "ConnectionExpressions", "ConnectionString", "ConnectionID", "IsProjectConnection"
                ])
            else:
                sheet_conn = workbook_details[conn_sheet_name]

            for conn in result.Connections:
                sheet_conn.append([
                    result.PackageName, result.PackagePath,
                    str(conn["ConnectionName"]), str(conn["ConnectionType"]),
                    str(conn["ConnectionExpressions"]), str(conn["ConnectionString"]),
                    str(conn["ConnectionID"]), str(conn["IsProjectConnection"])
                ])

            workbook_details.save(details_file_path)
            format_excel_file(details_file_path)


        elif self.DataSaveType.upper() == "SQL":
            conn = pyodbc.connect(self._connection_string)
            cursor = conn.cursor()

            insert_package = """
            INSERT INTO PackageAnalysisResults 
            (PackageName, CreatedDate, TaskCount, ConnectionCount, ExecutionTime, PackageFolder, ContainerCount, DTSXXML, CreatedBy, DataFlowTaskComponentCount)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """
            cursor.execute(insert_package, (
                result.PackageName,
                result.CreatedDate,
                task_count,
                len(result.Connections),
                result.ExecutionTime,
                result.PackagePath,
                result.Containers + self.container_count,
                result.DTSXXML,
                result.CreatedBy,
                self.ComponentCount
            ))

            for var in result.Variables:
                insert_var = """
                INSERT INTO PackageVariableParameterDetails 
                (PackageName, VariableOrParameterName, DataType, Value, PackagePath, IsParameter)
                VALUES (?, ?, ?, ?, ?, ?)
                """
                cursor.execute(insert_var, (
                    result.PackageName, var.Name, var.DataType, var.Value, result.PackagePath, var.IsParameter
                ))

            for conn in result.Connections:
                insert_conn = """
                INSERT INTO PackageConnectionDetails 
                (PackageName, ConnectionName, ConnectionType, PackagePath, ConnectionExpressions, ConnectionString, ConnectionDTSID, IsProjectConnection)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """
                cursor.execute(insert_conn, (
                    result.PackageName,
                    conn.ConnectionName,
                    conn.ConnectionType,
                    result.PackagePath,
                    conn.ConnectionExpressions,
                    conn.ConnectionString,
                    conn.ConnectionID,
                    conn.IsProjectConnection
                ))

            conn.commit()
            cursor.close()
            conn.close()

    def save_dataflow_metadata(self, result, file_path):
        if self.DataSaveType.upper() == "EXCEL":
            for dataflow in result.DataFlowTaskDetails:
                dataflow_file = os.path.join(
                    file_path, dataflow.PackageName.replace(".dtsx", "_DFM.xlsx")
                )
                workbook_exists = os.path.exists(dataflow_file)

                if workbook_exists:
                    workbook = openpyxl.load_workbook(dataflow_file)
                else:
                    workbook = openpyxl.Workbook()
                    workbook.remove(workbook.active)

                sheet_name = "DataFlowTaskMappingDetails"
                if sheet_name not in workbook.sheetnames:
                    ws = workbook.create_sheet(sheet_name)
                    ws.append([
                        "PackageName", "PackagePath", "TaskName", "ColumnName", "ColumnType", "DataType",
                        "ComponentName", "DataConversion", "ComponentPropertyDetails",
                        "ColumnPropertyDetails", "isEventHandler"
                    ])
                else:
                    ws = workbook[sheet_name]

                # Check for existing record
                exists = False
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if (
                        row[0] == dataflow.PackageName and
                        row[1] == dataflow.PackagePath and
                        row[2] == dataflow.TaskName and
                        row[3] == dataflow.ColumnName and
                        row[4] == dataflow.ColumnType and
                        row[5] == dataflow.DataType and
                        row[6] == dataflow.componentName and
                        row[7] == dataflow.DataConversion and
                        row[8] == dataflow.componentPropertyDetails and
                        row[9] == dataflow.ColumnPropertyDetails and
                        row[10] == dataflow.isEventHandler
                    ):
                        exists = True
                        break

                if not exists:
                    ws.append([
                        dataflow.PackageName,
                        dataflow.PackagePath,
                        dataflow.TaskName,
                        dataflow.ColumnName,
                        dataflow.ColumnType,
                        dataflow.DataType,
                        dataflow.componentName,
                        dataflow.DataConversion,
                        dataflow.componentPropertyDetails,
                        dataflow.ColumnPropertyDetails,
                        dataflow.isEventHandler
                    ])
                    workbook.save(dataflow_file)
                    format_excel_file(dataflow_file)


        elif self.DataSaveType.upper() == "SQL":
            conn = pyodbc.connect(self._connection_string)
            cursor = conn.cursor()

            for dataflow in result.DataFlowTaskDetails:
                insert_query = """
                    INSERT INTO DataFlowTaskMappingDetails (
                        PackageName, TaskName, ColumnName, DataType, ComponentName, DataConversion, PackagePath, 
                        ColumnType, isEventHandler, ComponentPropertyDetails, ColumnPropertyDetails
                    )
                    SELECT ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
                    WHERE NOT EXISTS (
                        SELECT 1 FROM DataFlowTaskMappingDetails
                        WHERE ISNULL(ColumnName, '') = ISNULL(?, '') AND ISNULL(DataType, '') = ISNULL(?, '')
                        AND ISNULL(PackageName, '') = ISNULL(?, '') AND ISNULL(PackagePath, '') = ISNULL(?, '')
                        AND ISNULL(ColumnType, '') = ISNULL(?, '') AND ISNULL(ComponentName, '') = ISNULL(?, '')
                        AND ISNULL(TaskName, '') = ISNULL(?, '') AND ISNULL(ComponentPropertyDetails, '') = ISNULL(?, '')
                        AND ISNULL(ColumnPropertyDetails, '') = ISNULL(?, '')
                    )
                """
                values = [
                    dataflow.PackageName,
                    dataflow.TaskName,
                    dataflow.ColumnName,
                    dataflow.DataType,
                    dataflow.componentName,
                    dataflow.DataConversion,
                    dataflow.PackagePath,
                    dataflow.ColumnType,
                    dataflow.isEventHandler,
                    dataflow.componentPropertyDetails,
                    dataflow.ColumnPropertyDetails,
                    # WHERE NOT EXISTS values
                    dataflow.ColumnName,
                    dataflow.DataType,
                    dataflow.PackageName,
                    dataflow.PackagePath,
                    dataflow.ColumnType,
                    dataflow.componentName,
                    dataflow.TaskName,
                    dataflow.componentPropertyDetails,
                    dataflow.ColumnPropertyDetails
                ]
                cursor.execute(insert_query, values)

            conn.commit()
            cursor.close()
            conn.close()

    def save_precedence_constraint_metadata(self, result, file_path):
        if self.DataSaveType.upper() == "EXCEL":
            workbook_exists = os.path.exists(file_path)
            wb = load_workbook(file_path) if workbook_exists else Workbook()

            if not wb.sheetnames:
                wb.remove(wb.active)

            sheet_name = "PrecedenceConstraintDetails"

            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
                ws.append([
                    "PackageName", "PackagePath", "PrecedenceConstraintFrom", "PrecedenceConstraintTo",
                    "PrecedenceConstraintValue", "PrecedenceConstraintExpression",
                    "PrecedenceConstraintLogicalAnd", "PrecedenceConstraintEvalOP", "ContainerName"
                ])

            for precedence in result.PrecedenceConstraintDetails:
                ws.append([
                    precedence.PackageName, 
                    precedence.PackagePath,
                    precedence.PrecedenceConstraintFrom, 
                    precedence.PrecedenceConstraintTo,
                    precedence.PrecedenceConstraintValue, 
                    precedence.PrecedenceConstraintExpression,
                    precedence.PrecedenceConstraintLogicalAnd,
                    precedence.PrecedenceConstraintEvalOP,
                    precedence.ContainerName
                ])
                
            wb.save(file_path)
            format_excel_file(file_path)
            print(f"Saved PrecedenceConstraintDetails to Excel: {file_path}")

        elif self.DataSaveType.upper() == "SQL":
            conn = pyodbc.connect(self._connection_string)
            cursor = conn.cursor()
            for precedence in result.PrecedenceConstraintDetails:
                insert_query = """
                    INSERT INTO PrecedenceConstraintDetails (
                        PackageName, PrecedenceConstraintFrom, PrecedenceConstraintTo, 
                        PrecedenceConstraintValue, PrecedenceConstraintExpression, PrecedenceConstraintLogicalAnd,
                        PrecedenceConstraintEvalOP, ContainerName, PackagePath
                    )
                    SELECT ?, ?, ?, ?, ?, ?, ?, ?, ?
                    WHERE NOT EXISTS (
                        SELECT 1 FROM PrecedenceConstraintDetails
                        WHERE PackageName = ? AND PrecedenceConstraintFrom = ? AND PrecedenceConstraintTo = ?
                        AND PrecedenceConstraintValue = ? AND PrecedenceConstraintExpression = ?
                        AND PrecedenceConstraintLogicalAnd = ? AND PrecedenceConstraintEvalOP = ?
                        AND ContainerName = ? AND PackagePath = ?
                    )
                """
                values = [
                    precedence.PackageName, precedence.PrecedenceConstraintFrom, precedence.PrecedenceConstraintTo,
                    precedence.PrecedenceConstraintValue, precedence.PrecedenceConstraintExpression,
                    precedence.PrecedenceConstraintLogicalAnd, precedence.PrecedenceConstraintEvalOP,
                    precedence.ContainerName, precedence.PackagePath,
                    # For WHERE NOT EXISTS part
                    precedence.PackageName, precedence.PrecedenceConstraintFrom, precedence.PrecedenceConstraintTo,
                    precedence.PrecedenceConstraintValue, precedence.PrecedenceConstraintExpression,
                    precedence.PrecedenceConstraintLogicalAnd, precedence.PrecedenceConstraintEvalOP,
                    precedence.ContainerName, precedence.PackagePath
                ]
                cursor.execute(insert_query, values)
            conn.commit()
            cursor.close()
            conn.close()

    
    def save_event_metadata(self, result, file_path):
        if self.DataSaveType.upper() == "EXCEL":
            workbook_exists = os.path.exists(file_path)
            wb = load_workbook(file_path) if workbook_exists else Workbook()
            if not wb.sheetnames:
                wb.remove(wb.active)

            sheet_name = "EventHandlerTaskDetails"
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
                ws.append([
                    "PackageName", "PackagePath", "EventHandlerName", "EventHandlerType", "EventType",
                    "TaskName", "TaskType", "ContainerName", "ContainerType", "ContainerExpression",
                    "TaskConnectionName", "SqlQuery", "Variables", "Parameters", "Expressions",
                    "DataFlowDaskSourceName", "DataFlowTaskSourceType", "DataFlowTaskTargetName", "DataFlowTaskTargetType",
                    "DataFlowTaskTargetTable", "DataFlowDaskSourceConnectionName", "DataFlowDaskTargetConnectionName",
                    "SendMailTaskDetails", "ResultSetDetails", "TaskComponentDetails"
                ])

            for task in result:
                ws.append([
                    task["PackageName"], task["PackagePath"], task["EventHandlerName"], task["EventHandlerType"], task["EventType"],
                    task["TaskName"], task["TaskType"], task["ContainerName"], task["ContainerType"], task["ContainerExpression"],
                    task["TaskConnectionName"], task["SqlQuery"], task["Variables"], task["Parameters"], task["Expressions"],
                    task["DataFlowDaskSourceName"], task["DataFlowTaskSourceType"], task["DataFlowTaskTargetName"], task["DataFlowTaskTargetType"],
                    task["DataFlowTaskTargetTable"], task["DataFlowDaskSourceConnectionName"], task["DataFlowDaskTargetConnectionName"],
                    task["SendMailTaskDetails"], task["ResultSetDetails"], task["TaskComponentDetails"]
                ])

            wb.save(file_path)
            format_excel_file(file_path)
            print(f"Saved EventHandler details to Excel: {file_path}")

        elif self.DataSaveType.upper() == "SQL":
            conn = pyodbc.connect(self._connection_string)
            cursor = conn.cursor()
            for task in result:
                insert_query = """
                    INSERT INTO EventTaskDetails (
                        PackageName, PackagePath, EventHandlerName, EventHandlerType, EventType,
                        TaskName, TaskType, ContainerName, ContainerType, ContainerExpression,
                        TaskConnectionName, SqlQuery, Variables, Parameters, Expressions,
                        DataFlowDaskSourceName, DataFlowTaskSourceType, DataFlowTaskTargetName, DataFlowTaskTargetType,
                        DataFlowTaskTargetTable, DataFlowDaskSourceConnectionName, DataFlowDaskTargetConnectionName,
                        SendMailTaskDetails, ResultSetDetails, TaskComponentDetails
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """
                cursor.execute(insert_query, (
                    task["PackageName"], task["PackagePath"], task["EventHandlerName"], task["EventHandlerType"], task["EventType"],
                    task["TaskName"], task["TaskType"], task["ContainerName"], task["ContainerType"], task["ContainerExpression"],
                    task["TaskConnectionName"], task["SqlQuery"], task["Variables"], task["Parameters"], task["Expressions"],
                    task["DataFlowDaskSourceName"], task["DataFlowTaskSourceType"], task["DataFlowTaskTargetName"], task["DataFlowTaskTargetType"],
                    task["DataFlowTaskTargetTable"], task["DataFlowDaskSourceConnectionName"], task["DataFlowDaskTargetConnectionName"],
                    task["SendMailTaskDetails"], task["ResultSetDetails"], task["TaskComponentDetails"]
                ))
            conn.commit()
            cursor.close()
            conn.close()
            print("Saved EventHandler details to SQL Server.")


    def save_package_task_metadata(self, result, file_path):
        if self.DataSaveType.upper() == "EXCEL":
            tasks = result.get("ExtractTaskDetails", [])
            workbook_exists = os.path.exists(file_path)

            for task in tasks:
                if task.get("TaskName"):
                    wb = load_workbook(file_path) if workbook_exists else Workbook()
                    if not wb.sheetnames:
                        wb.remove(wb.active)

                    sheet_name = "PackageTaskDetails"
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                    else:
                        ws = wb.create_sheet(sheet_name)
                        ws.append([
                            "PackageName", "PackagePath", "TaskName", "TaskType", "ContainerName",
                            "TaskConnectionName", "SqlQuery", "Variables", "Parameters", "Expressions",
                            "DataFlowTaskSourceName", "DataFlowTaskSourceType", "DataFlowTaskTargetName",
                            "DataFlowTaskTargetType", "DataFlowTaskTargetTable", "DataFlowTaskSourceConnectionName",
                            "DataFlowTaskTargetConnectionName", "ResultSetDetails", "TaskComponentDetails"
                        ])

                    rows = list(ws.iter_rows(min_row=2, values_only=True))
                    record_exists = any(
                        row[0] == task.get("PackageName") and
                        row[1] == task.get("PackagePath") and
                        row[2] == task.get("TaskName") and
                        row[3] == task.get("TaskType") and
                        row[4] == task.get("ContainerName")
                        for row in rows
                    )

                    if not record_exists:
                        ws.append([
                            task.get("PackageName"), 
                            task.get("PackagePath"), 
                            task.get("TaskName"),
                            task.get("TaskType"), 
                            task.get("ContainerName"), 
                            task.get("TaskConnectionName"),
                            task.get("TaskSqlQuery"), 
                            task.get("Variables"), 
                            task.get("Parameters"),
                            task.get("Expressions"), 
                            task.get("SourceComponent"), 
                            task.get("SourceType"),
                            task.get("TargetComponent"), 
                            task.get("TargetType"), 
                            task.get("TargetTable"),
                            task.get("SourceConnectionName"),
                            task.get("TargetConnectionName"),
                            task.get("ResultSetDetails"),
                            task.get("TaskComponentDetails")
                        ])
                        wb.save(file_path)
                        format_excel_file(file_path)


        elif self.DataSaveType.upper() == "SQL":
            conn = pyodbc.connect(self._connection_string)
            cursor = conn.cursor()

            for task in result.get("ExtractTaskDetails", []):
                if task.get("TaskName"):
                    task_query = """
                        INSERT INTO PackageTaskDetails (
                            PackageName, TaskName, TaskType, SqlQuery, ContainerName, PackagePath,
                            Variables, Parameters, Expressions, DataFlowDaskSourceName, DataFlowTaskSourceType,
                            DataFlowTaskTargetName, DataFlowTaskTargetType, DataFlowTaskTargetTable,
                            ResultSetDetails, DataFlowDaskSourceConnectionName,
                            DataFlowDaskTargetConnectionName, TaskConnectionName, TaskComponentDetails
                        )
                        SELECT ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
                        WHERE NOT EXISTS (
                            SELECT 1 FROM PackageTaskDetails
                            WHERE ContainerName = ? AND PackageName = ? AND PackagePath = ? AND TaskName = ?
                        )
                    """
                    values = [
                        task.get("PackageName"),
                        task.get("TaskName"),
                        task.get("TaskType"),
                        task.get("TaskSqlQuery"),
                        task.get("ContainerName"),
                        task.get("PackagePath"),
                        task.get("Variables"), 
                        task.get("Parameters"), 
                        task.get("Expressions"),
                        task.get("SourceComponent"), 
                        task.get("SourceType"), 
                        task.get("TargetComponent"),
                        task.get("TargetType"), 
                        task.get("TargetTable"), 
                        task.get("ResultSetDetails"),
                        task.get("SourceConnectionName"), 
                        task.get("TargetConnectionName"),
                        task.get("ConnectionName"), 
                        task.get("TaskComponentDetails"),
                        task.get("ContainerName"), 
                        task.get("PackageName"), 
                        task.get("PackagePath"), 
                        task.get("TaskName")
                    ]
                    cursor.execute(task_query, values)

            conn.commit()
            cursor.close()
            conn.close()
            print("Saved Package Task details to SQL Server.")

    def save_package_container_metadata(self, containers, file_path):
        if self.DataSaveType.upper() == "EXCEL":
            workbook_exists = os.path.exists(file_path)

            for container in containers:
                wb = load_workbook(file_path) if workbook_exists else Workbook()
                if not wb.sheetnames:
                    wb.remove(wb.active)

                sheet_name = "PackageContainerDetails"
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    ws = wb.create_sheet(sheet_name)
                    ws.append([
                        "PackageName", "PackagePath", "ContainerName",
                        "ContainerType", "ContainerExpressions", "ContainerEnumerator"
                    ])

                rows = list(ws.iter_rows(min_row=2, values_only=True))
                record_exists = any(
                    row[0] == container.get("PackageName") and
                    row[1] == container.get("PackagePath") and
                    row[2] == container.get("ContainerName") and
                    row[3] == container.get("ContainerType") and
                    row[4] == container.get("ContainerExpression") and
                    row[5] == container.get("ContainerEnum")
                    for row in rows
                )

                if not record_exists:
                    ws.append([
                        container.get("PackageName"), 
                        container.get("PackagePath"), 
                        container.get("ContainerName"),
                        container.get("ContainerType"), 
                        container.get("ContainerExpression"), 
                        container.get("ContainerEnum")
                    ])
                    wb.save(file_path)
                    format_excel_file(file_path)

        elif self.DataSaveType.upper() == "SQL":
            conn = pyodbc.connect(self._connection_string)
            cursor = conn.cursor()

            for container in containers:
                container_query = """
                    INSERT INTO PackageContainerDetails (
                        PackageName, ContainerName, ContainerType,
                        ContainerExpressions, ContainerEnumerator, PackagePath
                    )
                    SELECT ?, ?, ?, ?, ?, ?
                    WHERE NOT EXISTS (
                        SELECT 1 FROM PackageContainerDetails
                        WHERE ContainerName = ? AND ContainerType = ? AND PackageName = ?
                        AND PackagePath = ? AND ContainerExpressions = ISNULL(?, '')
                        AND ContainerEnumerator = ISNULL(?, '')
                    )
                """
                values = [
                    container.get("PackageName"), 
                    container.get("ContainerName"), 
                    container.get("ContainerType"),
                    container.get("ContainerExpression"), 
                    container.get("ContainerEnum"), 
                    container.get("PackagePath"),
                    container.get("ContainerName"), 
                    container.get("ContainerType"), 
                    container.get("PackageName"),
                    container.get("PackagePath"), 
                    container.get("ContainerExpression"), 
                    container.get("ContainerEnum")
                ]
                cursor.execute(container_query, values)

            conn.commit()
            cursor.close()
            conn.close()
            print("Saved Package Container metadata to SQL Server.")       


    def save_project_parameter_metadata(self, result, file_path):
        if self.DataSaveType.upper() == "EXCEL":
            # Open or create workbook
            wb = load_workbook(file_path) if os.path.exists(file_path) else Workbook()
            if not wb.sheetnames:
                wb.remove(wb.active)

            # Get or create sheet
            sheet_name = "ProjectParameterDetails"
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
                ws.append(["ProjectPath", "ParameterName", "ParameterValue", "ParameterDataType"])

            # Collect existing rows for duplicate check
            existing_rows = list(ws.iter_rows(min_row=2, values_only=True))

            for param_info in result.get("ProjectParameterDetails", []):
                row_data = (
                    result.get("ProjectPath", ""),
                    param_info.get("ParameterName", ""),
                    param_info.get("Value", ""),
                    param_info.get("DataType", "")
                )

                if row_data not in existing_rows:
                    ws.append(row_data)

            wb.save(file_path)
            format_excel_file(file_path)
            print(f"Saved project parameters to Excel: {file_path}")

        elif self.DataSaveType.upper() == "SQL":
            conn = pyodbc.connect(self._connection_string)
            cursor = conn.cursor()

            for param_info in result.get("ProjectParameterDetails", []):
                cursor.execute("""
                    IF NOT EXISTS (
                        SELECT 1 FROM ProjectParameterDetails
                        WHERE ParameterName = ? AND ProjectPath = ?
                    )
                    INSERT INTO ProjectParameterDetails (
                        ParameterName, ParameterValue, ParameterDataType, ProjectPath
                    ) VALUES (?, ?, ?, ?)
                """, (
                    param_info.get("ParameterName", ""),
                    result.get("ProjectPath", ""),
                    param_info.get("ParameterName", ""),
                    param_info.get("Value", ""),
                    param_info.get("DataType", ""),
                    result.get("ProjectPath", "")
                ))

            conn.commit()
            cursor.close()
            conn.close()
            print("Saved project parameters to SQL Server.")

    def save_connections_metadata(self, result, file_path):
        if self.DataSaveType.upper() == "EXCEL":
            for conn in result.get("Connections", []):
                if os.path.exists(file_path):
                    wb = load_workbook(file_path)
                else:
                    wb = Workbook()
                    wb.remove(wb.active)

                if "PackageConnectionDetails" in wb.sheetnames:
                    ws = wb["PackageConnectionDetails"]
                else:
                    ws = wb.create_sheet("PackageConnectionDetails")
                    ws.append([
                        "PackageName", "PackagePath", "ConnectionName", "ConnectionType",
                        "ConnectionExpressions", "ConnectionString", "ConnectionID", "IsProjectConnection"
                    ])

                row = [
                    result.get("PackageName", ""),
                    result.get("PackagePath", ""),
                    conn.get("ConnectionName", ""),
                    conn.get("ConnectionType", ""),
                    conn.get("ConnectionExpressions", ""),
                    conn.get("ConnectionString", ""),
                    conn.get("ConnectionID", ""),
                    conn.get("IsProjectConnection", "")
                ]
                ws.append(row)
                wb.save(file_path)
                format_excel_file(file_path)

            print(f"Saved connection metadata to Excel: {file_path}")

        elif self.DataSaveType.upper() == "SQL":
            conn_db = pyodbc.connect(self._connection_string)
            cursor = conn_db.cursor()

            for conn_info in result.get("Connections", []):
                cursor.execute("""
                    INSERT INTO PackageConnectionDetails (
                        PackageName, ConnectionName, ConnectionType, PackagePath, 
                        ConnectionExpressions, ConnectionString, ConnectionDTSID, IsProjectConnection
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    result.get("PackageName", ""),
                    conn_info.get("ConnectionName", ""),
                    conn_info.get("ConnectionType", ""),
                    result.get("PackagePath", ""),
                    conn_info.get("ConnectionExpressions", ""),
                    conn_info.get("ConnectionString", ""),
                    conn_info.get("ConnectionID", ""),
                    conn_info.get("IsProjectConnection", "")
                ))

            conn_db.commit()
            cursor.close()
            conn_db.close()
            print("Saved connection metadata to SQL Server.")

    def log_error(self, file_path, exception):
        print(f"Error processing {file_path}: {str(exception)}")
        traceback.print_exc()

    def save_update_connection_name(self, file_path):
        if self.DataSaveType.upper() == "EXCEL":
            # Check and create workbook if it doesn't exist
            if not self.does_workbook_exist(file_path):
                wb = Workbook()
                # Rename default sheet and create required ones
                wb.create_sheet("PackageTaskDetails")
                wb.create_sheet("PackageConnectionDetails")
                wb.create_sheet("EventHandlerTaskDetails")
                if "Sheet" in wb.sheetnames:
                    del wb["Sheet"]
                wb.save(file_path)
                format_excel_file(file_path)
                print(f"Created new Excel workbook at: {file_path}")

            wb = load_workbook(file_path)

            # Ensure required sheets exist
            for sheet_name in ["PackageTaskDetails", "PackageConnectionDetails", "EventHandlerTaskDetails"]:
                if sheet_name not in wb.sheetnames:
                    wb.create_sheet(sheet_name)
            
            # ✅ Ensure required sheets exist
            sheet1 = wb["PackageTaskDetails"]
            sheet2 = wb["PackageConnectionDetails"]
            sheet3 = wb["EventHandlerTaskDetails"]

            for row2 in range(2, sheet2.max_row + 1):
                conn_pkg_path = sheet2.cell(row=row2, column=2).value
                conn_name = sheet2.cell(row=row2, column=3).value
                conn_dtsid = sheet2.cell(row=row2, column=7).value

                for row1 in range(2, sheet1.max_row + 1):
                    task_pkg_path = sheet1.cell(row=row1, column=2).value
                    if conn_pkg_path == task_pkg_path:
                        if conn_dtsid == sheet1.cell(row=row1, column=6).value:
                            sheet1.cell(row=row1, column=6).value = conn_name
                        elif conn_dtsid == sheet1.cell(row=row1, column=16).value:
                            sheet1.cell(row=row1, column=16).value = conn_name
                        elif conn_dtsid == sheet1.cell(row=row1, column=17).value:
                            sheet1.cell(row=row1, column=17).value = conn_name

                for row3 in range(2, sheet3.max_row + 1):
                    handler_pkg_path = sheet3.cell(row=row3, column=2).value
                    if conn_pkg_path == handler_pkg_path:
                        if conn_dtsid == sheet3.cell(row=row3, column=11).value:
                            sheet3.cell(row=row3, column=11).value = conn_name
                        elif conn_dtsid == sheet3.cell(row=row3, column=21).value:
                            sheet3.cell(row=row3, column=21).value = conn_name
                        elif conn_dtsid == sheet3.cell(row=row3, column=22).value:
                            sheet3.cell(row=row3, column=22).value = conn_name

            wb.save(file_path)
            format_excel_file(file_path)
            print("Connection names updated in Excel.")

        elif self.DataSaveType.upper() == "SQL":
            conn = pyodbc.connect(self._connection_string)
            cursor = conn.cursor()

            connection_query = """
            UPDATE task SET task.TaskConnectionName = conn.ConnectionName
            FROM PackageTaskDetails task
            INNER JOIN PackageConnectionDetails conn WITH (NOLOCK)
                ON conn.PackagePath = task.PackagePath
                AND task.TaskConnectionName = conn.ConnectionDTSID
            WHERE ISNULL(task.TaskConnectionName, '') <> '';

            UPDATE task SET 
                task.DataFlowDaskSourceConnectionName = sconn.ConnectionName, 
                task.DataFlowDaskTargetConnectionName = tconn.ConnectionName
            FROM PackageTaskDetails task
            INNER JOIN PackageConnectionDetails sconn WITH (NOLOCK)
                ON sconn.PackagePath = task.PackagePath
                AND task.DataFlowDaskSourceConnectionName = sconn.ConnectionDTSID
            INNER JOIN PackageConnectionDetails tconn WITH (NOLOCK)
                ON tconn.PackagePath = task.PackagePath
                AND task.DataFlowDaskTargetConnectionName = tconn.ConnectionDTSID
            WHERE ISNULL(task.DataFlowDaskSourceConnectionName, '') <> '';

            UPDATE task SET task.TaskConnectionName = conn.ConnectionName
            FROM EventTaskDetails task
            INNER JOIN PackageConnectionDetails conn WITH (NOLOCK)
                ON conn.PackagePath = task.PackagePath
                AND task.TaskConnectionName = conn.ConnectionDTSID
            WHERE ISNULL(task.TaskConnectionName, '') <> '';

            UPDATE task SET 
                task.DataFlowDaskSourceConnectionName = sconn.ConnectionName,
                task.DataFlowDaskTargetConnectionName = tconn.ConnectionName
            FROM EventTaskDetails task
            INNER JOIN PackageConnectionDetails sconn WITH (NOLOCK)
                ON sconn.PackagePath = task.PackagePath
                AND task.DataFlowDaskSourceConnectionName = sconn.ConnectionDTSID
            INNER JOIN PackageConnectionDetails tconn WITH (NOLOCK)
                ON tconn.PackagePath = task.PackagePath
                AND task.DataFlowDaskTargetConnectionName = tconn.ConnectionDTSID
            WHERE ISNULL(task.DataFlowDaskSourceConnectionName, '') <> '';

            -- Reset precedence constraints
            UPDATE task SET
                task.ONSuccessPrecedenceConstrainttoTask = '',
                task.ONSuccessPrecedenceConstraintExpression = '',
                task.ONSuccessPrecedenceConstraintEvalOP = '',
                task.ONSuccessPrecedenceConstraintLogicalAnd = '',
                task.ONFailurePrecedenceConstrainttoTask = '',
                task.ONFailurePrecedenceConstraintExpression = '',
                task.ONFailurePrecedenceConstraintEvalOP = '',
                task.ONFailurePrecedenceConstraintLogicalAnd = '',
                task.ONCompletionPrecedenceConstrainttoTask = '',
                task.ONCompletionPrecedenceConstraintExpression = '',
                task.ONCompletionPrecedenceConstraintEvalOP = '',
                task.ONCompletionPrecedenceConstraintLogicalAnd = ''
            FROM PackageTaskDetails task;

            -- Apply precedence constraints (Success, Failure, Completion)
            UPDATE task SET
                task.ONSuccessPrecedenceConstrainttoTask = pcd.PrecedenceConstraintto,
                task.ONSuccessPrecedenceConstraintExpression = pcd.PrecedenceConstraintExpression,
                task.ONSuccessPrecedenceConstraintEvalOP = pcd.PrecedenceConstraintEvalOP,
                task.ONSuccessPrecedenceConstraintLogicalAnd = pcd.PrecedenceConstraintLogicalAnd
            FROM PackageTaskDetails task
            INNER JOIN PrecedenceConstraintDetails pcd WITH (NOLOCK)
                ON pcd.PrecedenceConstraintFrom = task.TaskName
                AND pcd.PackageName = task.PackageName
                AND pcd.PackagePath = task.PackagePath
                AND ISNULL(pcd.ContainerName, '') = ISNULL(task.ContainerName, '')
            WHERE pcd.PrecedenceConstraintValue = 'Success';

            UPDATE task SET
                task.ONFailurePrecedenceConstrainttoTask = pcd.PrecedenceConstraintto,
                task.ONFailurePrecedenceConstraintExpression = pcd.PrecedenceConstraintExpression,
                task.ONFailurePrecedenceConstraintEvalOP = pcd.PrecedenceConstraintEvalOP,
                task.ONFailurePrecedenceConstraintLogicalAnd = pcd.PrecedenceConstraintLogicalAnd
            FROM PackageTaskDetails task
            INNER JOIN PrecedenceConstraintDetails pcd WITH (NOLOCK)
                ON pcd.PrecedenceConstraintFrom = task.TaskName
                AND pcd.PackageName = task.PackageName
                AND pcd.PackagePath = task.PackagePath
                AND ISNULL(pcd.ContainerName, '') = ISNULL(task.ContainerName, '')
            WHERE pcd.PrecedenceConstraintValue = 'Failure';

            UPDATE task SET
                task.ONCompletionPrecedenceConstrainttoTask = pcd.PrecedenceConstraintto,
                task.ONCompletionPrecedenceConstraintExpression = pcd.PrecedenceConstraintExpression,
                task.ONCompletionPrecedenceConstraintEvalOP = pcd.PrecedenceConstraintEvalOP,
                task.ONCompletionPrecedenceConstraintLogicalAnd = pcd.PrecedenceConstraintLogicalAnd
            FROM PackageTaskDetails task
            INNER JOIN PrecedenceConstraintDetails pcd WITH (NOLOCK)
                ON pcd.PrecedenceConstraintFrom = task.TaskName
                AND pcd.PackageName = task.PackageName
                AND pcd.PackagePath = task.PackagePath
                AND ISNULL(pcd.ContainerName, '') = ISNULL(task.ContainerName, '')
            WHERE pcd.PrecedenceConstraintValue = 'Completion';

            -- Set Complexity Classification
            UPDATE PA SET PA.Complexcity = CASE 
                WHEN Final.TaskCount + Final.ContainerCount + Final.ComponentCount < 5 THEN 'Simple'
                WHEN Final.TaskCount + Final.ContainerCount + Final.ComponentCount BETWEEN 5 AND 10 THEN 'Medium'
                WHEN Final.TaskCount + Final.ContainerCount + Final.ComponentCount > 10 THEN 'Complex'
                ELSE 'Simple'
            END
            FROM PackageAnalysisResults PA
            LEFT JOIN (
                SELECT PackageName, PackagePath,
                       SUM(TaskCount) TaskCount,
                       SUM(ContainerCount) ContainerCount,
                       SUM(ComponentCount) ComponentCount
                FROM (
                    SELECT PT.PackageName, PT.PackagePath,
                           SUM(CASE WHEN TaskType <> 'ExecutePackageTask' THEN 1 ELSE 0 END) AS TaskCount,
                           0 AS ContainerCount,
                           0 AS ComponentCount
                    FROM PackageTaskDetails PT
                    GROUP BY PT.PackageName, PT.PackagePath

                    UNION ALL

                    SELECT PT.PackageName, PT.PackagePath,
                           1 AS TaskCount, 0, 0
                    FROM PackageTaskDetails PT
                    WHERE TaskType = 'ExecutePackageTask'

                    UNION ALL

                    SELECT PC.PackageName, PC.PackagePath,
                           0, 1, 0
                    FROM PackageContainerDetails PC
                    WHERE PC.ContainerType = 'Sequence'

                    UNION ALL

                    SELECT PC.PackageName, PC.PackagePath,
                           0, COUNT(1), 0
                    FROM PackageContainerDetails PC
                    WHERE PC.ContainerType <> 'Sequence'
                    GROUP BY PackageName, PackagePath

                    UNION ALL

                    SELECT PC.PackageName, PC.PackagePath,
                           0, 0, COUNT(DISTINCT ComponentName)
                    FROM DataFlowTaskMappingDetails PC
                    GROUP BY PackageName, PackagePath
                ) A
                GROUP BY PackageName, PackagePath
            ) Final ON Final.PackageName = PA.PackageName
            AND Final.PackagePath = PA.PackageFolder;
            """

            cursor.execute(connection_query)
            conn.commit()
            cursor.close()
            conn.close()
            print("Connection names updated in SQL.")

class Program:
    @staticmethod
    def delete_all_files_in_directory(directory_path):
        try:
            if os.path.exists(directory_path) and os.path.isdir(directory_path):
                for file_name in os.listdir(directory_path):
                    file_path = os.path.join(directory_path, file_name)
                    if os.path.isfile(file_path):
                        os.remove(file_path)
            else:
                print("Directory does not exist.")
        except Exception as ex:
            print(f"An error occurred: {ex}")

    @staticmethod
    def main():
        connection_string = ""
        output_folder = ""

        package_folder = input("Enter the Package Folder path:\n")
        data_save_type = input("Enter the Data Save Type (SQL or EXCEL):\n").strip().upper()

        if data_save_type == "SQL":
            connection_string = input("Enter the Connection String:\n")
        elif data_save_type == "EXCEL":
            output_folder = input("Enter the Output Folder path:\n")
        else:
            print("Wrong Input")
            time.sleep(5)
            return

        package_analysis_file_path = os.path.join(output_folder, "PackageAnalysisResult.xlsx")
        dataflow_file_path = output_folder
        package_details_file_path = os.path.join(output_folder, "PackageDetails.xlsx")

        if data_save_type == "EXCEL":
            Program.delete_all_files_in_directory(dataflow_file_path)

        analyzer = SSISPackageAnalyzer(
            package_folder,
            connection_string,
            package_analysis_file_path,
            dataflow_file_path,
            package_details_file_path,
            data_save_type
        )
        analyzer.analyze_all_packages(package_folder)

        print("Running...")


if __name__ == "__main__":
    Program.main()
